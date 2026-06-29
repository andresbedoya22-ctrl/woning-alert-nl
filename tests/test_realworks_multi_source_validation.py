from __future__ import annotations

import ast
from datetime import UTC, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR / "scraper" / "src"))

from domek_wonen.compliance import robots_gate  # noqa: E402
from domek_wonen.pilots.realworks_multi_source_validation import (  # noqa: E402
    RealworksValidationSource,
    decide_realworks_family,
    run_realworks_multi_source_validation,
    select_realworks_validation_sources,
    validation_source_from_evidence_row,
)


MODULE_PATH = BASE_DIR / "scraper" / "src" / "domek_wonen" / "pilots" / "realworks_multi_source_validation.py"
OBSERVED = datetime(2026, 6, 29, 8, 0, tzinfo=UTC)


def _row(
    *,
    source_id: str = "oldenkotte.com__tilburg",
    source_name: str = "Oldenkotte",
    domain: str = "oldenkotte.com",
    listing_url: str = "https://www.oldenkotte.com/aanbod/woningaanbod/koop/",
    gemeente: str = "Tilburg",
    delivery_mode: str = "realworks_public",
    parser_family_candidate: str = "realworks_public",
    access_status: str = "allowed",
    platform_guess: str = "realworks",
    needs_review: str = "false",
    notes: str = "local Realworks evidence",
) -> dict[str, str]:
    return {
        "source_id": source_id,
        "source_name": source_name,
        "domain": domain,
        "listing_url": listing_url,
        "gemeente": gemeente,
        "delivery_mode": delivery_mode,
        "parser_family_candidate": parser_family_candidate,
        "access_status": access_status,
        "platform_guess": platform_guess,
        "needs_review": needs_review,
        "notes": notes,
    }


def _source(source_id: str = "oldenkotte.com__tilburg", domain: str = "oldenkotte.com") -> RealworksValidationSource:
    return RealworksValidationSource(
        source_id=source_id,
        source_name=source_id,
        domain=domain,
        listing_url=f"https://{domain}/aanbod/woningaanbod/koop",
        gemeente="Tilburg",
        platform_guess="realworks",
        delivery_mode="realworks_public",
        parser_family_candidate="realworks_public",
        access_policy_status="allowed",
        selection_reason="test",
    )


def _listing_html(domain: str, *, href_suffix: str = "huis-1001-teststraat-1") -> str:
    return f"""
    <html><body>
      <li class="aanbodEntry">
        <a href="/aanbod/woningaanbod/breda/koop/{href_suffix}">
          <span class="street-address">Teststraat 1</span>
          <span class="locality">Breda</span>
          <span class="price">EUR 425.000 k.k.</span>
          <span class="objectstatusbanner">Beschikbaar</span>
          <span>Woonhuis 123 m2 5 kamers Te koop {domain}</span>
        </a>
      </li>
    </body></html>
    """


def _detail_html(
    *,
    postcode: str = "4811 AA",
    energy: str = "A",
    property_type: str = "Woonhuis",
    source_published: str = "2026-06-29",
) -> str:
    return f"""
    <html>
      <head>
        <meta name="description" content="Compacte omschrijving.">
        <script type="application/ld+json">{{"datePublished":"{source_published}"}}</script>
      </head>
      <body>
        <div>Teststraat 1 | {postcode} Breda</div>
        <span class="kenmerk"><span class="kenmerkName">Soort object</span><span class="kenmerkValue">{property_type}</span></span>
        <span class="kenmerk"><span class="kenmerkName">Vraagprijs</span><span class="kenmerkValue">EUR 425.000 k.k.</span></span>
        <span class="kenmerk"><span class="kenmerkName">Status</span><span class="kenmerkValue">Beschikbaar</span></span>
        <span class="kenmerk"><span class="kenmerkName">Woonoppervlakte</span><span class="kenmerkValue">123 m2</span></span>
        <span class="kenmerk"><span class="kenmerkName">Perceeloppervlakte</span><span class="kenmerkValue">234 m2</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal kamers</span><span class="kenmerkValue">5 kamers</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal slaapkamers</span><span class="kenmerkValue">3 slaapkamers</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal badkamers</span><span class="kenmerkValue">1 badkamer</span></span>
        <span class="kenmerk"><span class="kenmerkName">Inhoud</span><span class="kenmerkValue">456 m3</span></span>
        <span class="kenmerk"><span class="kenmerkName">Energieklasse</span><span class="kenmerkValue">{energy}</span></span>
        <span class="kenmerk"><span class="kenmerkName">Bouwjaar</span><span class="kenmerkValue">1998</span></span>
        <span class="kenmerk"><span class="kenmerkName">Verwarming</span><span class="kenmerkValue">CV-ketel</span></span>
        <span class="kenmerk"><span class="kenmerkName">Tuin</span><span class="kenmerkValue">Achtertuin</span></span>
        <span class="kenmerk"><span class="kenmerkName">Parkeerfaciliteiten</span><span class="kenmerkValue">Openbaar parkeren</span></span>
        <span class="kenmerk"><span class="kenmerkName">Eigendomssituatie</span><span class="kenmerkValue">Volle eigendom</span></span>
      </body>
    </html>
    """


def _fetch(url: str) -> str:
    domain = url.split("//", 1)[1].split("/", 1)[0].removeprefix("www.")
    if url.rstrip("/").endswith("/koop") or url.rstrip("/").endswith("/woningaanbod"):
        return _listing_html(domain)
    return _detail_html()


def test_selects_oldenkotte_control_from_explicit_local_input() -> None:
    result = select_realworks_validation_sources([_row()])

    assert result[0].source_id == "oldenkotte.com__tilburg"
    assert result[0].selection_reason


def test_selects_second_realworks_source_when_valid_evidence_exists() -> None:
    result = select_realworks_validation_sources(
        [
            _row(),
            _row(
                source_id="olden.nl__heusden",
                source_name="Olden",
                domain="olden.nl",
                listing_url="https://www.olden.nl/aanbod/woningaanbod",
                gemeente="Heusden",
            ),
        ]
    )

    assert [source.source_id for source in result] == ["oldenkotte.com__tilburg", "olden.nl__heusden"]


def test_excludes_funda_and_pararius() -> None:
    rows = [
        _row(source_id="funda", domain="funda.nl", listing_url="https://www.funda.nl/koop/tilburg/"),
        _row(source_id="pararius", domain="pararius.nl", listing_url="https://www.pararius.nl/koopwoningen"),
    ]

    assert select_realworks_validation_sources(rows) == ()


def test_excludes_blocked_legal_review_and_permission_required() -> None:
    rows = [
        _row(source_id="blocked", domain="blocked.nl", access_status="blocked"),
        _row(source_id="legal", domain="legal.nl", access_status="legal_review"),
        _row(source_id="permission", domain="permission.nl", access_status="permission_required"),
    ]

    assert select_realworks_validation_sources(rows) == ()


def test_does_not_invent_listing_url() -> None:
    assert validation_source_from_evidence_row(_row(listing_url="")) is None


def test_caps_max_sources() -> None:
    rows = [
        _row(),
        _row(source_id="olden.nl__heusden", domain="olden.nl", listing_url="https://olden.nl/aanbod/woningaanbod"),
        _row(source_id="third.nl__breda", domain="third.nl", listing_url="https://third.nl/aanbod/woningaanbod"),
    ]

    assert len(select_realworks_validation_sources(rows, max_sources=2)) == 2


def test_caps_max_detail_fetches_per_source(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)
    calls: list[str] = []

    def fetch(url: str) -> str:
        calls.append(url)
        if url.rstrip("/").endswith("/koop"):
            return (
                _listing_html("example.nl", href_suffix="huis-1001-teststraat-1")
                + _listing_html("example.nl", href_suffix="huis-1002-teststraat-2")
            )
        return _detail_html()

    result = run_realworks_multi_source_validation(
        sources=(_source(domain="example.nl"),),
        observed_at=OBSERVED,
        fetch_html=fetch,
        max_detail_fetches_per_source=1,
    )

    assert result.metrics[0].detail_attempted == 1
    assert len(calls) == 2


def test_calls_robots_gate_before_listing_fetch(monkeypatch) -> None:
    events: list[str] = []

    def can_fetch(domain: str, path: str) -> bool:
        events.append(f"robots:{path}")
        return True

    def fetch(url: str) -> str:
        events.append(f"fetch:{url}")
        return _fetch(url)

    monkeypatch.setattr(robots_gate, "can_fetch", can_fetch)
    run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=fetch)

    assert events[0].startswith("robots:")
    assert events[1].startswith("fetch:")


def test_calls_robots_gate_before_detail_fetch(monkeypatch) -> None:
    events: list[str] = []

    def can_fetch(domain: str, path: str) -> bool:
        events.append(f"robots:{path}")
        return True

    def fetch(url: str) -> str:
        events.append(f"fetch:{url}")
        return _fetch(url)

    monkeypatch.setattr(robots_gate, "can_fetch", can_fetch)
    run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=fetch)

    detail_fetch_index = next(index for index, event in enumerate(events) if "huis-1001-teststraat-1" in event and event.startswith("fetch:"))
    assert events[detail_fetch_index - 1].startswith("robots:")


def test_aggregates_per_source_parser_metrics(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=_fetch)

    assert result.metrics[0].parser_total == 1
    assert result.metrics[0].parser_qa_clean == 1


def test_aggregates_per_source_readiness_metrics(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=_fetch)

    assert result.metrics[0].facts_records_built == 1
    assert result.metrics[0].readiness_rows_built == 1
    assert result.metrics[0].export_ready == 1


def test_aggregates_postcode_vve_energy_non_residential_and_status_metrics(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    def fetch(url: str) -> str:
        if url.rstrip("/").endswith("/koop"):
            return _listing_html("example.nl")
        return _detail_html(energy="Niet aanwezig", property_type="Garage")

    result = run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=fetch)
    metric = result.metrics[0]

    assert metric.postcode_usable == 1
    assert metric.energy_label_review_missing == 1
    assert metric.non_residential_rows == 1
    assert metric.active_inventory_eligible_false == 1
    assert ("store_excluded_non_residential", 1) in metric.db_persistence_action_counts


def test_aggregates_lifecycle_and_freshness_metrics(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(sources=(_source(domain="example.nl"),), observed_at=OBSERVED, fetch_html=_fetch)
    metric = result.metrics[0]

    assert metric.source_published_at_usable == 1
    assert ("new_listing", 1) in metric.lifecycle_event_counts
    assert ("new_today", 1) in metric.freshness_bucket_counts


def test_family_decision_usable_when_two_sources_pass(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(
        sources=(
            _source("oldenkotte.com__tilburg", "oldenkotte.com"),
            _source("olden.nl__heusden", "olden.nl"),
        ),
        observed_at=OBSERVED,
        fetch_html=_fetch,
    )

    assert result.family_decision.family_decision == "realworks_family_usable_for_broader_audit"
    assert result.family_decision.second_makelaar_validated is True


def test_family_decision_needs_hardening_when_second_source_has_parser_failure(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    def fetch(url: str) -> str:
        if "olden.nl" in url and (url.rstrip("/").endswith("/koop") or url.rstrip("/").endswith("/woningaanbod")):
            return "<html></html>"
        return _fetch(url)

    result = run_realworks_multi_source_validation(
        sources=(
            _source("oldenkotte.com__tilburg", "oldenkotte.com"),
            _source("olden.nl__heusden", "olden.nl"),
        ),
        observed_at=OBSERVED,
        fetch_html=fetch,
    )

    assert result.family_decision.family_decision == "realworks_family_needs_hardening_v2"


def test_family_decision_insufficient_when_only_one_source_is_valid(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(
        sources=(_source("oldenkotte.com__tilburg", "oldenkotte.com"),),
        observed_at=OBSERVED,
        fetch_html=_fetch,
    )

    assert result.family_decision.family_decision == "insufficient_sources_to_decide"


def test_writes_workbook_and_summary_to_explicit_temp_path(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(
        sources=(_source(domain="example.nl"),),
        observed_at=OBSERVED,
        fetch_html=_fetch,
        output_dir=tmp_path / "generated",
    )

    assert result.workbook_path == tmp_path / "generated" / "realworks_multi_source_validation_v1.xlsx"
    assert result.summary_csv_path == tmp_path / "generated" / "realworks_multi_source_validation_v1_summary.csv"
    assert result.workbook_path.exists()
    assert result.summary_csv_path.exists()
    workbook = load_workbook(result.workbook_path)
    assert tuple(workbook.sheetnames) == (
        "Sources Summary",
        "Per-source Metrics",
        "All Realworks Properties",
        "Field Gaps",
        "Warnings",
        "Problem Rows",
        "Family Decision",
    )


def test_does_not_write_raw_html_or_json(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_multi_source_validation(
        sources=(_source(domain="example.nl"),),
        observed_at=OBSERVED,
        fetch_html=_fetch,
        output_dir=tmp_path / "generated",
    )

    written_files = {path.name for path in (tmp_path / "generated").iterdir()}
    assert written_files == {
        "realworks_multi_source_validation_v1.xlsx",
        "realworks_multi_source_validation_v1_summary.csv",
    }
    assert not any(path.suffix.lower() in {".html", ".json", ".har"} for path in (tmp_path / "generated").iterdir())


def test_does_not_include_long_descriptions(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)
    long_description = "Ruime woning " * 100

    def fetch(url: str) -> str:
        if url.rstrip("/").endswith("/koop"):
            return _listing_html("example.nl")
        return _detail_html().replace("Compacte omschrijving.", long_description)

    result = run_realworks_multi_source_validation(
        sources=(_source(domain="example.nl"),),
        observed_at=OBSERVED,
        fetch_html=fetch,
        output_dir=tmp_path / "generated",
    )
    text = "\n".join(
        str(cell.value)
        for worksheet in load_workbook(result.workbook_path).worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert long_description.strip() not in text


def test_does_not_create_parser_per_makelaar() -> None:
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    class_names = {node.name.casefold() for node in ast.walk(tree) if isinstance(node, ast.ClassDef)}

    assert not any("oldenkotte" in name or "olden" in name for name in class_names)


def test_no_disallowed_imports() -> None:
    disallowed = {"requests", "httpx", "playwright", "selenium"}
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    imported_modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.update(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_modules.add(node.module)

    assert disallowed.isdisjoint({module.split(".")[0] for module in imported_modules})
