from __future__ import annotations

import ast
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR / "scraper" / "src"))

from domek_wonen.facts import build_property_fact_value, build_property_facts_record  # noqa: E402
from domek_wonen.parsers.models import ParsedListing  # noqa: E402
from domek_wonen.pilots.kin_excel_export import (  # noqa: E402
    PROPERTY_COLUMNS,
    export_kin_readiness_rows_to_excel,
)
from domek_wonen.pilots.kin_full_property_readiness import build_kin_property_readiness_row  # noqa: E402


MODULE_PATH = BASE_DIR / "scraper" / "src" / "domek_wonen" / "pilots" / "kin_excel_export.py"
NOW = datetime(2026, 6, 28, 10, 0, tzinfo=UTC)


def _listing(index: int = 1, **overrides: object) -> ParsedListing:
    listing = ParsedListing(
        source_id="kinmakelaars.nl__breda",
        source_domain="kinmakelaars.nl",
        canonical_url=f"https://kinmakelaars.nl/aanbod/wonen/breda/export-{index}/kin-export-{index:03d}",
        address_raw=f"Exportstraat {index}",
        street="Exportstraat",
        house_number=str(index),
        postcode="4811AA",
        city="Breda",
        asking_price_eur=425000 + index,
        transaction_type="koop",
        status="beschikbaar",
        living_area_m2=120 + index,
        rooms_count=5,
        bedrooms_count=3,
        property_type="Tussenwoning",
        energy_label="A",
        confidence_score=0.95,
    )
    values = {field: getattr(listing, field) for field in listing.__dataclass_fields__}
    values.update(overrides)
    return ParsedListing(**values)


def _fact(
    field: str,
    value: object,
    *,
    normalized_value: object | None = None,
    status: str = "usable",
    warnings: tuple[str, ...] = (),
):
    return build_property_fact_value(
        field=field,
        value=value,
        normalized_value=value if normalized_value is None else normalized_value,
        source="metadata",
        confidence=1.0,
        status=status,
        evidence_preview=str(value),
        warnings=warnings,
    )


def _record(index: int = 1, *facts, warnings: tuple[str, ...] = ()):
    return build_property_facts_record(
        source_id="kinmakelaars.nl__breda",
        source_domain="kinmakelaars.nl",
        canonical_url=f"https://kinmakelaars.nl/aanbod/wonen/breda/export-{index}/kin-export-{index:03d}",
        address_raw=f"Exportstraat {index}",
        city="Breda",
        status="beschikbaar",
        fetched_at="2026-06-28T10:00:00Z",
        expires_at="2026-07-12T10:00:00Z",
        facts=facts,
        warnings=warnings,
    )


def _full_record(index: int = 1):
    return _record(
        index,
        _fact("asking_price", 425000),
        _fact("property_type", "Tussenwoning", normalized_value="tussenwoning"),
        _fact("living_area_m2", 123),
        _fact("plot_area_m2", 234),
        _fact("rooms", 5),
        _fact("bedrooms", 3),
        _fact("bathrooms", 1),
        _fact("energy_label", "A++"),
        _fact("eigendomssituatie", "Volle eigendom", normalized_value="volle_eigendom"),
        _fact("vve_monthly_cost", 125),
        _fact("vve_active", True),
        _fact("heating_type", "CV-ketel"),
        _fact("cv_ketel_ownership", "eigendom"),
        _fact("garden", True),
        _fact("balcony", True),
        _fact("storage", True),
        _fact("garage", True),
        _fact("parking", "eigen oprit"),
        _fact("availability_date", "in overleg"),
        _fact("open_huis_badge_or_event", "Open huis"),
    )


def _akkerstraat_record(index: int = 1):
    return _record(
        index,
        _fact("asking_price", 450000),
        _fact("property_type", "Eengezinswoning", normalized_value="woonhuis"),
        _fact("living_area_m2", 130),
        _fact("plot_area_m2", 218),
        _fact("rooms", 4),
        _fact("bedrooms", 3),
        _fact("bathrooms", 1),
        _fact("floors", 3),
        _fact("volume_m3", 497),
        _fact("energy_label", "C"),
        _fact("heating_type", "Cv ketel", normalized_value="cv_ketel"),
        _fact("hot_water", "Cv ketel", normalized_value="cv_ketel"),
        _fact("cv_ketel_brand", "Remeha"),
        _fact("cv_ketel_ownership", "Huur", normalized_value="huur"),
        _fact("garden", "Achtertuin"),
        _fact("main_garden_area_m2", 63),
        _fact("parking", "Parkeervergunningen, Op eigen terrein"),
        _fact("garage", "vrijstaand_hout"),
        _fact("garage_count", 1),
    )


def _row(index: int = 1, record=None, **listing_overrides: object):
    return build_kin_property_readiness_row(_listing(index, **listing_overrides), record or _full_record(index))


def _export(tmp_path: Path, *rows):
    output_path = tmp_path / "nested" / "kin_export.xlsx"
    result = export_kin_readiness_rows_to_excel(
        rows=tuple(rows),
        output_path=output_path,
        source_id="kinmakelaars.nl__breda",
        generated_at=NOW,
    )
    workbook = load_workbook(output_path)
    return result, workbook


def _headers(worksheet):
    return [cell.value for cell in worksheet[1]]


def _col(headers, name: str) -> int:
    return headers.index(name) + 1


def _all_workbook_text(workbook) -> str:
    values: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
    return "\n".join(values)


def test_exports_one_synthetic_row_to_xlsx(tmp_path: Path) -> None:
    result, workbook = _export(tmp_path, _row(1))

    assert result.output_path.exists()
    assert result.rows_written == 1
    assert workbook["KIN Properties"].max_row == 2


def test_workbook_contains_required_worksheets(tmp_path: Path) -> None:
    result, workbook = _export(tmp_path, _row(1))

    assert result.worksheet_names == ("KIN Properties", "Summary", "Field Gaps", "Warnings", "Problem Rows")
    assert set(result.worksheet_names).issubset(set(workbook.sheetnames))


def test_required_columns_exist(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1))

    assert _headers(workbook["KIN Properties"]) == list(PROPERTY_COLUMNS)


def test_canonical_url_is_written_as_full_url_text(tmp_path: Path) -> None:
    row = _row(1)
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["KIN Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "canonical_url")).value == row.canonical_url


def test_property_link_is_real_hyperlink_with_open_listing_text(tmp_path: Path) -> None:
    row = _row(1)
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["KIN Properties"]
    headers = _headers(worksheet)
    cell = worksheet.cell(2, _col(headers, "property_link"))

    assert cell.value == "Open listing"
    assert cell.hyperlink.target == row.canonical_url


def test_missing_canonical_url_row_is_export_blocked_with_warning(tmp_path: Path) -> None:
    row = _row(1, canonical_url="")
    result, workbook = _export(tmp_path, row)
    worksheet = workbook["KIN Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "export_readiness")).value == "export_blocked"
    assert "missing_or_invalid_canonical_url" in result.warnings


def test_attention_missing_fields_and_warnings_are_joined_deterministically(tmp_path: Path) -> None:
    record = _record(
        1,
        _fact("asking_price", 425000),
        _fact("property_type", "Tussenwoning", normalized_value="tussenwoning", status="review", warnings=("needs_review",)),
        warnings=("record_warning",),
    )
    row = _row(1, record)
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["KIN Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "attention_points")).value.startswith("property_type staat in review.")
    assert worksheet.cell(2, _col(headers, "missing_key_fields")).value == "property_type; living_area_m2; bedrooms; energy_label; eigendomssituatie"
    assert worksheet.cell(2, _col(headers, "warnings")).value == "missing_coordinates; record_warning; needs_review"


def test_no_raw_html_json_or_image_urls_are_exported(tmp_path: Path) -> None:
    record = _record(
        1,
        _fact("asking_price", 425000),
        _fact("property_type", "<html>{\"docs\":[]}</html>", normalized_value="tussenwoning", status="review"),
        _fact("parking", "https://cdn.example.test/photo.jpg"),
    )
    _, workbook = _export(tmp_path, _row(1, record))
    text = _all_workbook_text(workbook).casefold()

    assert "<html" not in text
    assert '"docs"' not in text
    assert "photo.jpg" not in text


def test_akkerstraat_synthetic_row_exports_correct_critical_facts(tmp_path: Path) -> None:
    row = _row(
        1,
        _akkerstraat_record(1),
        canonical_url="https://www.kinmakelaars.nl/aanbod/wonen/6a3e86c10b3069e4614b02c1",
        address_raw="Akkerstraat 69",
        postcode="5025MG",
        city="Tilburg",
    )
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["KIN Properties"]
    headers = _headers(worksheet)
    text = _all_workbook_text(workbook).casefold()

    assert worksheet.cell(2, _col(headers, "bedrooms")).value == 3
    assert worksheet.cell(2, _col(headers, "energy_label")).value == "C"
    assert worksheet.cell(2, _col(headers, "heating_type")).value == "cv_ketel"
    assert worksheet.cell(2, _col(headers, "cv_ketel_ownership")).value == "huur"
    assert worksheet.cell(2, _col(headers, "eigendomssituatie")).value in (None, "")
    assert worksheet.cell(2, _col(headers, "property_link")).hyperlink.target == row.canonical_url
    assert "heating\":{\"heating\"" not in text
    assert "10 slaapkamers" not in text
    assert "erfpacht" not in text


def test_summary_counts_match_input_rows(tmp_path: Path) -> None:
    blocked = _row(2, _record(2, _fact("property_type", "Tussenwoning", normalized_value="tussenwoning")), canonical_url="")
    _, workbook = _export(tmp_path, _row(1), blocked)
    values = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}

    assert values["rows_written"] == 2
    assert values["export_ready_count"] == 1
    assert values["export_blocked_count"] == 1
    assert values["client_ready_count"] == 1


def test_field_gap_sheet_counts_missing_review_usable_correctly(tmp_path: Path) -> None:
    review = _row(2, _record(2, _fact("property_type", "Tussenwoning", normalized_value="tussenwoning", status="review")))
    _, workbook = _export(tmp_path, _row(1), review, _row(3, _record(3)))
    rows = {row[0].value: [cell.value for cell in row] for row in workbook["Field Gaps"].iter_rows(min_row=2)}

    property_type = rows["property_type"]
    assert property_type[1:5] == [1, 1, 1, 3]
    assert property_type[5] == 0.3333
    assert property_type[6] == 0.3333


def test_warning_sheet_aggregates_warnings(tmp_path: Path) -> None:
    row_one = _row(1, _record(1, warnings=("shared_warning",)))
    row_two = _row(2, _record(2, warnings=("shared_warning",)))
    _, workbook = _export(tmp_path, row_one, row_two)
    values = {row[0].value: row[1].value for row in workbook["Warnings"].iter_rows(min_row=2)}

    assert values["shared_warning"] == 2
    assert values["missing_coordinates"] == 2


def test_output_path_parent_is_created_and_no_other_files_are_written(tmp_path: Path) -> None:
    result, _ = _export(tmp_path, _row(1))

    assert result.output_path == tmp_path / "nested" / "kin_export.xlsx"
    assert sorted(path.relative_to(tmp_path).as_posix() for path in tmp_path.rglob("*") if path.is_file()) == [
        "nested/kin_export.xlsx"
    ]


def test_table_formatting_is_applied(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1))
    worksheet = workbook["KIN Properties"]

    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref is not None
    assert worksheet["A1"].font.bold is True


def test_does_not_import_playwright_selenium_requests_httpx() -> None:
    disallowed = {"requests", "httpx", "playwright", "selenium"}
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    imported_modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.update(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_modules.add(node.module)

    assert disallowed.isdisjoint({module.split(".")[0] for module in imported_modules})


def test_export_result_counts_are_reported(tmp_path: Path) -> None:
    review = replace(_row(2, _record(2, _fact("asking_price", 425000))), quality_status="advisor_review")
    result, _ = _export(tmp_path, _row(1), review)

    assert result.export_ready_count == 1
    assert result.export_review_count == 1
    assert result.client_ready_count == 1
    assert result.advisor_review_count == 1
    assert ("property_type", 1) in result.missing_field_counts
