from __future__ import annotations

import ast
import csv
from datetime import UTC, datetime
from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR / "scraper" / "src"))

from domek_wonen.compliance import robots_gate  # noqa: E402
from domek_wonen.facts import build_property_fact_value, build_property_facts_record  # noqa: E402
from domek_wonen.parsers.models import ParsedListing  # noqa: E402
from domek_wonen.pilots.noord_brabant_realworks_audit import (  # noqa: E402
    EXPECTED_NB_REALWORKS_AUDIT_ROWS,
    NoordBrabantRealworksAggregateMetrics,
    NoordBrabantRealworksAuditInputError,
    NoordBrabantRealworksAuditMetrics,
    decide_noord_brabant_realworks_family,
    load_noord_brabant_realworks_audit_sources,
    resolve_noord_brabant_realworks_audit,
    run_noord_brabant_realworks_audit,
    validate_noord_brabant_realworks_audit_input_rows,
)
from domek_wonen.pilots.realworks_property_readiness import (  # noqa: E402
    RealworksPropertyReadinessResult,
    build_realworks_property_readiness_row,
)


MODULE_PATH = BASE_DIR / "scraper" / "src" / "domek_wonen" / "pilots" / "noord_brabant_realworks_audit.py"
NOW = datetime(2026, 7, 1, tzinfo=UTC)


def _input_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "source_id": "example.nl__breda",
        "source_name": "Example Makelaars",
        "domain": "example.nl",
        "accepted_aanbod_url": "https://example.nl/aanbod/woningaanbod/breda/koop",
        "coverage_city": "Breda",
        "coverage_gemeente": "Breda",
        "coverage_province": "Noord-Brabant",
        "aanbod_scope_status": "confirmed_nb_scope",
        "realworks_verification_status": "verified",
        "realworks_evidence_strength": "strong",
        "audit_input_status": "ready_for_noord_brabant_realworks_audit",
        "parser_family_candidate": "realworks_public",
    }
    row.update(overrides)
    return row


def _write_input(path: Path, rows: list[dict[str, object]]) -> Path:
    columns = list(_input_row().keys())
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})
    return path


def _valid_rows(count: int = EXPECTED_NB_REALWORKS_AUDIT_ROWS) -> list[dict[str, object]]:
    return [
        _input_row(
            source_id=f"example{i}.nl__breda",
            source_name=f"Example {i}",
            domain=f"example{i}.nl",
            accepted_aanbod_url=f"https://example{i}.nl/aanbod/woningaanbod/breda/koop",
        )
        for i in range(count)
    ]


def _listing(**overrides: object) -> ParsedListing:
    listing = ParsedListing(
        source_id="example0.nl__breda",
        source_domain="example0.nl",
        canonical_url="https://example0.nl/aanbod/woningaanbod/breda/koop/huis-123-teststraat-1",
        address_raw="Teststraat 1",
        street="Teststraat",
        house_number="1",
        postcode="4811 AA",
        city="Breda",
        asking_price_eur=425000,
        transaction_type="koop",
        status="beschikbaar",
        living_area_m2=123,
        rooms_count=5,
        bedrooms_count=3,
        property_type="woonhuis",
        confidence_score=0.95,
    )
    values = {field: getattr(listing, field) for field in listing.__dataclass_fields__}
    values.update(overrides)
    return ParsedListing(**values)


def _fact(field: str, value: object, *, normalized_value: object | None = None, status: str = "usable"):
    return build_property_fact_value(
        field=field,
        value=value,
        normalized_value=value if normalized_value is None else normalized_value,
        source="realworks_kenmerk",
        confidence=0.90,
        status=status,
        evidence_preview=str(value or ""),
    )


def _record():
    return build_property_facts_record(
        source_id="example0.nl__breda",
        source_domain="example0.nl",
        canonical_url="https://example0.nl/aanbod/woningaanbod/breda/koop/huis-123-teststraat-1",
        address_raw="Teststraat 1",
        city="Breda",
        status="beschikbaar",
        fetched_at="2026-07-01T00:00:00Z",
        expires_at="2026-07-15T00:00:00Z",
        facts=(
            _fact("asking_price", 425000),
            _fact("property_type", "Woonhuis", normalized_value="woonhuis"),
            _fact("living_area_m2", 123),
            _fact("bedrooms", 3),
            _fact("energy_label", "A"),
            _fact("eigendomssituatie", "Volle eigendom", normalized_value="volle_eigendom"),
        ),
    )


def _readiness(
    *,
    source_id: str = "example0.nl__breda",
    source_domain: str = "example0.nl",
    parser_total: int = 1,
    qa_clean: int = 1,
    rows_built: bool = True,
    warnings: tuple[str, ...] = (),
) -> RealworksPropertyReadinessResult:
    rows = ()
    if rows_built:
        row = build_realworks_property_readiness_row(
            _listing(source_id=source_id, source_domain=source_domain),
            _record(),
            observed_at=NOW,
        )
        rows = (row,)
    return RealworksPropertyReadinessResult(
        source_id=source_id,
        source_domain=source_domain,
        listing_parser_total=parser_total,
        listing_qa_clean=qa_clean,
        listing_qa_review=0,
        listing_qa_rejected=0,
        detail_attempted=1 if rows_built else 0,
        detail_succeeded=1 if rows_built else 0,
        detail_failed=0,
        facts_records_built=1 if rows_built else 0,
        readiness_rows_built=len(rows),
        quality_status_counts=(("advisor_review", len(rows)),),
        export_readiness_counts=(("export_review", len(rows)),),
        field_completion_counts=(("postcode", 0, 0, 1),),
        missing_key_fields_counts=(("postcode", 1),) if rows else (),
        review_fields_counts=(),
        warning_counts=(),
        sample_rows_compact=(),
        problem_rows_compact=(),
        excel_validation_ready=False,
        rows=rows,
        warnings=warnings,
    )


def test_loads_only_realworks_audit_input_and_rejects_missing_csv(tmp_path: Path) -> None:
    with pytest.raises(NoordBrabantRealworksAuditInputError, match="missing input CSV"):
        load_noord_brabant_realworks_audit_sources(tmp_path / "missing.csv")

    path = _write_input(tmp_path / "input.csv", _valid_rows())
    sources = load_noord_brabant_realworks_audit_sources(path)
    assert len(sources) == 65
    assert {source.parser_family_candidate for source in sources} == {"realworks_public"}
    assert all(source.accepted_aanbod_url for source in sources)


@pytest.mark.parametrize(
    ("overrides", "message"),
    [
        ({"audit_input_status": "needs_manual_scope_check"}, "audit_input_non_ready_count=1"),
        ({"parser_family_candidate": "wordpress_json"}, "audit_input_non_realworks_count=1"),
        ({"source_id": "kinmakelaars.nl__breda", "domain": "kinmakelaars.nl"}, "audit_input_kin_count=1"),
        ({"accepted_aanbod_url": ""}, "audit_input_missing_accepted_aanbod_url_count=1"),
        ({"accepted_aanbod_url": "https://example0.nl/aanbod/woningaanbod/koop/huis-123-test"}, "audit_input_property_detail_url_count=1"),
        ({"accepted_aanbod_url": "https://www.funda.nl/koop/breda/huis-123"}, "audit_input_property_detail_url_count=1"),
    ],
)
def test_rejects_contaminated_input_rows(tmp_path: Path, overrides: dict[str, object], message: str) -> None:
    rows = _valid_rows()
    rows[0].update(overrides)
    path = _write_input(tmp_path / "input.csv", rows)
    with pytest.raises(NoordBrabantRealworksAuditInputError, match=message):
        load_noord_brabant_realworks_audit_sources(path)


def test_reports_duplicate_input_rows(tmp_path: Path) -> None:
    rows = _valid_rows()
    rows[1]["domain"] = rows[0]["domain"]
    rows[1]["accepted_aanbod_url"] = rows[0]["accepted_aanbod_url"]
    path = _write_input(tmp_path / "input.csv", rows)
    with pytest.raises(NoordBrabantRealworksAuditInputError, match="duplicate domain"):
        load_noord_brabant_realworks_audit_sources(path)


def test_missing_required_columns_are_not_reported_as_non_realworks_rows(tmp_path: Path) -> None:
    rows = _valid_rows()
    path = tmp_path / "legacy.csv"
    columns = [column for column in _input_row().keys() if column not in {"parser_family_candidate", "coverage_province"}]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})

    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        loaded = list(reader)
        validation = validate_noord_brabant_realworks_audit_input_rows(loaded, columns=reader.fieldnames or ())

    assert validation.audit_input_missing_required_columns_count == 2
    assert validation.audit_input_missing_parser_family_candidate_column == 1
    assert validation.audit_input_missing_coverage_province_column == 1
    assert validation.audit_input_non_realworks_count == 0
    with pytest.raises(NoordBrabantRealworksAuditInputError, match="parser_family_candidate"):
        load_noord_brabant_realworks_audit_sources(path)


def test_rejects_non_official_domain_url(tmp_path: Path) -> None:
    rows = _valid_rows()
    rows[0]["accepted_aanbod_url"] = "https://other.nl/aanbod/woningaanbod/breda/koop"
    path = _write_input(tmp_path / "input.csv", rows)
    with pytest.raises(NoordBrabantRealworksAuditInputError, match="official domain"):
        load_noord_brabant_realworks_audit_sources(path)


def test_runner_continues_after_source_failure_and_writes_outputs(tmp_path: Path) -> None:
    input_csv = _write_input(tmp_path / "input.csv", _valid_rows())
    calls: list[str] = []

    def fake_runner(**kwargs):
        calls.append(kwargs["source_id"])
        if len(calls) == 1:
            return _readiness(source_id=kwargs["source_id"], source_domain=kwargs["source_domain"])
        raise RuntimeError("boom")

    result = run_noord_brabant_realworks_audit(
        input_csv=input_csv,
        output_workbook=tmp_path / "audit.xlsx",
        output_summary=tmp_path / "summary.csv",
        output_problem_sources=tmp_path / "problems.csv",
        max_sources=2,
        runtime_budget_seconds=60,
        observed_at=NOW,
        readiness_runner=fake_runner,
    )

    assert calls == ["example0.nl__breda", "example1.nl__breda"]
    assert result.aggregate_metrics.sources_attempted == 2
    assert result.aggregate_metrics.sources_fetch_failed == 1
    assert result.workbook_path and result.workbook_path.exists()
    assert result.summary_csv_path and result.summary_csv_path.exists()
    assert result.problem_sources_csv_path and result.problem_sources_csv_path.exists()
    workbook = load_workbook(result.workbook_path)
    assert {
        "Source Summary",
        "Realworks Properties",
        "Field Gaps",
        "Warnings",
        "Problem Sources",
        "Parser Failure Patterns",
        "Access Policy",
        "Audit Input",
        "Family Audit Decision",
        "Manual Verification",
    }.issubset(set(workbook.sheetnames))
    manual_headers = [cell.value for cell in workbook["Manual Verification"][1]]
    assert set(
        [
            "source_id",
            "domain",
            "address",
            "city",
            "price",
            "status",
            "accepted_aanbod_url",
            "canonical_url",
            "export_readiness",
            "quality_status",
            "missing_key_fields",
            "warnings",
            "manual_check_result",
            "manual_check_notes",
        ]
    ).issubset(set(manual_headers))


def test_stable_partial_result_on_runtime_budget_exhaustion(tmp_path: Path) -> None:
    input_csv = _write_input(tmp_path / "input.csv", _valid_rows())
    result = run_noord_brabant_realworks_audit(
        input_csv=input_csv,
        max_sources=65,
        runtime_budget_seconds=0,
        observed_at=NOW,
        readiness_runner=lambda **kwargs: _readiness(source_id=kwargs["source_id"], source_domain=kwargs["source_domain"]),
    )
    assert result.aggregate_metrics.sources_attempted == 0
    assert result.aggregate_metrics.runtime_budget_exhausted is True
    assert result.family_decision.family_decision == "insufficient_successful_sources"


def test_no_current_listings_are_separate_from_parser_failure(tmp_path: Path) -> None:
    input_csv = _write_input(tmp_path / "input.csv", _valid_rows())
    result = run_noord_brabant_realworks_audit(
        input_csv=input_csv,
        max_sources=1,
        observed_at=NOW,
        readiness_runner=lambda **kwargs: _readiness(
            source_id=kwargs["source_id"],
            source_domain=kwargs["source_domain"],
            parser_total=0,
            qa_clean=0,
            rows_built=False,
            warnings=("no_realworks_detail_urls_found",),
        ),
    )
    metric = result.metrics[0]
    assert metric.validation_status == "no_current_listings"
    assert metric.no_current_listings is True


def test_calls_robots_gate_before_listing_and_detail_fetch(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_csv = _write_input(tmp_path / "input.csv", _valid_rows())
    robots_calls: list[tuple[str, str]] = []
    fetch_calls: list[str] = []
    listing_url = "https://example0.nl/aanbod/woningaanbod/breda/koop"
    detail_url = "https://example0.nl/aanbod/woningaanbod/breda/koop/huis-123-teststraat-1"

    def fake_can_fetch(domain: str, path: str = "/") -> bool:
        robots_calls.append((domain, path))
        return True

    def fake_fetch(url: str) -> str:
        fetch_calls.append(url)
        if url == listing_url:
            return """
            <html><body>
              <article class="aanbodEntry">
                <a href="/aanbod/woningaanbod/breda/koop/huis-123-teststraat-1">
                  <h2>Teststraat 1</h2>
                  <span class="locality">Breda</span>
                  <span class="price">EUR 425.000 k.k.</span>
                  <span class="status">Beschikbaar</span>
                  <span>123 m2 5 kamers Woonhuis</span>
                </a>
              </article>
            </body></html>
            """
        if url == detail_url:
            return """
            <html><body>
              <span class="kenmerk"><span class="kenmerkName">Soort object</span><span class="kenmerkValue">Woonhuis</span></span>
              <span class="kenmerk"><span class="kenmerkName">Vraagprijs</span><span class="kenmerkValue">EUR 425.000 k.k.</span></span>
              <span class="kenmerk"><span class="kenmerkName">Status</span><span class="kenmerkValue">Beschikbaar</span></span>
              <span class="kenmerk"><span class="kenmerkName">Woonoppervlakte</span><span class="kenmerkValue">123 m2</span></span>
            </body></html>
            """
        raise AssertionError(url)

    monkeypatch.setattr(robots_gate, "can_fetch", fake_can_fetch)
    result = run_noord_brabant_realworks_audit(
        input_csv=input_csv,
        max_sources=1,
        max_detail_per_source=1,
        observed_at=NOW,
        fetch_html=fake_fetch,
    )

    assert fetch_calls == [listing_url, detail_url]
    assert robots_calls[0] == ("example0.nl", "/aanbod/woningaanbod/breda/koop")
    assert ("example0.nl", "/aanbod/woningaanbod/breda/koop/huis-123-teststraat-1") in robots_calls
    assert result.aggregate_metrics.total_detail_attempted == 1


def test_does_not_fetch_when_robots_disallow_listing(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_csv = _write_input(tmp_path / "input.csv", _valid_rows())
    fetch_calls: list[str] = []

    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path="/": False)

    result = run_noord_brabant_realworks_audit(
        input_csv=input_csv,
        max_sources=1,
        observed_at=NOW,
        fetch_html=lambda url: fetch_calls.append(url) or "",
    )

    assert fetch_calls == []
    assert result.metrics[0].validation_status == "blocked_by_robots"
    assert result.family_decision.family_decision == "blocked_by_access_policy"


def test_decision_variants() -> None:
    ready_metric = _metric("passed_with_review_gaps")
    assert (
        decide_noord_brabant_realworks_family(
            (ready_metric, ready_metric, ready_metric),
            _aggregate((ready_metric, ready_metric, ready_metric)),
        ).family_decision
        == "realworks_ready_for_nb_family_coverage"
    )
    blocked_metric = _metric("blocked_by_robots")
    assert (
        decide_noord_brabant_realworks_family(
            (ready_metric, blocked_metric, blocked_metric),
            _aggregate((ready_metric, blocked_metric, blocked_metric)),
        ).family_decision
        == "blocked_by_access_policy"
    )
    hard_metric = _metric("needs_realworks_hardening")
    assert (
        decide_noord_brabant_realworks_family(
            (ready_metric, hard_metric, hard_metric, hard_metric),
            _aggregate((ready_metric, hard_metric, hard_metric, hard_metric)),
        ).family_decision
        == "realworks_needs_hardening_v2"
    )
    failed_metric = _metric("listing_fetch_failed")
    assert (
        decide_noord_brabant_realworks_family(
            (ready_metric, ready_metric, failed_metric),
            _aggregate((ready_metric, ready_metric, failed_metric)),
        ).family_decision
        == "realworks_partially_ready_with_exclusions"
    )


def test_resolution_maps_source_statuses_and_writes_outputs(tmp_path: Path) -> None:
    summary = tmp_path / "summary.csv"
    _write_resolution_summary(
        summary,
        [
            _summary_row("ready.nl__breda", "passed"),
            _summary_row("review.nl__breda", "passed_with_review_gaps"),
            _summary_row("empty.nl__breda", "no_current_listings", parser_total=0, qa_clean=0, rows=0),
            _summary_row("failed.nl__breda", "listing_fetch_failed", parser_total=0, qa_clean=0, rows=0),
        ],
    )
    workbook = _write_resolution_workbook(tmp_path / "audit.xlsx", summary)
    result = resolve_noord_brabant_realworks_audit(
        audit_summary_csv=summary,
        audit_workbook=workbook,
        output_csv=tmp_path / "resolution.csv",
        output_workbook=tmp_path / "resolution.xlsx",
    )
    assert result.metrics["source_resolution_ready_count"] == 1
    assert result.metrics["source_resolution_ready_with_review_gaps_count"] == 1
    assert result.metrics["source_resolution_monitor_no_current_count"] == 1
    assert result.metrics["source_resolution_exclude_fetch_failed_count"] == 1
    assert result.metrics["systemic_hardening_pattern_count"] == 0
    assert result.decision.final_decision == "realworks_partially_ready_with_exclusions"
    assert result.output_csv_path and result.output_csv_path.exists()
    output = load_workbook(result.output_workbook_path)
    assert {
        "Source Resolution",
        "Ready Sources",
        "Ready With Review Gaps",
        "No Current Listings",
        "Excluded Sources",
        "Manual Review",
        "Hardening Candidates",
        "Property QA",
        "Duplicate Properties",
        "Readiness Label Checks",
        "Status Consistency",
        "Field Sanity",
        "Decision Summary",
    }.issubset(output.sheetnames)


def test_resolution_triggers_hardening_v2_on_repeated_parser_failures(tmp_path: Path) -> None:
    summary = tmp_path / "summary.csv"
    _write_resolution_summary(
        summary,
        [
            _summary_row(f"hard{i}.nl__breda", "needs_realworks_hardening", parser_total=5, qa_clean=0, rows=0)
            for i in range(3)
        ],
    )
    result = resolve_noord_brabant_realworks_audit(
        audit_summary_csv=summary,
        audit_workbook=_write_resolution_workbook(tmp_path / "audit.xlsx", summary),
    )
    assert result.metrics["source_resolution_needs_hardening_count"] == 3
    assert result.metrics["systemic_hardening_pattern_count"] > 0
    assert result.decision.final_decision == "realworks_needs_hardening_v2"


def test_property_qa_detects_duplicates_and_hard_gate_issues(tmp_path: Path) -> None:
    summary = tmp_path / "summary.csv"
    _write_resolution_summary(summary, [_summary_row("example.nl__breda", "passed")])
    workbook = _write_resolution_workbook(
        tmp_path / "audit.xlsx",
        summary,
        manual_rows=[
            _manual_row(),
            _manual_row(),
            _manual_row(canonical_url="https://example.nl/aanbod/woningaanbod/breda/koop/huis-456", price="", missing_key_fields="asking_price", export_readiness="export_ready"),
            _manual_row(canonical_url="https://example.nl/aanbod/woningaanbod/breda/koop/huis-789", status="inactive_sold", export_readiness="export_ready"),
        ],
    )
    result = resolve_noord_brabant_realworks_audit(audit_summary_csv=summary, audit_workbook=workbook)
    assert result.metrics["duplicate_canonical_url_count"] >= 2
    assert result.metrics["duplicate_source_canonical_url_count"] >= 2
    assert result.metrics["same_source_possible_duplicate_property_count"] >= 2
    assert result.metrics["cross_source_possible_duplicate_property_count"] >= 2
    assert result.metrics["export_ready_with_critical_missing_count"] == 1
    assert result.metrics["property_qa_duplicate_review_count"] >= 2
    assert result.metrics["property_qa_blocked_count"] >= 1


def test_static_scope_guardrails() -> None:
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    imported = {
        alias.name
        for node in ast.walk(tree)
        if isinstance(node, (ast.Import, ast.ImportFrom))
        for alias in node.names
    }
    source = MODULE_PATH.read_text(encoding="utf-8")
    forbidden_import_markers = ("matching", "n8n", "dashboard", "migrations", "sqlite", "playwright", "selenium")
    assert not any(marker in " ".join(imported).casefold() for marker in forbidden_import_markers)
    assert "data/raw" not in source.replace("\\", "/")
    assert "download" not in source.casefold()
    assert "ParserFamilyRunner" not in source
    assert "raw_html" not in source and "raw_json" not in source


def _summary_row(
    source_id: str,
    status: str,
    *,
    parser_total: int = 1,
    qa_clean: int = 1,
    rows: int = 1,
) -> dict[str, object]:
    domain = source_id.split("__", 1)[0]
    return {
        "source_id": source_id,
        "source_name": source_id,
        "domain": domain,
        "accepted_aanbod_url": f"https://{domain}/aanbod/woningaanbod/breda/koop",
        "coverage_city": "Breda",
        "coverage_gemeente": "Breda",
        "aanbod_scope_status": "confirmed_nb_scope",
        "access_policy_status": "allowed",
        "robots_listing_allowed": "True",
        "listing_fetch_status": "success",
        "listing_http_status": "ok",
        "parser_total": parser_total,
        "parser_qa_clean": qa_clean,
        "parser_qa_review": 0,
        "parser_qa_rejected": 0,
        "detail_attempted": rows,
        "detail_succeeded": rows,
        "detail_failed": 0,
        "readiness_rows_built": rows,
        "export_ready": rows,
        "export_review": 0,
        "export_blocked": 0,
        "active_inventory_eligible": rows,
        "inactive_status_count": 0,
        "non_residential_count": 0,
        "no_current_listings": status == "no_current_listings",
        "top_warning": "",
        "validation_status": status,
        "validation_decision": status,
        "recommended_next_action": status,
    }


def _write_resolution_summary(path: Path, rows: list[dict[str, object]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(_summary_row("example.nl__breda", "passed").keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})
    return path


def _manual_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "source_id": "example.nl__breda",
        "domain": "example.nl",
        "address": "Teststraat 1",
        "city": "Breda",
        "price": "425000",
        "status": "active_available",
        "accepted_aanbod_url": "https://example.nl/aanbod/woningaanbod/breda/koop",
        "canonical_url": "https://example.nl/aanbod/woningaanbod/breda/koop/huis-123",
        "property_link": "https://example.nl/aanbod/woningaanbod/breda/koop/huis-123",
        "export_readiness": "export_ready",
        "quality_status": "client_ready",
        "missing_key_fields": "",
        "warnings": "",
        "manual_check_result": "",
        "manual_check_notes": "",
    }
    row.update(overrides)
    return row


def _write_resolution_workbook(path: Path, summary: Path, manual_rows: list[dict[str, object]] | None = None) -> Path:
    with summary.open(encoding="utf-8", newline="") as handle:
        source_rows = list(csv.DictReader(handle))
    manual_rows = manual_rows or [_manual_row()]
    workbook = load_workbook(summary) if False else None
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Source Summary"
    _append_dict_sheet(ws, source_rows)
    manual_ws = wb.create_sheet("Manual Verification")
    _append_dict_sheet(manual_ws, manual_rows)
    props_ws = wb.create_sheet("Realworks Properties")
    property_rows = []
    for row in manual_rows:
        property_rows.append(
            {
                "source_id": row["source_id"],
                "source_domain": row["domain"],
                "canonical_url": row["canonical_url"],
                "address": row["address"],
                "postcode": "4811 AA",
                "city": row["city"],
                "asking_price": row["price"],
                "status_bucket": row["status"],
                "quality_status": row["quality_status"],
                "export_readiness": row["export_readiness"],
                "active_inventory_eligible": "False",
                "residential_classification": "residential",
                "freshness_bucket": "unknown_age",
                "lifecycle_events": "",
                "missing_key_fields": row["missing_key_fields"],
                "review_fields": "",
                "warnings": row["warnings"],
            }
        )
    _append_dict_sheet(props_ws, property_rows)
    wb.save(path)
    return path


def _append_dict_sheet(worksheet, rows: list[dict[str, object]]) -> None:
    columns = list(rows[0].keys()) if rows else ["empty"]
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])


def _metric(status: str) -> NoordBrabantRealworksAuditMetrics:
    return NoordBrabantRealworksAuditMetrics(
        source_id=f"{status}.nl__breda",
        source_name=status,
        domain=f"{status}.nl",
        accepted_aanbod_url=f"https://{status}.nl/aanbod/woningaanbod/breda/koop",
        coverage_city="Breda",
        coverage_gemeente="Breda",
        aanbod_scope_status="confirmed_nb_scope",
        access_policy_status="allowed",
        robots_listing_allowed=status != "blocked_by_robots",
        listing_fetch_status=status,
        listing_http_status="ok",
        parser_total=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        parser_qa_clean=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        parser_qa_review=0,
        parser_qa_rejected=0,
        detail_attempted=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        detail_succeeded=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        detail_failed=0,
        readiness_rows_built=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        export_ready=0,
        export_review=1 if status in {"passed", "passed_with_review_gaps"} else 0,
        export_blocked=0,
        active_inventory_eligible=0,
        inactive_status_count=0,
        non_residential_count=0,
        no_current_listings=status == "no_current_listings",
        top_warning=status,
        warning_counts=((status, 1),),
        field_gap_counts=(("postcode", 1),),
        failure_patterns=(status,),
        validation_status=status,
        validation_decision=status,
        recommended_next_action=status,
    )


def _aggregate(metrics: tuple[NoordBrabantRealworksAuditMetrics, ...]) -> NoordBrabantRealworksAggregateMetrics:
    from domek_wonen.pilots.noord_brabant_realworks_audit import aggregate_noord_brabant_realworks_metrics

    return aggregate_noord_brabant_realworks_metrics(
        input_sources=len(metrics),
        attempted_sources=len(metrics),
        metrics=metrics,
        runtime_budget_exhausted=False,
    )
