from __future__ import annotations

import csv
from pathlib import Path
import subprocess
import sys

from openpyxl import load_workbook

from domek_wonen.sources.makelaar_universe import (
    build_noord_brabant_makelaar_universe,
    compute_quality_metrics,
    load_funda_makelaar_observations,
    write_makelaar_universe_csv,
    write_makelaar_universe_workbook,
)


FUNDA_HEADER = (
    "city",
    "province",
    "funda_search_url",
    "page_number",
    "benchmark_count",
    "makelaar_name_raw",
    "makelaar_name_clean_candidate",
    "name_truncated",
    "generic_multiple_makelaars",
    "name_needs_manual_review",
    "evidence_text",
    "collection_status",
    "collected_at",
)


def _write_csv(path: Path, header: tuple[str, ...], rows: list[dict[str, object]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _funda_row(
    name: str,
    city: str = "Tilburg",
    page_number: int = 1,
    benchmark_count: str = "1.115 koopwoningen in Tilburg",
    name_truncated: str = "false",
    generic_multiple_makelaars: str = "false",
    name_needs_manual_review: str = "false",
) -> dict[str, object]:
    return {
        "city": city,
        "province": "Noord-Brabant",
        "funda_search_url": f"https://www.funda.nl/zoeken/koop?selected_area=[{city.lower()}]&page={page_number}",
        "page_number": page_number,
        "benchmark_count": benchmark_count,
        "makelaar_name_raw": name,
        "makelaar_name_clean_candidate": name,
        "name_truncated": name_truncated,
        "generic_multiple_makelaars": generic_multiple_makelaars,
        "name_needs_manual_review": name_needs_manual_review,
        "evidence_text": name,
        "collection_status": "ok",
        "collected_at": "2026-07-02T20:42:44.640Z",
    }


def _completion_row(
    source_id: str,
    source_name: str,
    domain: str,
    accepted_aanbod_url: str = "",
    aanbod_url_status: str = "unknown_explicit",
    delivery_mode: str = "",
    parser_family_candidate: str = "",
) -> dict[str, str]:
    return {
        "source_id": source_id,
        "source_name": source_name,
        "domain": domain,
        "root_url": f"https://{domain}" if domain else "",
        "office_city": "",
        "office_gemeente": "",
        "office_province": "",
        "office_location_status": "",
        "coverage_city": "Tilburg",
        "coverage_gemeente": "Tilburg",
        "coverage_province": "Noord-Brabant",
        "coverage_location_status": "known",
        "outside_office_coverage_status": "",
        "has_noord_brabant_coverage": "yes",
        "accepted_aanbod_url": accepted_aanbod_url,
        "aanbod_url_status": aanbod_url_status,
        "aanbod_url_confidence": "0.95",
        "aanbod_scope_status": "",
        "platform_guess": delivery_mode,
        "delivery_mode": delivery_mode,
        "parser_family_candidate": parser_family_candidate,
        "family_confidence": "0.95",
        "family_terminal_status": "",
        "realworks_verification_status": "",
        "access_policy_status": "",
        "terminal_status": "",
        "validation_status": "",
        "recommended_next_action": "build_source_config" if parser_family_candidate else "",
        "manual_check_result": "",
        "manual_check_notes": "",
    }


def _master_row(source_id: str, office_name: str, domain: str, aanbod_url: str = "") -> dict[str, str]:
    return {
        "source_id": source_id,
        "office_name": office_name,
        "root_domain": domain,
        "website": f"https://{domain}" if domain else "",
        "gemeente": "Tilburg",
        "province": "Noord-Brabant",
        "source_origin": "fixture",
        "aanbod_url": aanbod_url,
        "aanbod_url_quality": "valid" if aanbod_url else "missing",
        "aanbod_url_type": "",
        "confidence_score": "0.9",
        "score": "100",
        "source_quality_status": "",
        "source_quality_reason": "",
        "needs_review": "false",
        "review_reason": "",
        "legal_status": "allowed_official_source",
        "last_seen_at": "",
        "last_audited_at": "",
        "run_id": "fixture",
        "is_active": "true",
    }


def _fingerprint_row(source_id: str, office_name: str, domain: str, detected_platform: str) -> dict[str, str]:
    return {
        "source_id": source_id,
        "office_name": office_name,
        "root_domain": domain,
        "website_url": f"https://{domain}" if domain else "",
        "aanbod_url": "",
        "detected_platform": detected_platform,
        "confidence": "0.9",
        "evidence": "",
        "parser_priority": "p1",
        "recommended_next_action": "build_source_config",
        "fetch_status": "",
        "error": "",
    }


def _seed_row(office_name: str, domain: str, koopaanbod_url: str = "") -> dict[str, str]:
    return {
        "office_name": office_name,
        "website": f"https://{domain}" if domain else "",
        "domain": domain,
        "root_domain": domain,
        "koopaanbod_url": koopaanbod_url,
        "koopaanbod_url_quality": "valid" if koopaanbod_url else "missing",
        "plaats": "Tilburg",
        "provincie": "Noord-Brabant",
        "source_type": "makelaar_site",
        "discovery_source": "fixture",
        "confidence": "0.8",
        "needs_review": "false",
        "review_reason": "",
        "notes": "",
    }


def _resolution_row(raw_source_name: str, resolved_domain: str, resolved_source_id: str = "") -> dict[str, str]:
    return {
        "raw_source_name": raw_source_name,
        "raw_source_id": "",
        "raw_gemeente": "Tilburg",
        "raw_province": "Noord-Brabant",
        "evidence_file": "fixture.csv",
        "initial_reason": "missing_official_domain",
        "resolution_status": "resolved_to_existing_source" if resolved_source_id else "resolved_to_new_source",
        "resolved_domain": resolved_domain,
        "resolved_source_id": resolved_source_id,
        "matched_existing_source": "True" if resolved_source_id else "False",
        "created_new_source": "False" if resolved_source_id else "True",
        "confidence": "0.9",
        "attempt_count": "2",
        "candidate_count": "2",
        "evidence_preview": "",
        "resolution_reason": "",
        "suggested_next_action": "manual_official_domain_research_required",
        "manual_check_result": "",
        "manual_check_notes": "",
    }


def _build_result(tmp_path: Path, funda_rows: list[dict[str, object]], completion_rows: list[dict[str, str]] | None = None, master_rows: list[dict[str, str]] | None = None, fingerprint_rows: list[dict[str, str]] | None = None, seed_rows: list[dict[str, str]] | None = None, resolution_rows: list[dict[str, str]] | None = None):
    funda_path = _write_csv(tmp_path / "tmp" / "manual_inputs" / "funda_makelaar_names" / "funda_nb_makelaar_names_raw.csv", FUNDA_HEADER, funda_rows)
    completion_path = _write_csv(
        tmp_path / "tmp" / "generated" / "noord_brabant_source_completion_scope_verification_v2.csv",
        tuple(_completion_row("x", "x", "x").keys()),
        completion_rows or [],
    )
    master_path = _write_csv(
        tmp_path / "data" / "discovery" / "runs" / "20260614T122022Z" / "makelaar_sources_master.csv",
        tuple(_master_row("x", "x", "x").keys()),
        master_rows or [],
    )
    fingerprint_path = _write_csv(
        tmp_path / "data" / "discovery" / "platform_fingerprint" / "platform_fingerprint_results.csv",
        tuple(_fingerprint_row("x", "x", "x", "realworks").keys()),
        fingerprint_rows or [],
    )
    seed_path = _write_csv(
        tmp_path / "data" / "processed" / "sources_seed_noord_brabant.csv",
        tuple(_seed_row("x", "x").keys()),
        seed_rows or [],
    )
    resolution_path = _write_csv(
        tmp_path / "tmp" / "generated" / "noord_brabant_missing_domain_external_resolution_v1.csv",
        tuple(_resolution_row("x", "x").keys()),
        resolution_rows or [],
    )
    return build_noord_brabant_makelaar_universe(
        funda_raw_csv_path=funda_path,
        source_completion_csv_path=completion_path,
        missing_domain_resolution_csv_path=resolution_path,
        source_master_csv_path=master_path,
        platform_fingerprint_csv_path=fingerprint_path,
        source_seed_csv_path=seed_path,
    )


def test_reads_raw_funda_makelaar_name_csv(tmp_path: Path) -> None:
    funda_path = _write_csv(
        tmp_path / "tmp" / "manual_inputs" / "funda_makelaar_names" / "funda_nb_makelaar_names_raw.csv",
        FUNDA_HEADER,
        [_funda_row("Alpha Makelaars"), _funda_row("Beta Wonen", city="Breda")],
    )

    rows = load_funda_makelaar_observations(funda_path)

    assert len(rows) == 2
    assert rows[0].makelaar_name_raw == "Alpha Makelaars"
    assert rows[1].city == "Breda"


def test_does_not_treat_funda_as_operational_property_source(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Alpha Makelaars")])

    assert result.quality_metrics["funda_operational_property_source_count"] == 0


def test_does_not_require_property_details(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Alpha Makelaars")])

    assert result.rows[0].display_name == "Alpha Makelaars"
    assert result.quality_metrics["funda_property_detail_extracted_count"] == 0


def test_normalizes_slogans(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row('Hoomz - "Smart Move."')])
    row = result.rows[0]

    assert row.display_name == "Hoomz"
    assert row.normalized_name == "hoomz"
    assert "Smart Move." in row.aliases


def test_rejects_funda_as_official_domain(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Quinten Vastgoed", city="Nuenen")],
        completion_rows=[_completion_row("funda.nl__oirschot", "Quinten Vastgoed", "funda.nl")],
    )
    row = result.rows[0]

    assert row.existing_domain == "funda.nl"
    assert row.official_domain == ""
    assert row.official_domain_status == "rejected_third_party_domain"


def test_rejects_funda_aanbod_url_as_official_operational_url(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Portal Makelaars")],
        completion_rows=[
            _completion_row(
                "portal.nl__tilburg",
                "Portal Makelaars",
                "portal.nl",
                accepted_aanbod_url="https://www.funda.nl/makelaars/tilburg/123-portal-makelaars/woningaanbod/koop",
                aanbod_url_status="valid",
            )
        ],
    )
    row = result.rows[0]

    assert row.official_aanbod_url == ""
    assert row.official_aanbod_url_status == "rejected_third_party_portal"


def test_egan_does_not_keep_funda_aanbod_as_valid(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Egan Makelaardij B.V.", city="Breda")],
        completion_rows=[_completion_row("egan.nl__breda", "Egan Makelaardij B.V.", "eganmakelaardij.nl")],
        master_rows=[
            _master_row(
                "egan.nl__breda",
                "Egan Makelaardij B.V.",
                "eganmakelaardij.nl",
                "https://www.funda.nl/makelaars/breda/19403-egan-makelaardij-bv/woningaanbod/koop/heel-nederland",
            )
        ],
        seed_rows=[
            _seed_row(
                "Egan Makelaardij B.V.",
                "eganmakelaardij.nl",
                "https://www.funda.nl/makelaars/breda/19403-egan-makelaardij-bv/woningaanbod/koop/heel-nederland",
            )
        ],
    )
    row = result.rows[0]

    assert row.official_domain == "eganmakelaardij.nl"
    assert row.official_aanbod_url == ""
    assert row.official_aanbod_url_status == "rejected_third_party_portal"
    assert row.recommended_next_action == "verify_public_aanbod_url"


def test_quinten_vastgoed_does_not_keep_funda_domain(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Quinten Vastgoed", city="Nuenen")],
        completion_rows=[
            _completion_row(
                "funda.nl__oirschot",
                "Quinten Vastgoed",
                "funda.nl",
                accepted_aanbod_url="https://www.funda.nl/makelaarsmatch",
                aanbod_url_status="suspect",
            )
        ],
        master_rows=[
            _master_row(
                "funda.nl__oirschot",
                "Quinten Vastgoed",
                "funda.nl",
                "https://www.funda.nl/makelaarsmatch",
            )
        ],
    )
    row = result.rows[0]

    assert row.official_domain == ""
    assert row.official_domain_status == "rejected_third_party_domain"
    assert row.official_aanbod_url == ""
    assert row.recommended_next_action == "legal_or_permission_review"


def test_cleans_via_paul_slogan(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row("Via Paul Makelaardij, via wie anders?!", city="Goirle")]).rows[0]

    assert row.display_name == "Via Paul Makelaardij"
    assert row.normalized_name == "via paul makelaardij"


def test_cleans_kin_funda_topmakelaar_slogan(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row("KIN Makelaars, Funda-topmakelaar van Brabant!", city="Oisterwijk")]).rows[0]

    assert row.display_name == "KIN Makelaars"
    assert row.normalized_name == "kin makelaars"


def test_cleans_allround_slogan(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row("Allround Makelaardij, beste makelaar in jouw regio", city="Goirle")]).rows[0]

    assert row.display_name == "Allround Makelaardij"
    assert row.normalized_name == "allround makelaardij"


def test_cleans_appels_qualis_label(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row("Appels Makelaardij B.V. | NVM-QUALIS", city="Goirle")]).rows[0]

    assert row.display_name == "Appels Makelaardij B.V."
    assert row.normalized_name == "appels makelaardij"


def test_cleans_lemmens_sales_slogan(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row("Lemmens Makelaardij , 200 verkopen met cijfer 9,8!", city="Tilburg")]).rows[0]

    assert row.display_name == "Lemmens Makelaardij"
    assert row.normalized_name == "lemmens makelaardij"


def test_cleans_hoomz_smart_move_slogan(tmp_path: Path) -> None:
    row = _build_result(tmp_path, [_funda_row('Hoomz - "Smart Move."')]).rows[0]

    assert row.display_name == "Hoomz"
    assert row.normalized_name == "hoomz"


def test_keeps_truncated_names_in_manual_review(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Nathalie van Gisbergen Lom...", name_truncated="true", name_needs_manual_review="true")])

    assert result.rows[0].priority_tier == "manual_review"
    assert "truncated_name_from_funda" in result.rows[0].manual_check_notes


def test_handles_meerdere_makelaars(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Meerdere makelaars", generic_multiple_makelaars="true")])
    row = result.rows[0]

    assert row.display_name == "Meerdere makelaars"
    assert row.priority_tier == "manual_review"
    assert row.official_domain == ""


def test_dedupes_strong_duplicate_names(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [
            _funda_row("Alpha Makelaars"),
            _funda_row("Alpha Makelaars", city="Breda", page_number=2, benchmark_count="1.159 koopwoningen in Breda"),
        ],
    )

    assert result.deduped_makelaars == 1
    assert result.rows[0].funda_seen_count == 2


def test_does_not_over_merge_weak_similar_names(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [
            _funda_row("Van Bruggen"),
            _funda_row("Van Brugg... / Van Bruggen", city="Breda", name_needs_manual_review="true"),
        ],
    )

    assert result.deduped_makelaars == 2


def test_matches_existing_master_by_name(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Alpha Makelaars")],
        master_rows=[_master_row("alpha.nl__tilburg", "Alpha Makelaars", "alpha.nl")],
    )
    row = result.rows[0]

    assert row.seen_in_existing_master == "yes"
    assert row.existing_source_id == "alpha.nl__tilburg"
    assert row.existing_domain == "alpha.nl"


def test_matches_existing_master_by_domain(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Alpha Makelaars")],
        completion_rows=[_completion_row("alpha.nl__tilburg", "Alpha Makelaars B.V.", "alpha.nl")],
        seed_rows=[_seed_row("Alpha Makelaars B.V.", "alpha.nl")],
    )
    row = result.rows[0]

    assert row.official_domain == "alpha.nl"
    assert row.seen_in_existing_master == "yes"


def test_kin_slogan_row_can_match_existing_brand_without_merging_unrelated_names(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("KIN Makelaars, Funda-topmakelaar van Brabant!", city="Oisterwijk")],
        completion_rows=[
            _completion_row(
                "kinmakelaars.nl__tilburg",
                "KIN Makelaars Tilburg B.V.",
                "kinmakelaars.nl",
                accepted_aanbod_url="https://kinmakelaars.nl/aanbod/wonen/te-koop",
                aanbod_url_status="accepted",
                delivery_mode="realworks",
                parser_family_candidate="realworks_public",
            )
        ],
        master_rows=[_master_row("kinmakelaars.nl__tilburg", "KIN Makelaars Tilburg B.V.", "kinmakelaars.nl", "https://kinmakelaars.nl/aanbod/wonen/te-koop")],
    )
    row = result.rows[0]

    assert row.display_name == "KIN Makelaars"
    assert row.official_domain == "kinmakelaars.nl"
    assert row.parser_family_candidate == "realworks_public"
    assert row.existing_source_id == "kinmakelaars.nl__tilburg"


def test_kin_brand_fallback_beats_unknown_domain_exact_name_match(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("KIN Makelaars, Funda-topmakelaar van Brabant!", city="Oisterwijk")],
        completion_rows=[
            _completion_row(
                "kinmakelaars.nl__tilburg",
                "KIN Makelaars Tilburg B.V.",
                "kinmakelaars.nl",
                accepted_aanbod_url="https://kinmakelaars.nl/aanbod/wonen/te-koop",
                aanbod_url_status="accepted",
                delivery_mode="realworks",
                parser_family_candidate="realworks_public",
            )
        ],
        resolution_rows=[_resolution_row("KIN Makelaars", "", resolved_source_id="unknown-domain__gilze-en-rijen")],
    )
    row = result.rows[0]

    assert row.existing_source_id == "kinmakelaars.nl__tilburg"
    assert row.official_domain == "kinmakelaars.nl"
    assert row.parser_family_candidate == "realworks_public"


def test_branch_variants_for_hendriks_and_van_de_water_remain_separate(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [
            _funda_row("Hendriks Makelaardij Breda", city="Breda"),
            _funda_row("Hendriks Makelaardij Tilburg", city="Tilburg"),
            _funda_row("Van de Water Makelaars Breda", city="Breda"),
            _funda_row("Van de Water Makelaars Tilburg", city="Tilburg"),
        ],
        master_rows=[
            _master_row("hendriks.nl__breda", "Hendriks Makelaardij Breda", "hendriks.nl"),
            _master_row("hendriks.nl__tilburg", "Hendriks Makelaardij Tilburg", "hendriks.nl"),
            _master_row("vandewater.nl__breda", "Van de Water Makelaars Breda", "vandewatergroep.nl"),
            _master_row("vandewater.nl__tilburg", "Van de Water Makelaars Tilburg", "vandewatergroep.nl"),
        ],
    )
    rows_by_name = {row.display_name: row for row in result.rows}

    assert result.deduped_makelaars == 4
    assert rows_by_name["Hendriks Makelaardij Breda"].existing_source_id == "hendriks.nl__breda"
    assert rows_by_name["Hendriks Makelaardij Tilburg"].existing_source_id == "hendriks.nl__tilburg"
    assert rows_by_name["Van de Water Makelaars Breda"].existing_source_id == "vandewater.nl__breda"
    assert rows_by_name["Van de Water Makelaars Tilburg"].existing_source_id == "vandewater.nl__tilburg"


def test_copies_known_accepted_aanbod_url_from_source_completion_when_available(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Alpha Makelaars")],
        completion_rows=[
            _completion_row(
                "alpha.nl__tilburg",
                "Alpha Makelaars",
                "alpha.nl",
                accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop",
                aanbod_url_status="valid",
                delivery_mode="realworks_public",
                parser_family_candidate="realworks_public",
            )
        ],
    )
    row = result.rows[0]

    assert row.official_aanbod_url == "https://alpha.nl/aanbod/woningaanbod/tilburg/koop"
    assert row.official_aanbod_url_status == "valid"


def test_keeps_official_domain_unknown_explicit(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Unknown Makelaars")])

    assert result.rows[0].official_domain_status == "unknown_explicit"


def test_computes_priority_tier(tmp_path: Path) -> None:
    result = _build_result(
        tmp_path,
        [_funda_row("Alpha Makelaars", city="Tilburg"), _funda_row("Alpha Makelaars", city="Eindhoven", page_number=2, benchmark_count="1.511 koopwoningen in Eindhoven")],
        completion_rows=[
            _completion_row(
                "alpha.nl__tilburg",
                "Alpha Makelaars",
                "alpha.nl",
                accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop",
                aanbod_url_status="valid",
                delivery_mode="realworks_public",
                parser_family_candidate="realworks_public",
            )
        ],
    )

    assert result.rows[0].priority_tier == "P0_city_critical"


def test_writes_workbook_sheets(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Alpha Makelaars")])
    workbook_path = write_makelaar_universe_workbook(result, tmp_path / "tmp" / "generated" / "universe.xlsx")

    workbook = load_workbook(workbook_path)
    try:
        assert workbook.sheetnames == [
            "Makelaar Universe",
            "Review Queue",
            "Per City Counts",
            "Quality Gates",
            "Run Summary",
        ]
    finally:
        workbook.close()


def test_does_not_commit_manual_inputs_or_generated_artifacts(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Alpha Makelaars")])
    csv_path = write_makelaar_universe_csv(result.rows, tmp_path / "tmp" / "generated" / "universe.csv")
    review_path = write_makelaar_universe_csv(result.review_queue_rows, tmp_path / "tmp" / "generated" / "review.csv")

    assert "tmp/generated" in str(csv_path).replace("\\", "/")
    assert "tmp/generated" in str(review_path).replace("\\", "/")
    assert "data/raw" not in str(csv_path).replace("\\", "/")


def test_runner_uses_merged_snapshot_when_present(tmp_path: Path) -> None:
    raw_dir = tmp_path / "tmp" / "manual_inputs" / "funda_makelaar_names"
    _write_csv(raw_dir / "funda_nb_makelaar_names_raw.csv", FUNDA_HEADER, [_funda_row("Alpha Makelaars")])
    _write_csv(raw_dir / "funda_nb_makelaar_names_raw_merged_snapshot.csv", FUNDA_HEADER, [_funda_row("Merged Makelaars")])
    _write_csv(tmp_path / "tmp" / "generated" / "noord_brabant_source_completion_scope_verification_v2.csv", tuple(_completion_row("x", "x", "x").keys()), [])
    _write_csv(tmp_path / "tmp" / "generated" / "noord_brabant_missing_domain_external_resolution_v1.csv", tuple(_resolution_row("x", "x").keys()), [])
    _write_csv(tmp_path / "data" / "discovery" / "runs" / "20260614T122022Z" / "makelaar_sources_master.csv", tuple(_master_row("x", "x", "x").keys()), [])
    _write_csv(tmp_path / "data" / "discovery" / "platform_fingerprint" / "platform_fingerprint_results.csv", tuple(_fingerprint_row("x", "x", "x", "realworks").keys()), [])
    _write_csv(tmp_path / "data" / "processed" / "sources_seed_noord_brabant.csv", tuple(_seed_row("x", "x").keys()), [])

    result = build_noord_brabant_makelaar_universe(
        funda_raw_csv_path=raw_dir / "funda_nb_makelaar_names_raw.csv",
        source_completion_csv_path=tmp_path / "tmp" / "generated" / "noord_brabant_source_completion_scope_verification_v2.csv",
        missing_domain_resolution_csv_path=tmp_path / "tmp" / "generated" / "noord_brabant_missing_domain_external_resolution_v1.csv",
        source_master_csv_path=tmp_path / "data" / "discovery" / "runs" / "20260614T122022Z" / "makelaar_sources_master.csv",
        platform_fingerprint_csv_path=tmp_path / "data" / "discovery" / "platform_fingerprint" / "platform_fingerprint_results.csv",
        source_seed_csv_path=tmp_path / "data" / "processed" / "sources_seed_noord_brabant.csv",
    )

    assert result.rows[0].display_name == "Merged Makelaars"
    assert result.effective_input_path.endswith("funda_nb_makelaar_names_raw_merged_snapshot.csv")


def test_cli_writes_outputs(tmp_path: Path) -> None:
    raw_dir = tmp_path / "tmp" / "manual_inputs" / "funda_makelaar_names"
    gen_dir = tmp_path / "tmp" / "generated"
    data_dir = tmp_path / "data"
    _write_csv(raw_dir / "funda_nb_makelaar_names_raw.csv", FUNDA_HEADER, [_funda_row("Alpha Makelaars")])
    _write_csv(gen_dir / "noord_brabant_source_completion_scope_verification_v2.csv", tuple(_completion_row("x", "x", "x").keys()), [])
    _write_csv(gen_dir / "noord_brabant_missing_domain_external_resolution_v1.csv", tuple(_resolution_row("x", "x").keys()), [])
    _write_csv(data_dir / "discovery" / "runs" / "20260614T122022Z" / "makelaar_sources_master.csv", tuple(_master_row("x", "x", "x").keys()), [])
    _write_csv(data_dir / "discovery" / "platform_fingerprint" / "platform_fingerprint_results.csv", tuple(_fingerprint_row("x", "x", "x", "realworks").keys()), [])
    _write_csv(data_dir / "processed" / "sources_seed_noord_brabant.csv", tuple(_seed_row("x", "x").keys()), [])

    output_xlsx = gen_dir / "noord_brabant_makelaar_universe_v1.xlsx"
    output_csv = gen_dir / "noord_brabant_makelaar_universe_v1.csv"
    output_review = gen_dir / "noord_brabant_makelaar_universe_v1_review_queue.csv"
    command = [
        sys.executable,
        "scripts/run_noord_brabant_makelaar_universe.py",
        "--funda-raw-csv",
        str(raw_dir / "funda_nb_makelaar_names_raw.csv"),
        "--source-completion-csv",
        str(gen_dir / "noord_brabant_source_completion_scope_verification_v2.csv"),
        "--missing-domain-resolution-csv",
        str(gen_dir / "noord_brabant_missing_domain_external_resolution_v1.csv"),
        "--source-master-csv",
        str(data_dir / "discovery" / "runs" / "20260614T122022Z" / "makelaar_sources_master.csv"),
        "--platform-fingerprint-csv",
        str(data_dir / "discovery" / "platform_fingerprint" / "platform_fingerprint_results.csv"),
        "--source-seed-csv",
        str(data_dir / "processed" / "sources_seed_noord_brabant.csv"),
        "--output-workbook",
        str(output_xlsx),
        "--output-csv",
        str(output_csv),
        "--output-review-queue-csv",
        str(output_review),
    ]

    result = subprocess.run(command, cwd=Path(__file__).resolve().parents[1], capture_output=True, text=True, check=False)

    assert result.returncode == 0
    assert output_xlsx.exists()
    assert output_csv.exists()
    assert output_review.exists()
    assert "raw_rows_imported: 1" in result.stdout


def test_quality_gates_stay_zero_for_valid_rows(tmp_path: Path) -> None:
    result = _build_result(tmp_path, [_funda_row("Alpha Makelaars")])

    assert compute_quality_metrics(result.rows) == {
        "makelaar_universe_rows_without_name_count": 0,
        "duplicate_high_confidence_makelaar_count": 0,
        "funda_property_detail_extracted_count": 0,
        "funda_operational_property_source_count": 0,
        "raw_html_json_persisted_count": 0,
        "long_descriptions_exported_count": 0,
        "browser_automation_used_count": 0,
    }
