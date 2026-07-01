from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from domek_wonen.sources.coverage_census import (
    AanbodUrlCandidate,
    CoverageCensusResult,
    CoverageSourceRecord,
    CoverageTerminalStatus,
    DomainResolutionRecord,
    MissingDomainQueueRecord,
    build_aanbod_scope_checks,
    build_domain_resolution_records,
    build_no_public_verification_records,
    build_office_location_verification_records,
    build_realworks_audit_input_records,
    classify_family_from_content,
    compute_quality_metrics,
    consolidate_coverage_source_seeds,
    derive_listing_index_candidates_from_detail_url,
    discover_aanbod_url,
    normalize_gemeente_name,
    finalize_coverage_record,
    run_investigation_loop,
    run_noord_brabant_coverage_source_census,
    write_realworks_audit_input_csv,
    write_coverage_census_outputs,
    _set_accepted_aanbod_url,
)


def _row(**overrides: str) -> dict[str, str]:
    base = {
        "source_id": "alpha.nl__tilburg",
        "office_name": "Alpha Makelaars",
        "root_domain": "alpha.nl",
        "website": "https://www.alpha.nl",
        "gemeente": "Tilburg",
        "province": "Noord-Brabant",
        "aanbod_url": "https://www.alpha.nl/aanbod/woningaanbod",
        "aanbod_url_quality": "valid",
        "legal_status": "allowed_official_source",
        "evidence_file": "fixture.csv",
    }
    base.update(overrides)
    return base


def _record(**overrides: object) -> CoverageSourceRecord:
    base = {
        "source_id": "alpha.nl__tilburg",
        "source_name": "Alpha Makelaars",
        "domain": "alpha.nl",
        "root_url": "https://alpha.nl",
        "coverage_gemeente": "Tilburg",
        "coverage_province": "Noord-Brabant",
        "has_noord_brabant_coverage": True,
        "access_policy_status": "allowed",
    }
    base.update(overrides)
    return CoverageSourceRecord(**base)


def _fetcher(pages: dict[str, str]):
    def fetch(url: str) -> str:
        return pages[url.rstrip("/")]

    return fetch


def test_consolidates_local_source_seeds() -> None:
    records, duplicates = consolidate_coverage_source_seeds([_row(), _row(root_domain="beta.nl", source_id="beta.nl__breda")])
    assert len(records) == 2
    assert duplicates == []


def test_dedupes_by_normalized_domain() -> None:
    records, duplicates = consolidate_coverage_source_seeds([_row(root_domain="www.alpha.nl"), _row(source_id="alias", root_domain="alpha.nl")])
    assert len(records) == 1
    assert len(duplicates) == 1


def test_preserves_source_aliases() -> None:
    records, _duplicates = consolidate_coverage_source_seeds([_row(office_name="Alpha"), _row(source_id="alias", office_name="Alpha Alias")])
    assert "alias" in records[0].source_ids
    assert "Alpha Alias" in records[0].aliases


def test_includes_outside_office_source_if_it_has_noord_brabant_coverage() -> None:
    records, _ = consolidate_coverage_source_seeds([
        _row(office_province="Zuid-Holland", province="Noord-Brabant", gemeente="Tilburg")
    ])
    assert records[0].has_noord_brabant_coverage is True


def test_separates_office_location_from_coverage_location() -> None:
    records, _ = consolidate_coverage_source_seeds([
        _row(office_gemeente="Rotterdam", gemeente="Tilburg", office_province="Zuid-Holland", province="Noord-Brabant")
    ])
    assert records[0].office_gemeente == "Rotterdam"
    assert records[0].coverage_gemeente == "Tilburg"


def test_accepts_explicit_valid_aanbod_url_from_local_evidence() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen")
    assert discover_aanbod_url(record, fetch_text=None)
    assert record.aanbod_url_status == "accepted"


def test_rejects_property_detail_url_as_aanbod_url() -> None:
    record = _record(aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop/huis-123-main")
    assert not discover_aanbod_url(record, fetch_text=None)
    assert record.aanbod_url_candidates[0].rejection_reason == "property_detail_url_rejected"


def test_rejects_funda_pararius_as_operational_aanbod_url() -> None:
    record = _record(domain="alpha.nl", aanbod_url="https://www.funda.nl/koop/tilburg/")
    assert not discover_aanbod_url(record, fetch_text=None)
    assert record.aanbod_url_candidates[0].rejection_reason == "funda_pararius_not_operational"


def test_discovers_aanbod_url_from_homepage_link() -> None:
    record = _record(aanbod_url="")
    fetch = _fetcher({
        "https://alpha.nl": '<a href="/woningen">Koopwoningen</a>',
        "https://alpha.nl/woningen": '<div class="property-card">Vraagprijs</div>',
    })
    assert discover_aanbod_url(record, fetch_text=fetch, can_fetch=lambda _d, _p: True)
    assert record.aanbod_url == "https://alpha.nl/woningen"


def test_discovers_aanbod_url_from_sitemap() -> None:
    record = _record(aanbod_url="")
    fetch = _fetcher({
        "https://alpha.nl": "<html></html>",
        "https://alpha.nl/sitemap.xml": "<loc>https://alpha.nl/aanbod/woningaanbod</loc>",
    })
    assert discover_aanbod_url(record, fetch_text=fetch, can_fetch=lambda _d, _p: True)
    assert record.aanbod_url == "https://alpha.nl/aanbod/woningaanbod"


def test_discovers_aanbod_url_from_common_paths() -> None:
    record = _record(aanbod_url="")
    fetch = _fetcher({
        "https://alpha.nl": "<html></html>",
        "https://alpha.nl/sitemap.xml": "<urlset></urlset>",
        "https://alpha.nl/aanbod": '<article class="listing-card">k.k.</article>',
    })
    assert discover_aanbod_url(record, fetch_text=fetch, can_fetch=lambda _d, _p: True)
    assert record.aanbod_url == "https://alpha.nl/aanbod"


def test_marks_no_public_aanbod_after_exhausted_attempts() -> None:
    record = _record(aanbod_url="")
    run_investigation_loop(record, fetch_text=None)
    assert record.parser_family_candidate == "no_public_aanbod"
    assert record.terminal_status == CoverageTerminalStatus.CONFIRMED_NO_PUBLIC_AANBOD.value


def test_classifies_realworks_from_realworks_signals() -> None:
    family, _delivery, _confidence, signal = classify_family_from_content(content='<li class="aanbodEntry"></li>', url="/aanbod/woningaanbod")
    assert family == "realworks_public"
    assert signal == "realworks_signal"


def test_classifies_ogonline_from_xhr_api_signals() -> None:
    family, _delivery, _confidence, _signal = classify_family_from_content(content='{"docs":[],"totalDocs":0}')
    assert family == "ogonline_xhr"


def test_classifies_wordpress_from_wp_json_wp_content() -> None:
    assert classify_family_from_content(content="/wp-json/wp/v2")[0] == "wordpress_json"
    assert classify_family_from_content(content="/wp-content/themes/site")[0] == "wordpress_static"


def test_classifies_custom_html_from_server_rendered_listing_cards() -> None:
    assert classify_family_from_content(content='<div class="property-card">Vraagprijs k.k.</div>')[0] == "custom_html"


def test_classifies_custom_xhr_from_json_listing_payload_signals() -> None:
    assert classify_family_from_content(content="fetch('/api/listings')")[0] == "custom_xhr"


def test_classifies_custom_js_app_when_static_html_has_app_shell_but_no_listing_payload() -> None:
    assert classify_family_from_content(content='<div id="app"></div><script src="/main.js"></script>')[0] == "custom_js_app"


def test_marks_blocked_or_legal_review_when_access_policy_blocks() -> None:
    record = _record(access_policy_status="blocked")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "blocked_or_legal_review"


def test_does_not_leave_final_parser_family_candidate_unknown() -> None:
    record = _record(aanbod_url="")
    run_investigation_loop(record)
    assert record.parser_family_candidate not in {"", "unknown", "missing", "tbd", "todo"}


def test_does_not_leave_missing_aanbod_url_without_terminal_reason() -> None:
    record = _record(aanbod_url="")
    run_investigation_loop(record)
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["missing_aanbod_url_without_terminal_reason_count"] == 0


def test_writes_workbook_with_all_required_sheets(tmp_path: Path) -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="realworks_public", delivery_mode="realworks_public")
    run_investigation_loop(record)
    write_coverage_census_outputs(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1), tmp_path / "census.xlsx", tmp_path / "master.csv", tmp_path / "review.csv")
    workbook = load_workbook(tmp_path / "census.xlsx")
    try:
        assert workbook.sheetnames == [
            "Master Sources",
            "Domain Resolution",
            "No Public Verification",
            "Aanbod Scope Check",
            "Realworks Audit Input",
            "Aanbod URL Evidence",
            "Family Fingerprints",
            "Investigation Attempts",
            "Coverage Matrix",
            "Realworks Candidates",
            "Realworks Verification",
            "OGonline Candidates",
            "Custom Needs Parser",
            "Custom JS Refingerprint",
            "Family Conflicts",
            "Office Location Verification",
            "Blocked or Legal Review",
            "Duplicates",
            "Missing Domain Queue",
            "Normalization Issues",
            "Review Queue",
            "Quality Gates",
        ]
    finally:
        workbook.close()


def test_writes_review_queue(tmp_path: Path) -> None:
    record = _record(access_policy_status="blocked")
    run_investigation_loop(record)
    write_coverage_census_outputs(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1), tmp_path / "census.xlsx", tmp_path / "master.csv", tmp_path / "review.csv")
    with (tmp_path / "review.csv").open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["domain"] == "alpha.nl"
    assert rows[0]["terminal_status"] == CoverageTerminalStatus.CONFIRMED_BLOCKED_OR_LEGAL_REVIEW.value


def test_does_not_persist_raw_html_json(tmp_path: Path) -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_html", family_evidence="<html><script>raw</script></html>")
    run_investigation_loop(record)
    _wb, master, review = write_coverage_census_outputs(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1), tmp_path / "census.xlsx", tmp_path / "master.csv", tmp_path / "review.csv")
    text = master.read_text(encoding="utf-8") + review.read_text(encoding="utf-8")
    assert "<html" not in text
    assert "<script" not in text


def test_does_not_include_long_descriptions(tmp_path: Path) -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_html", coverage_evidence="x" * 900)
    run_investigation_loop(record)
    _wb, master, _review = write_coverage_census_outputs(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1), tmp_path / "census.xlsx", tmp_path / "master.csv", tmp_path / "review.csv")
    assert "x" * 600 not in master.read_text(encoding="utf-8")


def test_does_not_create_parser_per_makelaar() -> None:
    family, *_ = classify_family_from_content(content="Alpha Makelaars bespoke page")
    assert family in {"custom_js_app", "custom_html", "custom_xhr"}
    assert "alpha" not in family


def test_does_not_import_matching() -> None:
    import domek_wonen.sources.coverage_census as coverage_census

    assert "domek_wonen.matching" not in coverage_census.__dict__


def test_does_not_touch_funda_pararius_operationally() -> None:
    record = _record(aanbod_url="https://pararius.nl/koopwoningen/tilburg")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "blocked_or_legal_review"


def test_quality_gate_fails_if_unknown_operational_family_remains() -> None:
    record = _record(parser_family_candidate="unknown", terminal_status="confirmed_source_ready")
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["operational_unknown_family_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_if_missing_aanbod_url_lacks_terminal_reason() -> None:
    record = _record(aanbod_url="", parser_family_candidate="custom_html", terminal_status="confirmed_source_ready")
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["missing_aanbod_url_without_terminal_reason_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_runner_writes_outputs_from_local_csvs(tmp_path: Path) -> None:
    data_dir = tmp_path / "data" / "processed"
    data_dir.mkdir(parents=True)
    seed = data_dir / "sources_seed_noord_brabant.csv"
    seed.write_text(
        "source_id,office_name,root_domain,website,gemeente,province,aanbod_url,aanbod_url_quality,legal_status\n"
        "alpha.nl__tilburg,Alpha,alpha.nl,https://alpha.nl,Tilburg,Noord-Brabant,https://alpha.nl/woningen,valid,allowed_official_source\n",
        encoding="utf-8",
    )
    result = run_noord_brabant_coverage_source_census(
        repo_root=tmp_path,
        evidence_paths=(Path("data/processed/sources_seed_noord_brabant.csv"),),
        output_dir=tmp_path / "tmp" / "generated",
    )
    assert result.workbook_path and result.workbook_path.exists()
    assert result.master_csv_path and result.master_csv_path.exists()
    assert result.review_queue_csv_path and result.review_queue_csv_path.exists()


def test_rejected_candidate_cannot_appear_as_accepted_aanbod_url() -> None:
    record = _record(aanbod_url="https://other.nl/woningen")
    assert not discover_aanbod_url(record, fetch_text=None)
    assert record.accepted_aanbod_url == ""
    assert record.best_rejected_candidate_reason == "candidate_not_on_official_domain"


def test_later_acceptance_reconciles_prior_rejected_duplicate_candidate() -> None:
    record = _record()
    record.aanbod_url_candidates.append(
        AanbodUrlCandidate(
            source_id=record.source_id,
            domain=record.domain,
            candidate_url="https://alpha.nl/aanbod",
            accepted=False,
            rejection_reason="page_does_not_look_like_listing_index",
            evidence_type="homepage_link",
            pass_name="pass_2_homepage_links",
            confidence=0.45,
        )
    )
    _set_accepted_aanbod_url(record, "https://alpha.nl/aanbod", "common_path", "listing cards", 0.62)
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert record.aanbod_url_candidates[0].accepted is True
    assert record.aanbod_url_candidates[0].rejection_reason == "reconciled_after_later_acceptance"
    assert metrics["rejected_candidate_used_as_master_aanbod_url_count"] == 0


def test_property_detail_url_cannot_appear_as_accepted_aanbod_url() -> None:
    record = _record(aanbod_url="https://alpha.nl/koop/huis-123-straat-1")
    discover_aanbod_url(record, fetch_text=None)
    assert record.accepted_aanbod_url == ""


def test_realworks_detail_url_derives_listing_index_candidates() -> None:
    candidates = derive_listing_index_candidates_from_detail_url(
        "https://alpha.nl/aanbod/woningaanbod/tilburg/koop/huis-123-straat-1"
    )
    assert "https://alpha.nl/aanbod/woningaanbod/tilburg/koop" in candidates
    assert "https://alpha.nl/aanbod/woningaanbod/koop" in candidates
    assert "https://alpha.nl/aanbod/woningaanbod" in candidates


def test_derived_realworks_listing_index_can_become_accepted_aanbod_url() -> None:
    record = _record(aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop/huis-123-straat-1")
    fetch = _fetcher({
        "https://alpha.nl/aanbod/woningaanbod/tilburg/koop": '<li class="aanbodEntry">Vraagprijs</li>',
    })
    assert discover_aanbod_url(record, fetch_text=fetch, can_fetch=lambda _d, _p: True)
    assert record.accepted_aanbod_url == "https://alpha.nl/aanbod/woningaanbod/tilburg/koop"


def test_platform_guess_realworks_is_not_enough_to_classify_realworks() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", platform_guess="realworks")
    run_investigation_loop(record)
    assert record.parser_family_candidate != "realworks_public"


def test_realworks_requires_strong_evidence() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", platform_guess="realworks", parser_family_candidate="realworks_public")
    run_investigation_loop(record)
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert record.parser_family_candidate != "realworks_public"
    assert metrics["realworks_without_strong_evidence_count"] == 0


def test_kin_is_not_classified_as_realworks_from_stale_local_evidence() -> None:
    record = _record(
        source_id="kinmakelaars-nl__breda",
        domain="kinmakelaars.nl",
        aanbod_url="https://www.kinmakelaars.nl/aanbod/wonen/te-koop",
        platform_guess="realworks",
        parser_family_candidate="realworks_public",
    )
    run_investigation_loop(record)
    assert record.parser_family_candidate == "ogonline_xhr"


def test_kin_resolves_to_ogonline_or_explicit_review_status() -> None:
    record = _record(source_id="kinmakelaars-nl__breda", domain="kinmakelaars.nl", aanbod_url="https://www.kinmakelaars.nl/aanbod/wonen/te-koop")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "ogonline_xhr"
    assert record.validation_status in {"needs_ogonline_revalidation", "passed_quality_gates"}


def test_family_conflicts_are_reported() -> None:
    record = _record(source_id="kinmakelaars-nl__breda", domain="kinmakelaars.nl", aanbod_url="https://www.kinmakelaars.nl/aanbod/wonen/te-koop", parser_family_candidate="realworks_public")
    run_investigation_loop(record)
    assert record.family_conflicts


def test_custom_js_app_sources_are_refingerprinted() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_js_app")
    run_investigation_loop(record)
    assert record.custom_js_refingerprint_attempts


def test_custom_js_app_with_wp_json_moves_to_wordpress_json() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_js_app", family_evidence="wp-json")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "wordpress_json"


def test_custom_js_app_with_server_rendered_cards_moves_to_custom_html() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_js_app", family_evidence="property-card Vraagprijs")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "custom_html"


def test_custom_js_app_with_realworks_strong_signals_moves_to_realworks_public() -> None:
    record = _record(aanbod_url="https://alpha.nl/aanbod/woningaanbod", parser_family_candidate="custom_js_app", family_evidence="aanbodEntry")
    run_investigation_loop(record)
    assert record.parser_family_candidate == "realworks_public"


def test_custom_js_app_with_only_app_shell_remains_custom_js_app_with_evidence() -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_js_app", family_evidence='<div id="app"></div>')
    run_investigation_loop(record)
    assert record.parser_family_candidate == "custom_js_app"
    assert record.custom_js_refingerprint_attempts[0].decision == "confirmed_custom_js_app"


def test_gemeente_names_are_normalized() -> None:
    assert normalize_gemeente_name("Bergen Op Zoom") == "Bergen op Zoom"
    assert normalize_gemeente_name("S Hertogenbosch") == "'s-Hertogenbosch"


def test_missing_domain_rows_go_to_missing_domain_queue(tmp_path: Path) -> None:
    seed = tmp_path / "seed.csv"
    seed.write_text("source_id,office_name,gemeente,province\nmissing,No Domain,Tilburg,Noord-Brabant\n", encoding="utf-8")
    result = run_noord_brabant_coverage_source_census(repo_root=tmp_path, evidence_paths=(Path("seed.csv"),))
    assert len(result.records) == 0
    assert len(result.missing_domain_queue) == 1


def test_office_location_is_not_fabricated_as_noord_brabant_when_missing() -> None:
    records, _ = consolidate_coverage_source_seeds([_row(office_province="", office_gemeente="")])
    assert records[0].office_province == ""
    assert records[0].office_location_status != "known"


def test_outside_office_coverage_fields_are_separate_from_office_fields() -> None:
    records, _ = consolidate_coverage_source_seeds([_row(office_province="Zuid-Holland", office_gemeente="Rotterdam")])
    assert records[0].office_province == "Zuid-Holland"
    assert records[0].coverage_province == "Noord-Brabant"
    assert records[0].outside_office_coverage_status == "outside_office_source_included"


def test_quality_gate_fails_when_rejected_candidate_is_used_in_master() -> None:
    record = _record()
    record.accepted_aanbod_url = "https://other.nl/woningen"
    record.aanbod_url_candidates.append(type("C", (), {"accepted": False, "candidate_url": "https://other.nl/woningen"})())
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["rejected_candidate_used_as_master_aanbod_url_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_when_property_detail_url_remains_accepted() -> None:
    record = _record()
    record.accepted_aanbod_url = "https://alpha.nl/koop/huis-123-straat-1"
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["property_detail_url_as_aanbod_url_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_when_realworks_lacks_strong_evidence() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/woningen", parser_family_candidate="realworks_public")
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["realworks_without_strong_evidence_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_when_kin_remains_realworks() -> None:
    record = _record(source_id="kinmakelaars-nl__breda", domain="kinmakelaars.nl", accepted_aanbod_url="https://kinmakelaars.nl/aanbod/wonen/te-koop", parser_family_candidate="realworks_public")
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["kin_family_conflict_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_when_custom_js_app_has_no_fingerprint_attempt() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/woningen", parser_family_candidate="custom_js_app")
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1))
    assert metrics["custom_js_app_without_fingerprint_attempt_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_quality_gate_fails_on_gemeente_normalization_conflicts() -> None:
    record = _record()
    result = CoverageCensusResult(records=(record,), normalization_issues=(), initial_source_count=1, deduped_source_count=1)
    metrics = compute_quality_metrics(result)
    assert metrics["gemeente_normalization_conflict_count"] == 0


def test_hardened_workbook_uses_accepted_aanbod_url_not_ambiguous_aanbod_url(tmp_path: Path) -> None:
    record = _record(aanbod_url="https://alpha.nl/woningen")
    run_investigation_loop(record)
    _wb, master, _review = write_coverage_census_outputs(CoverageCensusResult(records=(record,), initial_source_count=1, deduped_source_count=1), tmp_path / "census.xlsx", tmp_path / "master.csv", tmp_path / "review.csv")
    columns = master.read_text(encoding="utf-8").splitlines()[0].split(",")
    assert "accepted_aanbod_url" in columns
    assert "aanbod_url" not in columns


def test_missing_domain_rows_get_resolution_attempts() -> None:
    missing = MissingDomainQueueRecord("Alpha Makelaars", "raw-1", "Tilburg", "Noord-Brabant", "seed.csv", "missing_official_domain", "Alpha Tilburg makelaar", "manual")
    rows = build_domain_resolution_records([missing], [_record(source_name="Alpha Makelaars", domain="alpha.nl")])
    assert rows[0].attempt_count == 1
    assert rows[0].resolution_status == "resolved_to_existing_source"
    assert rows[0].resolved_domain == "alpha.nl"


def test_missing_domain_can_resolve_to_new_source_from_official_lookup() -> None:
    missing = MissingDomainQueueRecord("Beta Wonen", "raw-2", "Breda", "Noord-Brabant", "seed.csv", "missing_official_domain", "Beta Breda makelaar", "manual")
    rows = build_domain_resolution_records([missing], [], official_domain_lookup={"Beta Wonen": "beta.nl"})
    assert rows[0].resolution_status == "resolved_to_new_source"
    assert rows[0].resolved_domain == "beta.nl"


def test_missing_domain_unresolved_row_has_suggested_next_action() -> None:
    missing = MissingDomainQueueRecord("Unknown Office", "raw-3", "Breda", "Noord-Brabant", "seed.csv", "missing_official_domain", "Unknown Office Breda makelaar", "manual")
    rows = build_domain_resolution_records([missing], [])
    assert rows[0].resolution_status == "needs_manual_domain_research"
    assert rows[0].suggested_next_action


def test_no_public_source_requires_full_attempt_history() -> None:
    record = _record(aanbod_url="")
    run_investigation_loop(record, fetch_text=None)
    rows = build_no_public_verification_records([record])
    assert rows[0].homepage_checked == "yes"
    assert rows[0].sitemap_checked == "yes"
    assert rows[0].common_paths_checked == "yes"
    assert rows[0].reason == "full_attempt_history"


def test_no_public_source_is_reclassified_when_official_aanbod_is_found() -> None:
    record = _record(parser_family_candidate="no_public_aanbod", accepted_aanbod_url="https://alpha.nl/aanbod")
    rows = build_no_public_verification_records([record])
    assert rows[0].final_status == "official_aanbod_found"


def test_accepted_aanbod_scope_detects_broad_official_index() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/aanbod")
    rows = build_aanbod_scope_checks([record])
    assert rows[0].aanbod_scope_status == "broad_official_index"


def test_accepted_aanbod_scope_detects_matching_nb_place() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop")
    rows = build_aanbod_scope_checks([record])
    assert rows[0].url_place_token == "tilburg"
    assert rows[0].aanbod_scope_status == "confirmed_nb_scope"


def test_accepted_aanbod_scope_flags_outside_place_for_review() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/amsterdam/koop")
    rows = build_aanbod_scope_checks([record])
    assert rows[0].aanbod_scope_status == "needs_scope_review"
    assert rows[0].reason


def test_out_of_scope_looking_url_cannot_be_silently_ready_for_audit() -> None:
    record = _record(
        accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/amsterdam/koop",
        parser_family_candidate="realworks_public",
        family_evidence="aanbodEntry",
    )
    rows = build_realworks_audit_input_records([record], build_aanbod_scope_checks([record]))
    assert rows[0].audit_input_status == "needs_manual_scope_check"


def test_realworks_audit_input_excludes_kin() -> None:
    record = _record(
        source_id="kinmakelaars-nl__breda",
        domain="kinmakelaars.nl",
        accepted_aanbod_url="https://kinmakelaars.nl/aanbod/woningaanbod/breda/koop",
        parser_family_candidate="realworks_public",
        family_evidence="aanbodEntry",
    )
    rows = build_realworks_audit_input_records([record], build_aanbod_scope_checks([record]))
    assert rows[0].audit_input_status == "exclude_not_realworks"


def test_realworks_audit_input_excludes_missing_accepted_aanbod_url() -> None:
    record = _record(parser_family_candidate="realworks_public", family_evidence="aanbodEntry")
    rows = build_realworks_audit_input_records([record])
    assert rows[0].audit_input_status == "exclude_no_public_aanbod"


def test_realworks_audit_input_excludes_property_detail_url() -> None:
    record = _record(
        accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop/huis-123-main",
        parser_family_candidate="realworks_public",
        family_evidence="aanbodEntry",
    )
    rows = build_realworks_audit_input_records([record], build_aanbod_scope_checks([record]))
    assert rows[0].audit_input_status == "exclude_not_realworks"


def test_realworks_audit_input_requires_strong_realworks_evidence() -> None:
    record = _record(accepted_aanbod_url="https://alpha.nl/woningen", parser_family_candidate="realworks_public")
    rows = build_realworks_audit_input_records([record], build_aanbod_scope_checks([record]))
    assert rows[0].audit_input_status == "exclude_not_realworks"


def test_realworks_audit_input_csv_contains_only_ready_records(tmp_path: Path) -> None:
    ready = _record(
        source_id="alpha.nl__tilburg",
        accepted_aanbod_url="https://alpha.nl/aanbod/woningaanbod/tilburg/koop",
        parser_family_candidate="realworks_public",
        family_evidence="aanbodEntry",
    )
    excluded = _record(source_id="beta.nl__tilburg", domain="beta.nl", parser_family_candidate="realworks_public")
    records = build_realworks_audit_input_records([ready, excluded], build_aanbod_scope_checks([ready, excluded]))
    output = write_realworks_audit_input_csv(records, tmp_path / "audit.csv")
    with output.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert [row["source_id"] for row in rows] == ["alpha.nl__tilburg"]


def test_office_location_can_be_confirmed_from_compact_evidence() -> None:
    record = _record(office_gemeente="Tilburg", office_province="Noord-Brabant")
    rows = build_office_location_verification_records([record])
    assert rows[0].office_location_status == "office_confirmed_noord_brabant"


def test_office_location_is_not_fabricated_gate() -> None:
    record = _record()
    rows = build_office_location_verification_records([record])
    assert rows[0].office_location_status == "office_unknown"
    metrics = compute_quality_metrics(CoverageCensusResult(records=(record,), office_location_verification=tuple(rows), initial_source_count=1, deduped_source_count=1))
    assert metrics["office_location_fabricated_count"] == 0


def test_new_quality_gate_fails_when_missing_domain_has_no_attempt() -> None:
    row = DomainResolutionRecord("Alpha", "raw", "Tilburg", "Noord-Brabant", "seed.csv", "needs_manual_domain_research", attempt_count=0)
    metrics = compute_quality_metrics(CoverageCensusResult(records=(), domain_resolution=(row,), initial_source_count=1, deduped_source_count=0))
    assert metrics["missing_domain_without_resolution_attempt_count"] == 1
    assert metrics["quality_gate_passed"] is False


def test_completion_runner_writes_expected_artifacts(tmp_path: Path) -> None:
    data_dir = tmp_path / "data" / "processed"
    data_dir.mkdir(parents=True)
    seed = data_dir / "sources_seed_noord_brabant.csv"
    seed.write_text(
        "source_id,office_name,root_domain,website,gemeente,province,aanbod_url,aanbod_url_quality,legal_status,parser_family_candidate,evidence\n"
        "alpha.nl__tilburg,Alpha,alpha.nl,https://alpha.nl,Tilburg,Noord-Brabant,https://alpha.nl/aanbod/woningaanbod/tilburg/koop,valid,allowed_official_source,realworks_public,aanbodEntry\n",
        encoding="utf-8",
    )
    result = run_noord_brabant_coverage_source_census(
        repo_root=tmp_path,
        evidence_paths=(Path("data/processed/sources_seed_noord_brabant.csv"),),
        output_dir=tmp_path / "tmp" / "generated",
        completion_scope_verification=True,
    )
    assert result.workbook_path and result.workbook_path.name == "noord_brabant_source_completion_scope_verification_v1.xlsx"
    assert result.master_csv_path and result.master_csv_path.name == "noord_brabant_source_completion_scope_verification_v1.csv"
    assert result.review_queue_csv_path and result.review_queue_csv_path.name == "noord_brabant_source_completion_scope_verification_v1_review_queue.csv"
    assert result.realworks_audit_input_csv_path and result.realworks_audit_input_csv_path.exists()
