from __future__ import annotations

import ast
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR / "scraper" / "src"))

from domek_wonen.facts import build_property_fact_value, build_property_facts_record  # noqa: E402
from domek_wonen.parsers.models import ParsedListing  # noqa: E402
from domek_wonen.pilots.realworks_excel_export import (  # noqa: E402
    FIELD_GAP_FIELDS,
    PROPERTY_COLUMNS,
    WORKSHEET_NAMES,
    export_realworks_readiness_rows_to_excel,
)
from domek_wonen.pilots.realworks_property_readiness import (  # noqa: E402
    RealworksPropertyReadinessResult,
    build_realworks_property_readiness_row,
)


MODULE_PATH = BASE_DIR / "scraper" / "src" / "domek_wonen" / "pilots" / "realworks_excel_export.py"
NOW = datetime(2026, 6, 28, 10, 0, tzinfo=UTC)


def _listing(index: int = 1, **overrides: object) -> ParsedListing:
    listing = ParsedListing(
        source_id="oldenkotte.com__tilburg",
        source_domain="oldenkotte.com",
        canonical_url=f"http://www.oldenkotte.com/aanbod/woningaanbod/tilburg/koop/huis-10145{index}-Exportstraat-{index}",
        address_raw=f"Exportstraat {index}",
        street="Exportstraat",
        house_number=str(index),
        postcode="",
        city="Tilburg",
        asking_price_eur=425000 + index,
        transaction_type="koop",
        status="beschikbaar",
        living_area_m2=120 + index,
        rooms_count=5,
        bedrooms_count=3,
        property_type="woonhuis",
        confidence_score=0.95,
    )
    values = {field: getattr(listing, field) for field in listing.__dataclass_fields__}
    values.update(overrides)
    return ParsedListing(**values)


def _fact(
    field: str,
    value: object,
    *,
    normalized_value: object | None = None,
    status: str = "usable",
    warnings: tuple[str, ...] = (),
):
    return build_property_fact_value(
        field=field,
        value=value,
        normalized_value=value if normalized_value is None else normalized_value,
        source="realworks_kenmerk",
        confidence=0.90,
        status=status,
        evidence_preview=str(value or ""),
        warnings=warnings,
    )


def _record(index: int = 1, *facts, warnings: tuple[str, ...] = ()):
    return build_property_facts_record(
        source_id="oldenkotte.com__tilburg",
        source_domain="oldenkotte.com",
        canonical_url=f"http://www.oldenkotte.com/aanbod/woningaanbod/tilburg/koop/huis-10145{index}-Exportstraat-{index}",
        address_raw=f"Exportstraat {index}",
        city="Tilburg",
        status="beschikbaar",
        fetched_at="2026-06-28T10:00:00Z",
        expires_at="2026-07-12T10:00:00Z",
        facts=facts,
        warnings=warnings,
    )


def _full_record(index: int = 1):
    return _record(
        index,
        _fact("asking_price", 425000),
        _fact("property_type", "Woonhuis", normalized_value="woonhuis"),
        _fact("availability_date", "Beschikbaar"),
        _fact("rooms", 5),
        _fact("bedrooms", 3),
        _fact("bathrooms", 1),
        _fact("living_area_m2", 123),
        _fact("plot_area_m2", 234),
        _fact("volume_m3", 456),
        _fact("energy_label", "A"),
        _fact("bouwjaar", 1998),
        _fact("heating_type", "CV-ketel", normalized_value="cv_ketel"),
        _fact("garden", "Achtertuin"),
        _fact("parking", "Openbaar parkeren"),
        _fact("garage", "geen_garage"),
        _fact("eigendomssituatie", "Volle eigendom", normalized_value="volle_eigendom"),
        _fact("description_length_bucket", "medium"),
        warnings=("description_not_stored",),
    )


def _apartment_record(index: int = 3):
    return _record(
        index,
        _fact("asking_price", 325000),
        _fact("property_type", "Appartement", normalized_value="appartement"),
        _fact("availability_date", "Beschikbaar"),
        _fact("rooms", 3),
        _fact("bedrooms", 2),
        _fact("bathrooms", 1),
        _fact("living_area_m2", 83),
        _fact("plot_area_m2", None, normalized_value=None, status="missing"),
        _fact("volume_m3", 250),
        _fact("energy_label", "Niet aanwezig", normalized_value=None, status="review", warnings=("energy_label_not_explicit",)),
        _fact("bouwjaar", 1998),
        _fact("heating_type", "CV-ketel", normalized_value="cv_ketel"),
        _fact("garden", None, normalized_value=None, status="missing"),
        _fact("parking", "Openbaar parkeren"),
        _fact("garage", None, normalized_value=None, status="missing"),
        _fact("eigendomssituatie", "Volle eigendom", normalized_value="volle_eigendom"),
        _fact("description_length_bucket", "medium"),
    )


def _problem_record(index: int = 2):
    return _record(
        index,
        _fact("asking_price", 149500),
        _fact(
            "property_type",
            "OverigOG",
            normalized_value="unknown",
            status="review",
            warnings=("unsupported_property_type_overigog",),
        ),
        _fact("availability_date", "Beschikbaar"),
        _fact("rooms", None, normalized_value=None, status="missing"),
        _fact("bedrooms", None, normalized_value=None, status="missing"),
        _fact("bathrooms", None, normalized_value=None, status="missing"),
        _fact("living_area_m2", None, normalized_value=None, status="missing"),
        _fact("plot_area_m2", 108),
        _fact("volume_m3", None, normalized_value=None, status="missing"),
        _fact("energy_label", "Niet aanwezig", normalized_value=None, status="review", warnings=("energy_label_not_explicit",)),
        _fact("bouwjaar", None, normalized_value=None, status="missing"),
        _fact("heating_type", None, normalized_value=None, status="missing"),
        _fact("garden", None, normalized_value=None, status="missing"),
        _fact("parking", None, normalized_value=None, status="missing"),
        _fact("garage", None, normalized_value=None, status="missing"),
        _fact("eigendomssituatie", "Volle eigendom", normalized_value="volle_eigendom"),
        _fact("description_length_bucket", "short"),
        warnings=("description_not_stored",),
    )


def _row(index: int = 1, record=None, **listing_overrides: object):
    return build_realworks_property_readiness_row(_listing(index, **listing_overrides), record or _full_record(index))


def _readiness_result(rows):
    rows = tuple(rows)
    return RealworksPropertyReadinessResult(
        source_id="oldenkotte.com__tilburg",
        source_domain="oldenkotte.com",
        listing_parser_total=len(rows),
        listing_qa_clean=len(rows),
        listing_qa_review=0,
        listing_qa_rejected=0,
        detail_attempted=len(rows),
        detail_succeeded=len(rows),
        detail_failed=0,
        facts_records_built=len(rows),
        readiness_rows_built=len(rows),
        quality_status_counts=tuple(sorted({row.quality_status for row in rows})),
        export_readiness_counts=tuple(sorted({row.export_readiness for row in rows})),
        field_completion_counts=(),
        missing_key_fields_counts=(),
        review_fields_counts=(),
        warning_counts=(),
        sample_rows_compact=(),
        problem_rows_compact=(),
        excel_validation_ready=True,
        rows=rows,
    )


def _export(tmp_path: Path, *rows):
    output_path = tmp_path / "generated" / "realworks.xlsx"
    result = export_realworks_readiness_rows_to_excel(
        rows=tuple(rows),
        output_path=output_path,
        readiness_result=_readiness_result(rows),
        generated_at=NOW,
    )
    workbook = load_workbook(output_path)
    return result, workbook


def _headers(worksheet):
    return [cell.value for cell in worksheet[1]]


def _col(headers, name: str) -> int:
    return headers.index(name) + 1


def _all_workbook_text(workbook) -> str:
    values: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
    return "\n".join(values)


def test_writes_workbook_to_temp_path(tmp_path: Path) -> None:
    result, workbook = _export(tmp_path, _row(1))

    assert result.output_path == tmp_path / "generated" / "realworks.xlsx"
    assert result.output_path.exists()
    assert workbook["Realworks Properties"].max_row == 2


def test_creates_expected_sheets(tmp_path: Path) -> None:
    result, workbook = _export(tmp_path, _row(1))

    assert result.worksheet_names == WORKSHEET_NAMES
    assert tuple(workbook.sheetnames) == WORKSHEET_NAMES


def test_writes_all_rows_including_export_review(tmp_path: Path) -> None:
    review = _row(2, postcode="")
    assert review.export_readiness == "export_review"

    result, workbook = _export(tmp_path, _row(1), review)

    assert result.rows_written == 2
    assert workbook["Realworks Properties"].max_row == 3


def test_required_property_columns_exist(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1))

    assert _headers(workbook["Realworks Properties"]) == list(PROPERTY_COLUMNS)


def test_preserves_canonical_url_as_full_text(tmp_path: Path) -> None:
    row = _row(1)
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "canonical_url")).value == row.canonical_url


def test_creates_clickable_property_link(tmp_path: Path) -> None:
    row = _row(1)
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)
    cell = worksheet.cell(2, _col(headers, "property_link"))

    assert cell.value == "Open listing"
    assert cell.hyperlink.target == row.canonical_url


def test_summary_counts_are_written(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1), _row(2, _problem_record(2)))
    values = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}

    assert values["listing_parser_total"] == 2
    assert values["readiness_rows_built"] == 2
    assert values["excel_rows_written"] == 2
    assert values["advisor_review_count"] == 1
    assert values["blocked_count"] == 1
    assert values["export_review_count"] == 1
    assert values["export_blocked_count"] == 1
    assert values["artifact_statement"] == "Excel validation artifact only"
    assert values["production_statement"] == "Not client-ready production output"


def test_field_gaps_are_written(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1), _row(2, _problem_record(2)))
    rows = {row[0].value: [cell.value for cell in row] for row in workbook["Field Gaps"].iter_rows(min_row=2)}

    assert set(FIELD_GAP_FIELDS).issubset(rows)
    assert rows["postcode"][1:4] == [0, 0, 2]
    assert rows["postcode_status"][1:4] == [0, 0, 2]
    assert rows["coordinates"][1:4] == [0, 0, 2]
    assert rows["property_type"][1:4] == [1, 1, 0]
    assert rows["energy_label_status"][1:4] == [1, 1, 0]
    assert rows["residential_classification"][1:4] == [1, 0, 1]


def test_warnings_are_aggregated_with_sample_urls(tmp_path: Path) -> None:
    row_one = _row(1)
    row_two = _row(2, _problem_record(2))
    _, workbook = _export(tmp_path, row_one, row_two)
    values = {row[0].value: [cell.value for cell in row] for row in workbook["Warnings"].iter_rows(min_row=2)}

    assert values["missing_coordinates"][1] == 2
    assert row_one.canonical_url in values["missing_coordinates"][2]
    assert values["description_not_stored"][1] == 2
    assert values["unsupported_property_type_overigog"][1] == 1
    assert values["non_residential_property_type"][1] == 1


def test_problem_rows_are_sorted_by_problem_severity(tmp_path: Path) -> None:
    low = _row(1)
    high = _row(2, _problem_record(2), address_raw="Corellistraat")
    _, workbook = _export(tmp_path, low, high)
    worksheet = workbook["Problem Rows"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "address")).value == "Corellistraat"
    assert worksheet.cell(2, _col(headers, "problem_score")).value > worksheet.cell(3, _col(headers, "problem_score")).value


def test_excel_includes_critical_status_columns(tmp_path: Path) -> None:
    _, workbook = _export(tmp_path, _row(1, _apartment_record(1), property_type="appartement"))
    headers = _headers(workbook["Realworks Properties"])

    for column in (
        "residential_classification",
        "postcode_status",
        "postcode_source",
        "postcode_review_reason",
        "source_status",
        "status_bucket",
        "active_inventory_eligible",
        "db_persistence_action",
        "source_published_at",
        "source_published_at_raw",
        "source_published_at_source",
        "source_published_at_status",
        "source_published_at_review_reason",
        "first_seen_at",
        "last_seen_at",
        "observed_at",
        "status_first_seen_at",
        "status_changed_at",
        "price_first_seen_at",
        "price_changed_at",
        "removed_at",
        "days_on_market_source",
        "days_since_first_seen",
        "freshness_bucket",
        "lifecycle_events",
        "vve_active",
        "vve_monthly_cost",
        "vve_status",
        "vve_review_reason",
        "energy_label_status",
        "energy_label_raw",
        "energy_label_review_reason",
    ):
        assert column in headers


def test_summary_has_source_published_freshness_and_lifecycle_counts(tmp_path: Path) -> None:
    row = build_realworks_property_readiness_row(
        _listing(1),
        _full_record(1),
        observed_at=NOW,
        source_published_at=datetime(2026, 6, 28, tzinfo=UTC),
        source_published_at_raw="2026-06-28",
        source_published_at_source="json_ld_datePublished",
        source_published_at_status="usable",
    )
    _, workbook = _export(tmp_path, row)
    values = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}

    assert values["source_published_at_usable_count"] == 1
    assert values["source_published_at_review_count"] == 0
    assert values["source_published_at_missing_count"] == 0
    assert values["freshness_bucket_counts"] == "new_today=1"
    assert values["lifecycle_event_counts"] == "new_listing=1"


def test_workbook_writes_rows_when_source_published_at_missing(tmp_path: Path) -> None:
    row = build_realworks_property_readiness_row(_listing(1), _full_record(1), observed_at=NOW)
    result, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert result.rows_written == 1
    assert worksheet.cell(2, _col(headers, "source_published_at")).value in (None, "")
    assert worksheet.cell(2, _col(headers, "source_published_at_status")).value == "missing"
    assert worksheet.cell(2, _col(headers, "first_seen_at")).value == "2026-06-28T10:00:00Z"
    assert worksheet.cell(2, _col(headers, "observed_at")).value == "2026-06-28T10:00:00Z"


def test_excel_writes_extracted_postcode_and_status_policy(tmp_path: Path) -> None:
    row = _row(1, postcode="5044 SN", status="beschikbaar")
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "postcode")).value == "5044 SN"
    assert worksheet.cell(2, _col(headers, "postcode_status")).value == "usable"
    assert worksheet.cell(2, _col(headers, "status_bucket")).value == "active_available"
    assert worksheet.cell(2, _col(headers, "active_inventory_eligible")).value is True
    assert worksheet.cell(2, _col(headers, "db_persistence_action")).value == "store_active_candidate"


def test_verkocht_rows_are_visible_but_not_active_eligible(tmp_path: Path) -> None:
    row = _row(1, postcode="5044 SN", status="verkocht")
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "source_status")).value == "verkocht"
    assert worksheet.cell(2, _col(headers, "status_bucket")).value == "inactive_sold"
    assert worksheet.cell(2, _col(headers, "active_inventory_eligible")).value is False
    assert worksheet.cell(2, _col(headers, "db_persistence_action")).value == "store_status_history"


def test_apartment_without_vve_and_energy_review_are_visible(tmp_path: Path) -> None:
    row = _row(1, _apartment_record(1), property_type="appartement")
    _, workbook = _export(tmp_path, row)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert worksheet.cell(2, _col(headers, "vve_status")).value == "missing"
    assert worksheet.cell(2, _col(headers, "vve_missing_reason")).value == "missing_vve_for_apartment"
    assert worksheet.cell(2, _col(headers, "energy_label")).value in (None, "")
    assert worksheet.cell(2, _col(headers, "energy_label_raw")).value == "Niet aanwezig"
    assert worksheet.cell(2, _col(headers, "energy_label_status")).value == "review"
    assert "missing_vve_for_apartment" in worksheet.cell(2, _col(headers, "warnings")).value


def test_blocked_rows_are_still_written_for_human_validation(tmp_path: Path) -> None:
    blocked = _row(2, _problem_record(2), address_raw="Corellistraat")
    result, workbook = _export(tmp_path, blocked)
    worksheet = workbook["Realworks Properties"]
    headers = _headers(worksheet)

    assert result.rows_written == 1
    assert worksheet.cell(2, _col(headers, "residential_classification")).value == "non_residential_blocked"
    assert worksheet.cell(2, _col(headers, "quality_status")).value == "blocked"
    assert worksheet.cell(2, _col(headers, "export_readiness")).value == "export_blocked"
    assert worksheet.cell(2, _col(headers, "active_inventory_eligible")).value is False
    assert worksheet.cell(2, _col(headers, "db_persistence_action")).value == "store_excluded_non_residential"


def test_does_not_include_raw_html_json_or_long_description_text(tmp_path: Path) -> None:
    long_text = "Omschrijving " + ("ruime woning " * 80)
    record = _record(
        1,
        _fact("asking_price", 425000),
        _fact("property_type", "<html>{\"docs\":[]}</html>", normalized_value="woonhuis", status="review"),
        _fact("description_length_bucket", long_text),
        _fact("parking", "https://cdn.example.test/photo.jpg"),
    )
    _, workbook = _export(tmp_path, _row(1, record))
    text = _all_workbook_text(workbook).casefold()

    assert "<html" not in text
    assert '"docs"' not in text
    assert "photo.jpg" not in text
    assert "ruime woning ruime woning ruime woning ruime woning ruime woning ruime woning" not in text


def test_handles_empty_warnings_and_missing_fields(tmp_path: Path) -> None:
    row = replace(
        _row(1),
        warnings=(),
        missing_key_fields=(),
        review_fields=(),
        quality_status="client_ready",
        export_readiness="export_ready",
        postcode_status="usable",
        postcode_source="parsed_listing",
        postcode_review_reason=None,
    )
    result, workbook = _export(tmp_path, row)

    assert result.warning_counts == ()
    assert workbook["Warnings"].max_row == 1
    assert workbook["Problem Rows"].max_row == 1


def test_does_not_require_live_network() -> None:
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    imported_modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.update(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_modules.add(node.module)

    disallowed = {"requests", "httpx", "playwright", "selenium"}
    assert disallowed.isdisjoint({module.split(".")[0] for module in imported_modules})
