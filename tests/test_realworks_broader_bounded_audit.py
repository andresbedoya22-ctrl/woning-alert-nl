from __future__ import annotations

import ast
from datetime import UTC, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR / "scraper" / "src"))

from domek_wonen.compliance import robots_gate  # noqa: E402
from domek_wonen.pilots.realworks_broader_bounded_audit import (  # noqa: E402
    OUTPUT_SUMMARY_NAME,
    OUTPUT_WORKBOOK_NAME,
    RealworksBroaderAuditSource,
    broader_audit_source_from_evidence_row,
    run_realworks_broader_bounded_audit,
    select_realworks_broader_audit_sources,
)


MODULE_PATH = BASE_DIR / "scraper" / "src" / "domek_wonen" / "pilots" / "realworks_broader_bounded_audit.py"
OBSERVED = datetime(2026, 6, 29, 9, 0, tzinfo=UTC)


def _row(
    *,
    source_id: str = "alpha.nl__breda",
    source_name: str = "Alpha",
    domain: str = "alpha.nl",
    listing_url: str = "https://www.alpha.nl/aanbod/woningaanbod/koop",
    gemeente: str = "Breda",
    province: str = "Noord-Brabant",
    access_status: str = "allowed",
    delivery_mode: str = "realworks_public",
    parser_family_candidate: str = "realworks_public",
    platform_guess: str = "realworks",
    needs_review: str = "false",
    evidence_file: str = "local.csv",
) -> dict[str, str]:
    return {
        "source_id": source_id,
        "source_name": source_name,
        "domain": domain,
        "listing_url": listing_url,
        "gemeente": gemeente,
        "province": province,
        "access_status": access_status,
        "delivery_mode": delivery_mode,
        "parser_family_candidate": parser_family_candidate,
        "platform_guess": platform_guess,
        "needs_review": needs_review,
        "confidence_score": "95",
        "evidence_file": evidence_file,
        "notes": "local Realworks evidence",
    }


def _source(source_id: str, domain: str) -> RealworksBroaderAuditSource:
    return RealworksBroaderAuditSource(
        source_id=source_id,
        source_name=source_id,
        domain=domain,
        listing_url=f"https://{domain}/aanbod/woningaanbod/koop",
        gemeente="Breda",
        province="Noord-Brabant",
        platform_guess="realworks",
        delivery_mode="realworks_public",
        parser_family_candidate="realworks_public",
        access_policy_status="allowed",
        selection_confidence="95",
        selection_evidence_file="local.csv",
        selection_reason="test",
    )


def _listing_html(domain: str, count: int = 1) -> str:
    cards = []
    for index in range(1, count + 1):
        cards.append(
            f"""
            <li class="aanbodEntry">
              <a href="/aanbod/woningaanbod/breda/koop/huis-100{index}-teststraat-{index}">
                <span class="street-address">Teststraat {index}</span>
                <span class="locality">Breda</span>
                <span class="price">EUR 425.000 k.k.</span>
                <span class="objectstatusbanner">Beschikbaar</span>
                <span>Woonhuis 123 m2 5 kamers Te koop {domain}</span>
              </a>
            </li>
            """
        )
    return f"<html><body>{''.join(cards)}</body></html>"


def _detail_html(*, property_type: str = "Woonhuis", source_published: str = "2026-06-29") -> str:
    return f"""
    <html>
      <head>
        <meta name="description" content="Compacte omschrijving.">
        <script type="application/ld+json">{{"datePublished":"{source_published}"}}</script>
      </head>
      <body>
        <div>Teststraat 1 | 4811 AA Breda</div>
        <span class="kenmerk"><span class="kenmerkName">Soort object</span><span class="kenmerkValue">{property_type}</span></span>
        <span class="kenmerk"><span class="kenmerkName">Vraagprijs</span><span class="kenmerkValue">EUR 425.000 k.k.</span></span>
        <span class="kenmerk"><span class="kenmerkName">Status</span><span class="kenmerkValue">Beschikbaar</span></span>
        <span class="kenmerk"><span class="kenmerkName">Woonoppervlakte</span><span class="kenmerkValue">123 m2</span></span>
        <span class="kenmerk"><span class="kenmerkName">Perceeloppervlakte</span><span class="kenmerkValue">234 m2</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal kamers</span><span class="kenmerkValue">5 kamers</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal slaapkamers</span><span class="kenmerkValue">3 slaapkamers</span></span>
        <span class="kenmerk"><span class="kenmerkName">Aantal badkamers</span><span class="kenmerkValue">1 badkamer</span></span>
        <span class="kenmerk"><span class="kenmerkName">Inhoud</span><span class="kenmerkValue">456 m3</span></span>
        <span class="kenmerk"><span class="kenmerkName">Energieklasse</span><span class="kenmerkValue">A</span></span>
        <span class="kenmerk"><span class="kenmerkName">Bouwjaar</span><span class="kenmerkValue">1998</span></span>
        <span class="kenmerk"><span class="kenmerkName">Verwarming</span><span class="kenmerkValue">CV-ketel</span></span>
        <span class="kenmerk"><span class="kenmerkName">Tuin</span><span class="kenmerkValue">Achtertuin</span></span>
        <span class="kenmerk"><span class="kenmerkName">Parkeerfaciliteiten</span><span class="kenmerkValue">Openbaar parkeren</span></span>
        <span class="kenmerk"><span class="kenmerkName">Eigendomssituatie</span><span class="kenmerkValue">Volle eigendom</span></span>
      </body>
    </html>
    """


def _fetch(url: str) -> str:
    domain = url.split("//", 1)[1].split("/", 1)[0].removeprefix("www.")
    if url.rstrip("/").endswith("/koop") or url.rstrip("/").endswith("/woningaanbod"):
        return _listing_html(domain)
    return _detail_html()


def _headers(workbook, sheet: str) -> list[str]:
    return [cell.value for cell in workbook[sheet][1]]


def _col(headers: list[str], name: str) -> int:
    return headers.index(name) + 1


def test_selects_additional_sources_excluding_oldenkotte_and_olden() -> None:
    result = select_realworks_broader_audit_sources(
        [
            _row(source_id="oldenkotte.com__tilburg", domain="oldenkotte.com"),
            _row(source_id="olden.nl__heusden", domain="olden.nl"),
            _row(source_id="alpha.nl__breda", domain="alpha.nl"),
            _row(source_id="beta.nl__tilburg", domain="beta.nl"),
        ]
    )

    assert [source.source_id for source in result] == ["alpha.nl__breda", "beta.nl__tilburg"]


def test_does_not_invent_listing_url() -> None:
    assert broader_audit_source_from_evidence_row(_row(listing_url="")) is None


def test_does_not_use_property_url_as_listing_url() -> None:
    row = _row(listing_url="https://www.alpha.nl/aanbod/woningaanbod/breda/koop/huis-1001-teststraat-1")

    assert broader_audit_source_from_evidence_row(row) is None


def test_excludes_funda_and_pararius() -> None:
    rows = [
        _row(source_id="funda", domain="funda.nl", listing_url="https://www.funda.nl/koop/tilburg/"),
        _row(source_id="pararius", domain="pararius.nl", listing_url="https://www.pararius.nl/koopwoningen"),
    ]

    assert select_realworks_broader_audit_sources(rows) == ()


def test_excludes_blocked_legal_review_permission_login_captcha_and_403() -> None:
    rows = [
        _row(source_id="blocked", domain="blocked.nl", access_status="blocked"),
        _row(source_id="legal", domain="legal.nl", access_status="legal_review"),
        _row(source_id="permission", domain="permission.nl", access_status="permission_required"),
        _row(source_id="login", domain="login.nl", needs_review="true"),
        _row(source_id="captcha", domain="captcha.nl", evidence_file="captcha.csv"),
        _row(source_id="forbidden", domain="forbidden.nl", evidence_file="403.csv"),
    ]

    assert select_realworks_broader_audit_sources(rows) == ()


def test_respects_max_sources() -> None:
    rows = [
        _row(source_id="alpha.nl__breda", domain="alpha.nl"),
        _row(source_id="beta.nl__breda", domain="beta.nl"),
        _row(source_id="gamma.nl__breda", domain="gamma.nl"),
    ]

    assert len(select_realworks_broader_audit_sources(rows, max_sources=2)) == 2


def test_respects_max_detail_fetches_per_source(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)
    calls: list[str] = []

    def fetch(url: str) -> str:
        calls.append(url)
        if url.rstrip("/").endswith("/koop"):
            return _listing_html("alpha.nl", count=3)
        return _detail_html()

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=fetch,
        max_detail_fetches_per_source=2,
    )

    assert result.metrics[0].detail_attempted == 2


def test_calls_robots_gate_before_listing_fetch(monkeypatch) -> None:
    events: list[str] = []

    def can_fetch(domain: str, path: str) -> bool:
        events.append(f"robots:{path}")
        return True

    def fetch(url: str) -> str:
        events.append(f"fetch:{url}")
        return _fetch(url)

    monkeypatch.setattr(robots_gate, "can_fetch", can_fetch)
    run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=fetch,
    )

    assert events[0].startswith("robots:")
    assert events[1].startswith("fetch:")


def test_calls_robots_gate_before_each_detail_fetch(monkeypatch) -> None:
    events: list[str] = []

    def can_fetch(domain: str, path: str) -> bool:
        events.append(f"robots:{path}")
        return True

    def fetch(url: str) -> str:
        events.append(f"fetch:{url}")
        return _fetch(url)

    monkeypatch.setattr(robots_gate, "can_fetch", can_fetch)
    run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=fetch,
    )

    detail_fetch_indexes = [
        index for index, event in enumerate(events) if "huis-1001-teststraat-1" in event and event.startswith("fetch:")
    ]
    assert detail_fetch_indexes
    assert all(events[index - 1].startswith("robots:") for index in detail_fetch_indexes)


def test_aggregates_per_source_and_property_metrics(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=_fetch,
    )

    assert result.metrics[0].parser_total == 1
    assert result.metrics[0].parser_qa_clean == 1
    assert result.metrics[0].facts_records_built == 1
    assert result.metrics[0].readiness_rows_built == 1
    assert result.metrics[0].export_ready == 1


def test_writes_manual_verification_workbook(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=_fetch,
        output_dir=tmp_path / "generated",
    )

    assert result.workbook_path == tmp_path / "generated" / OUTPUT_WORKBOOK_NAME
    assert result.summary_csv_path == tmp_path / "generated" / OUTPUT_SUMMARY_NAME
    workbook = load_workbook(result.workbook_path)
    assert tuple(workbook.sheetnames) == (
        "Sources Summary",
        "Manual Verification",
        "Field Gaps",
        "Warnings",
        "Problem Rows",
        "Family Audit Decision",
    )
    headers = _headers(workbook, "Manual Verification")
    assert "property_link" in headers
    assert "manual_check_result" in headers
    assert "manual_check_notes" in headers
    assert workbook["Manual Verification"].cell(2, _col(headers, "manual_check_result")).value in (None, "")
    assert workbook["Manual Verification"].cell(2, _col(headers, "manual_check_notes")).value in (None, "")


def test_family_decision_ready_when_two_additional_sources_pass(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=_fetch,
    )

    assert result.family_decision.family_decision == "realworks_ready_for_noord_brabant_realworks_audit"
    assert result.family_decision.second_wave_sources_validated == 2


def test_family_decision_needs_hardening_when_source_fails_parser(monkeypatch) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    def fetch(url: str) -> str:
        if "beta.nl" in url and url.rstrip("/").endswith("/koop"):
            return "<html></html>"
        return _fetch(url)

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=fetch,
    )

    assert result.family_decision.family_decision == "realworks_needs_hardening_v2"


def test_family_decision_insufficient_when_less_than_two_sources() -> None:
    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"),),
        observed_at=OBSERVED,
        fetch_html=_fetch,
    )

    assert result.metrics == ()
    assert result.family_decision.family_decision == "insufficient_additional_sources"


def test_does_not_write_raw_html_or_json(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)

    run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=_fetch,
        output_dir=tmp_path / "generated",
    )

    written_files = {path.name for path in (tmp_path / "generated").iterdir()}
    assert written_files == {OUTPUT_WORKBOOK_NAME, OUTPUT_SUMMARY_NAME}
    assert not any(path.suffix.lower() in {".html", ".json", ".har"} for path in (tmp_path / "generated").iterdir())


def test_does_not_include_long_descriptions(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(robots_gate, "can_fetch", lambda domain, path: True)
    long_description = "Ruime woning " * 100

    def fetch(url: str) -> str:
        if url.rstrip("/").endswith("/koop"):
            return _listing_html("alpha.nl")
        return _detail_html().replace("Compacte omschrijving.", long_description)

    result = run_realworks_broader_bounded_audit(
        sources=(_source("alpha.nl__breda", "alpha.nl"), _source("beta.nl__breda", "beta.nl")),
        observed_at=OBSERVED,
        fetch_html=fetch,
        output_dir=tmp_path / "generated",
    )
    text = "\n".join(
        str(cell.value)
        for worksheet in load_workbook(result.workbook_path).worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert long_description.strip() not in text


def test_does_not_create_parser_per_makelaar() -> None:
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    class_names = {node.name.casefold() for node in ast.walk(tree) if isinstance(node, ast.ClassDef)}

    assert not any("oldenkotte" in name or "olden" in name for name in class_names)


def test_no_matching_imports() -> None:
    tree = ast.parse(MODULE_PATH.read_text(encoding="utf-8"))
    imported_modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.update(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_modules.add(node.module)

    assert "domek_wonen.matching" not in imported_modules
