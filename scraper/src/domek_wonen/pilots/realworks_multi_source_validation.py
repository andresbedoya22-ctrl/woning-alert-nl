from __future__ import annotations

import csv
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.pilots.realworks_property_readiness import (
    RealworksPropertyReadinessResult,
    RealworksPropertyReadinessRow,
    run_realworks_property_readiness,
)


MAX_REALWORKS_VALIDATION_SOURCES = 4
DEFAULT_MAX_DETAIL_FETCHES_PER_SOURCE = 12
DEFAULT_TIMEOUT_SECONDS = 15.0
REALWORKS_FAMILY = "realworks_public"
VALIDATION_WORKSHEET_NAMES = (
    "Sources Summary",
    "Per-source Metrics",
    "All Realworks Properties",
    "Field Gaps",
    "Warnings",
    "Problem Rows",
    "Family Decision",
)
BLOCKED_ACCESS_STATUSES = {"blocked", "legal_review", "permission_required"}
ALLOWED_ACCESS_STATUSES = {"allowed", "limited"}
RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
LONG_TEXT_LIMIT = 500


@dataclass(frozen=True, slots=True)
class RealworksValidationSource:
    source_id: str
    source_name: str
    domain: str
    listing_url: str
    gemeente: str
    platform_guess: str
    delivery_mode: str
    parser_family_candidate: str
    access_policy_status: str
    selection_reason: str


@dataclass(frozen=True, slots=True)
class RealworksSourceValidationMetrics:
    source_id: str
    source_name: str
    domain: str
    listing_url: str
    listing_fetch_status: str
    robots_allowed_listing: bool
    parser_total: int
    parser_qa_clean: int
    parser_qa_review: int
    parser_qa_rejected: int
    detail_attempted: int
    detail_succeeded: int
    detail_failed: int
    facts_records_built: int
    readiness_rows_built: int
    client_ready: int
    advisor_review: int
    blocked: int
    export_ready: int
    export_review: int
    export_blocked: int
    active_inventory_eligible_true: int
    active_inventory_eligible_false: int
    db_persistence_action_counts: tuple[tuple[str, int], ...]
    postcode_usable: int
    postcode_missing: int
    vve_missing_for_apartment: int
    energy_label_usable: int
    energy_label_review_missing: int
    non_residential_rows: int
    source_published_at_usable: int
    source_published_at_review: int
    source_published_at_missing: int
    freshness_bucket_counts: tuple[tuple[str, int], ...]
    lifecycle_event_counts: tuple[tuple[str, int], ...]
    top_warnings: tuple[tuple[str, int], ...]
    top_field_gaps: tuple[tuple[str, int], ...]
    problem_row_count: int
    validation_status: str
    validation_decision: str


@dataclass(frozen=True, slots=True)
class RealworksFamilyDecision:
    validated_source_count: int
    passed_count: int
    passed_with_review_gaps_count: int
    needs_hardening_count: int
    blocked_count: int
    failed_count: int
    second_makelaar_validated: bool
    oldenkotte_control_passed: bool
    family_decision: str
    recommended_next_action: str


@dataclass(frozen=True, slots=True)
class RealworksMultiSourceValidationResult:
    sources: tuple[RealworksValidationSource, ...]
    metrics: tuple[RealworksSourceValidationMetrics, ...]
    readiness_results: tuple[RealworksPropertyReadinessResult, ...]
    family_decision: RealworksFamilyDecision
    workbook_path: Path | None = None
    summary_csv_path: Path | None = None
    observed_at: datetime | None = None
    warnings: tuple[str, ...] = ()


def load_realworks_validation_sources_from_local_evidence(
    *,
    repo_root: Path,
    max_sources: int = MAX_REALWORKS_VALIDATION_SOURCES,
) -> tuple[RealworksValidationSource, ...]:
    rows: list[dict[str, Any]] = []
    processed_seed = repo_root / "data" / "processed" / "sources_seed_noord_brabant.csv"
    overrides = repo_root / "data" / "discovery" / "reference" / "property_discovery_source_overrides.csv"
    rows.extend(_read_csv_rows(processed_seed, evidence_file=str(processed_seed)))
    rows.extend(_read_csv_rows(overrides, evidence_file=str(overrides)))
    return select_realworks_validation_sources(rows, max_sources=max_sources)


def select_realworks_validation_sources(
    evidence_rows: Iterable[Mapping[str, Any]],
    *,
    max_sources: int = MAX_REALWORKS_VALIDATION_SOURCES,
) -> tuple[RealworksValidationSource, ...]:
    if max_sources <= 0:
        return ()

    candidates: dict[str, RealworksValidationSource] = {}
    for row in evidence_rows:
        source = validation_source_from_evidence_row(row)
        if source is None:
            continue
        candidates.setdefault(source.source_id, source)

    selected: list[RealworksValidationSource] = []
    for preferred in ("oldenkotte.com__tilburg", "olden.nl__heusden"):
        source = candidates.pop(preferred, None)
        if source is not None:
            selected.append(source)

    remaining = sorted(candidates.values(), key=lambda source: (source.domain, source.source_id))
    selected.extend(remaining[: max(0, max_sources - len(selected))])
    return tuple(selected[:max_sources])


def validation_source_from_evidence_row(row: Mapping[str, Any]) -> RealworksValidationSource | None:
    normalized = {_normalize_key(key): _clean_text(value) for key, value in row.items()}
    domain = _normalize_domain(_first_value(normalized, ("domain", "root_domain", "source_domain", "website_domain")))
    listing_url = _clean_text(_first_value(normalized, ("listing_url", "aanbod_url", "koopaanbod_url", "source_url")))
    source_id = _clean_text(_first_value(normalized, ("source_id", "id"))) or _source_id(domain, normalized)
    source_name = _clean_text(_first_value(normalized, ("source_name", "office_name", "makelaar_name", "name")))
    gemeente = _clean_text(_first_value(normalized, ("gemeente", "plaats", "city", "municipality")))
    access_status = _access_policy_status(normalized)
    delivery_mode = _delivery_mode(normalized, listing_url)
    parser_family = _parser_family(normalized, delivery_mode)
    platform_guess = _platform_guess(normalized, delivery_mode, listing_url)

    if not source_id or not domain or not listing_url:
        return None
    if not _is_http_url(listing_url):
        return None
    if _is_funda_or_pararius(domain, listing_url):
        return None
    if access_status in BLOCKED_ACCESS_STATUSES:
        return None
    if access_status not in ALLOWED_ACCESS_STATUSES:
        return None
    if _has_blocked_signal(normalized):
        return None
    if delivery_mode != REALWORKS_FAMILY or parser_family != REALWORKS_FAMILY:
        return None

    return RealworksValidationSource(
        source_id=source_id,
        source_name=source_name or source_id,
        domain=domain,
        listing_url=listing_url,
        gemeente=gemeente,
        platform_guess=platform_guess,
        delivery_mode=delivery_mode,
        parser_family_candidate=parser_family,
        access_policy_status=access_status,
        selection_reason=_selection_reason(normalized),
    )


def run_realworks_multi_source_validation(
    *,
    sources: Sequence[RealworksValidationSource],
    observed_at: datetime,
    output_dir: Path | None = None,
    max_sources: int = MAX_REALWORKS_VALIDATION_SOURCES,
    max_listing_fetches_per_source: int = 1,
    max_detail_fetches_per_source: int = DEFAULT_MAX_DETAIL_FETCHES_PER_SOURCE,
    timeout_seconds: float = DEFAULT_TIMEOUT_SECONDS,
    fetch_html: Callable[[str], str] | None = None,
) -> RealworksMultiSourceValidationResult:
    observed = _to_utc(observed_at)
    capped_sources = tuple(sources[: max(0, min(max_sources, MAX_REALWORKS_VALIDATION_SOURCES))])
    capped_detail_fetches = max(0, min(max_detail_fetches_per_source, DEFAULT_MAX_DETAIL_FETCHES_PER_SOURCE))
    readiness_results: list[RealworksPropertyReadinessResult] = []
    metrics: list[RealworksSourceValidationMetrics] = []
    warnings: list[str] = []

    for source in capped_sources:
        readiness = run_realworks_property_readiness(
            source_id=source.source_id,
            source_domain=source.domain,
            listing_url=source.listing_url,
            max_listing_fetches=max_listing_fetches_per_source,
            max_detail_fetches=capped_detail_fetches,
            timeout_seconds=timeout_seconds,
            fetch_html=fetch_html,
            observed_at=observed,
        )
        readiness_results.append(readiness)
        metrics.append(_metrics_for_source(source, readiness))
        warnings.extend(readiness.warnings)

    family_decision = decide_realworks_family(metrics)
    result = RealworksMultiSourceValidationResult(
        sources=capped_sources,
        metrics=tuple(metrics),
        readiness_results=tuple(readiness_results),
        family_decision=family_decision,
        observed_at=observed,
        warnings=_dedupe(warnings),
    )
    if output_dir is None:
        return result

    output_dir = Path(output_dir)
    workbook_path = output_dir / "realworks_multi_source_validation_v1.xlsx"
    summary_csv_path = output_dir / "realworks_multi_source_validation_v1_summary.csv"
    write_realworks_multi_source_workbook(result, workbook_path)
    write_realworks_multi_source_summary_csv(result, summary_csv_path)
    return RealworksMultiSourceValidationResult(
        sources=result.sources,
        metrics=result.metrics,
        readiness_results=result.readiness_results,
        family_decision=result.family_decision,
        workbook_path=workbook_path,
        summary_csv_path=summary_csv_path,
        observed_at=result.observed_at,
        warnings=result.warnings,
    )


def decide_realworks_family(metrics: Sequence[RealworksSourceValidationMetrics]) -> RealworksFamilyDecision:
    statuses = Counter(metric.validation_status for metric in metrics)
    validated = sum(1 for metric in metrics if metric.validation_status in {"passed", "passed_with_review_gaps"})
    oldenkotte_passed = any(
        metric.source_id == "oldenkotte.com__tilburg"
        and metric.validation_status in {"passed", "passed_with_review_gaps"}
        for metric in metrics
    )
    second_validated = any(
        metric.source_id != "oldenkotte.com__tilburg"
        and metric.validation_status in {"passed", "passed_with_review_gaps"}
        for metric in metrics
    )
    if validated >= 2 and oldenkotte_passed and second_validated and statuses["needs_hardening"] == 0:
        decision = "realworks_family_usable_for_broader_audit"
        next_action = "Proceed to bounded broader Realworks audit before Noord-Brabant-wide application."
    elif statuses["needs_hardening"] > 0 or statuses["failed"] > 0:
        decision = "realworks_family_needs_hardening_v2"
        next_action = "Run Realworks Family Hardening v2 before broader audit."
    elif statuses["blocked"] and statuses["blocked"] == len(metrics):
        decision = "realworks_family_blocked_by_access"
        next_action = "Do not proceed until access policy permits enough sources."
    else:
        decision = "insufficient_sources_to_decide"
        next_action = "Find at least one additional locally evidenced Realworks source before deciding."
    return RealworksFamilyDecision(
        validated_source_count=validated,
        passed_count=statuses["passed"],
        passed_with_review_gaps_count=statuses["passed_with_review_gaps"],
        needs_hardening_count=statuses["needs_hardening"],
        blocked_count=statuses["blocked"],
        failed_count=statuses["failed"],
        second_makelaar_validated=second_validated,
        oldenkotte_control_passed=oldenkotte_passed,
        family_decision=decision,
        recommended_next_action=next_action,
    )


def write_realworks_multi_source_workbook(
    result: RealworksMultiSourceValidationResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Sources Summary"
    _write_sources_summary(workbook["Sources Summary"], result.sources)
    _write_per_source_metrics(workbook.create_sheet("Per-source Metrics"), result.metrics)
    _write_all_properties(workbook.create_sheet("All Realworks Properties"), result.readiness_results)
    _write_field_gaps(workbook.create_sheet("Field Gaps"), result.readiness_results)
    _write_warnings(workbook.create_sheet("Warnings"), result.readiness_results)
    _write_problem_rows(workbook.create_sheet("Problem Rows"), result.readiness_results)
    _write_family_decision(workbook.create_sheet("Family Decision"), result.family_decision)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def write_realworks_multi_source_summary_csv(
    result: RealworksMultiSourceValidationResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = (
        "source_id",
        "domain",
        "listing_url",
        "validation_status",
        "validation_decision",
        "parser_total",
        "parser_qa_clean",
        "detail_attempted",
        "detail_succeeded",
        "detail_failed",
        "facts_records_built",
        "readiness_rows_built",
        "export_ready",
        "export_review",
        "export_blocked",
    )
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for metric in result.metrics:
            writer.writerow({column: _safe_cell(getattr(metric, column)) for column in columns})
    return output_path


def _metrics_for_source(
    source: RealworksValidationSource,
    readiness: RealworksPropertyReadinessResult,
) -> RealworksSourceValidationMetrics:
    rows = tuple(readiness.rows)
    quality_counts = Counter(row.quality_status for row in rows)
    export_counts = Counter(row.export_readiness for row in rows)
    active_counts = Counter(row.active_inventory_eligible for row in rows)
    db_counts = _counter_pairs(row.db_persistence_action for row in rows)
    freshness_counts = _counter_pairs(row.freshness_bucket for row in rows)
    lifecycle_counts = _counter_pairs(event for row in rows for event in row.lifecycle_events)
    warning_counts = _top_counter_pairs(row.warnings for row in rows)
    field_gaps = _top_field_gaps(rows)
    non_residential_rows = sum(1 for row in rows if row.residential_classification.startswith("non_residential"))
    validation_status = _validation_status(readiness)
    return RealworksSourceValidationMetrics(
        source_id=source.source_id,
        source_name=source.source_name,
        domain=source.domain,
        listing_url=source.listing_url,
        listing_fetch_status=_listing_fetch_status(readiness),
        robots_allowed_listing="listing_blocked_by_robots" not in readiness.warnings,
        parser_total=readiness.listing_parser_total,
        parser_qa_clean=readiness.listing_qa_clean,
        parser_qa_review=readiness.listing_qa_review,
        parser_qa_rejected=readiness.listing_qa_rejected,
        detail_attempted=readiness.detail_attempted,
        detail_succeeded=readiness.detail_succeeded,
        detail_failed=readiness.detail_failed,
        facts_records_built=readiness.facts_records_built,
        readiness_rows_built=readiness.readiness_rows_built,
        client_ready=quality_counts["client_ready"],
        advisor_review=quality_counts["advisor_review"],
        blocked=quality_counts["blocked"],
        export_ready=export_counts["export_ready"],
        export_review=export_counts["export_review"],
        export_blocked=export_counts["export_blocked"],
        active_inventory_eligible_true=active_counts[True],
        active_inventory_eligible_false=active_counts[False],
        db_persistence_action_counts=db_counts,
        postcode_usable=sum(1 for row in rows if row.postcode_status == "usable"),
        postcode_missing=sum(1 for row in rows if row.postcode_status == "missing"),
        vve_missing_for_apartment=sum(1 for row in rows if row.vve_status == "missing"),
        energy_label_usable=sum(1 for row in rows if row.energy_label_status == "usable"),
        energy_label_review_missing=sum(1 for row in rows if row.energy_label_status in {"review", "missing"}),
        non_residential_rows=non_residential_rows,
        source_published_at_usable=sum(1 for row in rows if row.source_published_at_status == "usable"),
        source_published_at_review=sum(1 for row in rows if row.source_published_at_status == "review"),
        source_published_at_missing=sum(1 for row in rows if row.source_published_at_status == "missing"),
        freshness_bucket_counts=freshness_counts,
        lifecycle_event_counts=lifecycle_counts,
        top_warnings=warning_counts,
        top_field_gaps=field_gaps,
        problem_row_count=len(readiness.problem_rows_compact),
        validation_status=validation_status,
        validation_decision=_validation_decision(validation_status),
    )


def _validation_status(readiness: RealworksPropertyReadinessResult) -> str:
    warnings = set(readiness.warnings)
    if "listing_blocked_by_robots" in warnings or "detail_blocked_by_robots" in warnings:
        return "blocked"
    if "listing_fetch_parse_or_qa_exception" in warnings:
        return "failed"
    if readiness.listing_parser_total <= 0 or readiness.listing_qa_clean <= 0:
        return "needs_hardening"
    if readiness.detail_attempted > 0 and readiness.detail_succeeded == 0:
        return "needs_hardening"
    if readiness.readiness_rows_built <= 0:
        return "needs_hardening"
    if readiness.detail_failed > readiness.detail_succeeded:
        return "needs_hardening"
    if _has_review_gaps(readiness.rows):
        return "passed_with_review_gaps"
    return "passed"


def _validation_decision(validation_status: str) -> str:
    return {
        "passed": "source passed without source-specific code",
        "passed_with_review_gaps": "source passed with explicit review gaps",
        "needs_hardening": "family needs hardening before broader audit",
        "blocked": "source blocked by access or robots policy",
        "failed": "technical failure unrelated to access policy",
    }.get(validation_status, "unknown validation state")


def _has_review_gaps(rows: Sequence[RealworksPropertyReadinessRow]) -> bool:
    return any(
        row.quality_status != "client_ready"
        or row.missing_key_fields
        or row.review_fields
        or row.vve_status in {"missing", "review"}
        or row.energy_label_status in {"missing", "review"}
        or row.source_published_at_status in {"missing", "review"}
        for row in rows
    )


def _listing_fetch_status(readiness: RealworksPropertyReadinessResult) -> str:
    if "listing_blocked_by_robots" in readiness.warnings:
        return "robots_blocked"
    if "listing_fetch_parse_or_qa_exception" in readiness.warnings:
        return "failed"
    if readiness.listing_parser_total > 0:
        return "success"
    return "empty"


def _write_sources_summary(worksheet: Worksheet, sources: Sequence[RealworksValidationSource]) -> None:
    columns = tuple(RealworksValidationSource.__dataclass_fields__)
    worksheet.append(columns)
    for source in sources:
        worksheet.append([_safe_cell(getattr(source, column)) for column in columns])


def _write_per_source_metrics(worksheet: Worksheet, metrics: Sequence[RealworksSourceValidationMetrics]) -> None:
    columns = tuple(RealworksSourceValidationMetrics.__dataclass_fields__)
    worksheet.append(columns)
    for metric in metrics:
        worksheet.append([_safe_cell(_metric_value(metric, column)) for column in columns])


def _write_all_properties(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    columns = (
        "source_id",
        "source_domain",
        "canonical_url",
        "property_link",
        "address",
        "postcode",
        "city",
        "quality_status",
        "export_readiness",
        "status_bucket",
        "active_inventory_eligible",
        "db_persistence_action",
        "residential_classification",
        "asking_price",
        "property_type",
        "energy_label",
        "energy_label_status",
        "vve_status",
        "source_published_at_status",
        "freshness_bucket",
        "lifecycle_events",
        "missing_key_fields",
        "review_fields",
        "warnings",
    )
    worksheet.append(columns)
    for readiness in readiness_results:
        for row in readiness.rows:
            values = {
                "source_id": row.source_id,
                "source_domain": row.source_domain,
                "canonical_url": row.canonical_url,
                "property_link": "Open listing" if _is_http_url(row.canonical_url) else "",
                "address": row.address,
                "postcode": row.postcode,
                "city": row.city,
                "quality_status": row.quality_status,
                "export_readiness": row.export_readiness,
                "status_bucket": row.status_bucket,
                "active_inventory_eligible": row.active_inventory_eligible,
                "db_persistence_action": row.db_persistence_action,
                "residential_classification": row.residential_classification,
                "asking_price": row.asking_price,
                "property_type": row.property_type,
                "energy_label": row.energy_label,
                "energy_label_status": row.energy_label_status,
                "vve_status": row.vve_status,
                "source_published_at_status": row.source_published_at_status,
                "freshness_bucket": row.freshness_bucket,
                "lifecycle_events": _join(row.lifecycle_events),
                "missing_key_fields": _join(row.missing_key_fields),
                "review_fields": _join(row.review_fields),
                "warnings": _join(row.warnings),
            }
            worksheet.append([_safe_cell(values[column]) for column in columns])
            link_cell = worksheet.cell(worksheet.max_row, columns.index("property_link") + 1)
            if _is_http_url(row.canonical_url):
                link_cell.hyperlink = row.canonical_url
                link_cell.style = "Hyperlink"


def _write_field_gaps(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "field", "usable_count", "review_count", "missing_count"))
    for readiness in readiness_results:
        for field, usable, review, missing in readiness.field_completion_counts:
            worksheet.append((readiness.source_id, field, usable, review, missing))
        for field, count in _top_field_gaps(readiness.rows, limit=100):
            worksheet.append((readiness.source_id, field, 0, 0, count))


def _write_warnings(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "warning", "count"))
    for readiness in readiness_results:
        for warning, count in readiness.warning_counts:
            worksheet.append((readiness.source_id, warning, count))


def _write_problem_rows(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "canonical_url", "address", "quality_status", "export_readiness", "issues"))
    for readiness in readiness_results:
        for row in readiness.problem_rows_compact:
            issues = _join((*row.missing_key_fields, *row.review_fields, *row.warnings))
            worksheet.append(
                (
                    readiness.source_id,
                    _safe_cell(row.canonical_url),
                    _safe_cell(row.address),
                    row.quality_status,
                    row.export_readiness,
                    _safe_cell(issues),
                )
            )


def _write_family_decision(worksheet: Worksheet, decision: RealworksFamilyDecision) -> None:
    worksheet.append(("metric", "value"))
    for field in RealworksFamilyDecision.__dataclass_fields__:
        worksheet.append((field, _safe_cell(getattr(decision, field))))


def _read_csv_rows(path: Path, *, evidence_file: str) -> tuple[dict[str, Any], ...]:
    if not path.exists():
        return ()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return tuple({**row, "evidence_file": evidence_file} for row in csv.DictReader(handle))


def _access_policy_status(row: Mapping[str, str]) -> str:
    value = _normalize_key(_first_value(row, ("access_policy_status", "access_status", "legal_status")))
    if value == "allowed_official_source":
        return "allowed"
    if value in {"requires_permission", "needs_permission"}:
        return "permission_required"
    if value in ALLOWED_ACCESS_STATUSES | BLOCKED_ACCESS_STATUSES:
        return value
    if _normalize_key(row.get("needs_review")) in {"true", "1", "yes", "ja"}:
        return "permission_required"
    if _normalize_key(_first_value(row, ("koopaanbod_url_quality", "aanbod_url_quality", "source_quality_status"))) == "valid":
        return "allowed"
    return "legal_review"


def _delivery_mode(row: Mapping[str, str], listing_url: str) -> str:
    value = _normalize_key(_first_value(row, ("delivery_mode", "detected_delivery_mode", "detected_delivery_mode_enriched")))
    if value:
        return value
    text = _joined_text(row, listing_url)
    if "realworks_public" in text or "realworks" in text or "/aanbod/woningaanbod" in listing_url.casefold():
        return REALWORKS_FAMILY
    return ""


def _parser_family(row: Mapping[str, str], delivery_mode: str) -> str:
    value = _normalize_key(_first_value(row, ("parser_family_candidate", "parser_family", "detected_parser_family")))
    return value or (REALWORKS_FAMILY if delivery_mode == REALWORKS_FAMILY else "")


def _platform_guess(row: Mapping[str, str], delivery_mode: str, listing_url: str) -> str:
    value = _normalize_key(_first_value(row, ("platform_guess", "detected_platform", "platform", "cms")))
    if value:
        return value
    if delivery_mode == REALWORKS_FAMILY or "/aanbod/woningaanbod" in listing_url.casefold():
        return "realworks"
    return ""


def _selection_reason(row: Mapping[str, str]) -> str:
    evidence_file = _clean_text(row.get("evidence_file"))
    reason = _first_value(row, ("source_quality_reason", "reason", "notes", "evidence"))
    parts = [part for part in (reason, evidence_file) if part]
    return "; ".join(parts) or "local Realworks evidence"


def _source_id(domain: str, row: Mapping[str, str]) -> str:
    gemeente = _normalize_key(_first_value(row, ("gemeente", "plaats", "city", "municipality")))
    if domain and gemeente:
        return f"{domain}__{gemeente}"
    return domain


def _has_blocked_signal(row: Mapping[str, str]) -> bool:
    text = _joined_text(row)
    return any(marker in text for marker in ("captcha", "login", "login_required", "403", "forbidden", "paywall"))


def _is_funda_or_pararius(domain: str, listing_url: str) -> bool:
    text = f"{domain} {listing_url}".casefold()
    return "funda.nl" in text or "pararius.nl" in text


def _top_field_gaps(rows: Sequence[RealworksPropertyReadinessRow], limit: int = 10) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for row in rows:
        counts.update(row.missing_key_fields)
        counts.update(f"review:{field}" for field in row.review_fields)
        if row.vve_status == "missing":
            counts["vve_active"] += 1
        if row.energy_label_status in {"review", "missing"}:
            counts["energy_label"] += 1
        if row.source_published_at_status in {"review", "missing"}:
            counts["source_published_at"] += 1
    return tuple(counts.most_common(limit))


def _top_counter_pairs(values: Iterable[Iterable[str]], limit: int = 10) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for items in values:
        counts.update(items)
    return tuple(counts.most_common(limit))


def _counter_pairs(values: Iterable[str]) -> tuple[tuple[str, int], ...]:
    return tuple(sorted(Counter(value for value in values if value).items()))


def _metric_value(metric: RealworksSourceValidationMetrics, column: str) -> object:
    value = getattr(metric, column)
    if isinstance(value, tuple):
        return _counts_tuple_text(value)
    return value


def _counts_tuple_text(values: tuple[tuple[str, int], ...]) -> str:
    return "; ".join(f"{key}={count}" for key, count in values)


def _format_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[letter].width = max(10, min(max_length + 2, 60))


def _first_value(row: Mapping[str, str], keys: Iterable[str]) -> str:
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return ""


def _normalize_domain(value: object) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    candidate = text if "://" in text else f"https://{text}"
    parts = urlsplit(candidate)
    host = (parts.netloc or parts.path).split("/", 1)[0].split(":", 1)[0].strip().lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def _is_http_url(value: object) -> bool:
    parts = urlsplit(str(value or "").strip())
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " ".join(_clean_text(item) for item in value if _clean_text(item))
    return str(value).strip()


def _normalize_key(value: object) -> str:
    return _clean_text(value).casefold().replace("-", "_").replace(" ", "_")


def _joined_text(row: Mapping[str, str], extra: str = "") -> str:
    return _normalize_key(" ".join((*row.values(), extra)))


def _join(values: Iterable[object]) -> str:
    return "; ".join(str(_safe_cell(value)) for value in values if _safe_cell(value) not in (None, ""))


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in RAW_MARKERS):
        return ""
    if len(text) > LONG_TEXT_LIMIT:
        return "[long_text_omitted]"
    return text


def _dedupe(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return tuple(result)


def _to_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC)
