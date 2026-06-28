from __future__ import annotations

from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.facts.models import FACT_STATUS_REVIEW, FACT_STATUS_USABLE
from domek_wonen.pilots.realworks_property_readiness import (
    RealworksPropertyReadinessResult,
    RealworksPropertyReadinessRow,
    run_realworks_property_readiness,
)


PROPERTY_COLUMNS = (
    "source_id",
    "source_domain",
    "canonical_url",
    "property_link",
    "address",
    "postcode",
    "postcode_status",
    "postcode_source",
    "postcode_review_reason",
    "city",
    "asking_price",
    "property_type",
    "status",
    "availability",
    "rooms",
    "bedrooms",
    "bathrooms",
    "living_area_m2",
    "plot_area_m2",
    "volume_m3",
    "energy_label",
    "energy_label_status",
    "energy_label_raw",
    "energy_label_review_reason",
    "bouwjaar",
    "heating",
    "garden",
    "parking",
    "garage",
    "ownership_or_erfpacht",
    "vve_active",
    "vve_monthly_cost",
    "vve_status",
    "vve_review_reason",
    "vve_missing_reason",
    "description_length_bucket",
    "residential_classification",
    "location_readiness",
    "quality_status",
    "export_readiness",
    "missing_key_fields",
    "review_fields",
    "warnings",
    "client_summary_headline",
    "client_summary_facts",
    "client_summary_financial",
    "client_summary_energy",
    "client_summary_outdoor",
    "client_summary_attention",
)
FIELD_GAP_FIELDS = (
    "property_type",
    "asking_price",
    "availability",
    "rooms",
    "bedrooms",
    "bathrooms",
    "living_area_m2",
    "plot_area_m2",
    "volume_m3",
    "energy_label",
    "bouwjaar",
    "heating",
    "garden",
    "parking",
    "garage",
    "ownership_or_erfpacht",
    "description_length_bucket",
    "postcode",
    "postcode_status",
    "coordinates",
    "vve_active",
    "vve_monthly_cost",
    "vve_status",
    "energy_label_status",
    "residential_classification",
)
WORKSHEET_NAMES = ("Realworks Properties", "Summary", "Field Gaps", "Warnings", "Problem Rows")
_FACT_FIELD_ALIASES = {
    "availability": "availability_date",
    "heating": "heating_type",
    "ownership_or_erfpacht": "eigendomssituatie",
}
_RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif", ".svg")
_LONG_TEXT_LIMIT = 500


@dataclass(frozen=True, slots=True)
class RealworksExcelExportResult:
    output_path: Path
    rows_written: int
    worksheet_names: tuple[str, ...]
    export_ready_count: int
    export_review_count: int
    export_blocked_count: int
    client_ready_count: int
    advisor_review_count: int
    blocked_count: int
    field_gaps: tuple[tuple[str, int, int, int, str], ...]
    warning_counts: tuple[tuple[str, int], ...]
    generated_at: str
    warnings: tuple[str, ...] = ()


def export_realworks_readiness_rows_to_excel(
    *,
    rows: Sequence[RealworksPropertyReadinessRow],
    output_path: Path,
    readiness_result: RealworksPropertyReadinessResult | None = None,
    source_id: str | None = None,
    source_domain: str | None = None,
    generated_at: datetime | None = None,
) -> RealworksExcelExportResult:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = tuple(rows)
    generated_at_text = _datetime_to_utc_iso(generated_at or datetime.now(UTC))
    source_id = source_id or (readiness_result.source_id if readiness_result else _first_attr(rows, "source_id"))
    source_domain = source_domain or (
        readiness_result.source_domain if readiness_result else _first_attr(rows, "source_domain")
    )

    workbook = Workbook()
    properties_ws = workbook.active
    properties_ws.title = "Realworks Properties"
    _write_properties_sheet(properties_ws, rows)

    summary_ws = workbook.create_sheet("Summary")
    _write_summary_sheet(
        summary_ws,
        rows,
        readiness_result=readiness_result,
        source_id=source_id,
        source_domain=source_domain,
        generated_at=generated_at_text,
    )

    field_gaps = _field_gaps(rows)
    gaps_ws = workbook.create_sheet("Field Gaps")
    _write_field_gaps_sheet(gaps_ws, field_gaps)

    warning_details = _warning_details(rows)
    warnings_ws = workbook.create_sheet("Warnings")
    _write_warnings_sheet(warnings_ws, warning_details)

    problem_ws = workbook.create_sheet("Problem Rows")
    _write_problem_rows_sheet(problem_ws, rows)

    for worksheet in workbook.worksheets:
        _format_table_sheet(worksheet)

    workbook.save(output_path)

    quality_counts = Counter(row.quality_status for row in rows)
    export_counts = Counter(row.export_readiness for row in rows)
    return RealworksExcelExportResult(
        output_path=output_path,
        rows_written=len(rows),
        worksheet_names=tuple(workbook.sheetnames),
        export_ready_count=export_counts["export_ready"],
        export_review_count=export_counts["export_review"],
        export_blocked_count=export_counts["export_blocked"],
        client_ready_count=quality_counts["client_ready"],
        advisor_review_count=quality_counts["advisor_review"],
        blocked_count=quality_counts["blocked"],
        field_gaps=tuple((gap.field, gap.usable_count, gap.review_count, gap.missing_count, gap.notes) for gap in field_gaps),
        warning_counts=tuple((item.warning, item.count) for item in warning_details),
        generated_at=generated_at_text,
        warnings=_dedupe(warning for row in rows for warning in row.warnings),
    )


def run_realworks_excel_export(
    *,
    source_id: str,
    source_domain: str,
    listing_url: str,
    output_path: Path,
    max_listing_fetches: int = 1,
    max_detail_fetches: int = 9,
    timeout_seconds: float = 15.0,
) -> tuple[RealworksPropertyReadinessResult, RealworksExcelExportResult]:
    readiness = run_realworks_property_readiness(
        source_id=source_id,
        source_domain=source_domain,
        listing_url=listing_url,
        max_listing_fetches=max_listing_fetches,
        max_detail_fetches=max_detail_fetches,
        timeout_seconds=timeout_seconds,
    )
    export = export_realworks_readiness_rows_to_excel(
        rows=readiness.rows,
        output_path=output_path,
        readiness_result=readiness,
        source_id=source_id,
        source_domain=source_domain,
    )
    return readiness, export


def _write_properties_sheet(worksheet: Worksheet, rows: Sequence[RealworksPropertyReadinessRow]) -> None:
    worksheet.append(PROPERTY_COLUMNS)
    for row in rows:
        values = _property_row_values(row)
        worksheet.append([values[column] for column in PROPERTY_COLUMNS])
        link_cell = worksheet.cell(worksheet.max_row, PROPERTY_COLUMNS.index("property_link") + 1)
        if _is_valid_http_url(row.canonical_url):
            link_cell.hyperlink = row.canonical_url
            link_cell.style = "Hyperlink"


def _write_summary_sheet(
    worksheet: Worksheet,
    rows: Sequence[RealworksPropertyReadinessRow],
    *,
    readiness_result: RealworksPropertyReadinessResult | None,
    source_id: str,
    source_domain: str,
    generated_at: str,
) -> None:
    quality_counts = Counter(row.quality_status for row in rows)
    export_counts = Counter(row.export_readiness for row in rows)
    metrics = (
        ("source_id", source_id),
        ("source_domain", source_domain),
        ("listing_parser_total", _metric(readiness_result, "listing_parser_total")),
        ("listing_qa_clean", _metric(readiness_result, "listing_qa_clean")),
        ("detail_attempted", _metric(readiness_result, "detail_attempted")),
        ("detail_succeeded", _metric(readiness_result, "detail_succeeded")),
        ("detail_failed", _metric(readiness_result, "detail_failed")),
        ("facts_records_built", _metric(readiness_result, "facts_records_built")),
        ("readiness_rows_built", _metric(readiness_result, "readiness_rows_built", len(rows))),
        ("excel_rows_written", len(rows)),
        ("quality_status_counts", _counts_text(quality_counts)),
        ("export_readiness_counts", _counts_text(export_counts)),
        ("client_ready_count", quality_counts["client_ready"]),
        ("advisor_review_count", quality_counts["advisor_review"]),
        ("blocked_count", quality_counts["blocked"]),
        ("export_ready_count", export_counts["export_ready"]),
        ("export_review_count", export_counts["export_review"]),
        ("export_blocked_count", export_counts["export_blocked"]),
        ("generated_at", generated_at),
        ("artifact_statement", "Excel validation artifact only"),
        ("production_statement", "Not client-ready production output"),
    )
    worksheet.append(("metric", "value"))
    for metric in metrics:
        worksheet.append(metric)


def _write_field_gaps_sheet(worksheet: Worksheet, gaps: Sequence["_FieldGap"]) -> None:
    worksheet.append(("field", "usable_count", "review_count", "missing_count", "notes"))
    for gap in gaps:
        worksheet.append((gap.field, gap.usable_count, gap.review_count, gap.missing_count, gap.notes))


def _write_warnings_sheet(worksheet: Worksheet, warnings: Sequence["_WarningDetail"]) -> None:
    worksheet.append(("warning", "count", "sample_canonical_urls", "notes"))
    for item in warnings:
        worksheet.append((item.warning, item.count, _join(item.sample_canonical_urls), item.notes))


def _write_problem_rows_sheet(worksheet: Worksheet, rows: Sequence[RealworksPropertyReadinessRow]) -> None:
    columns = (
        "canonical_url",
        "property_link",
        "address",
        "city",
        "asking_price",
        "quality_status",
        "export_readiness",
        "problem_score",
        "issue_summary",
        "missing_key_fields",
        "review_fields",
        "warnings",
    )
    worksheet.append(columns)
    for row in sorted(rows, key=_problem_score, reverse=True):
        score = _problem_score(row)
        if score <= 0:
            continue
        worksheet.append(
            (
                _safe_cell(row.canonical_url),
                "Open listing" if _is_valid_http_url(row.canonical_url) else "",
                _safe_cell(row.address),
                _safe_cell(row.city),
                row.asking_price,
                row.quality_status,
                row.export_readiness,
                score,
                _issue_summary(row),
                _join(row.missing_key_fields),
                _join(row.review_fields),
                _join(row.warnings),
            )
        )
        link_cell = worksheet.cell(worksheet.max_row, 2)
        if _is_valid_http_url(row.canonical_url):
            link_cell.hyperlink = row.canonical_url
            link_cell.style = "Hyperlink"


def _property_row_values(row: RealworksPropertyReadinessRow) -> dict[str, object]:
    values = {
        "source_id": row.source_id,
        "source_domain": row.source_domain,
        "canonical_url": row.canonical_url,
        "property_link": "Open listing" if _is_valid_http_url(row.canonical_url) else "",
        "address": row.address,
        "postcode": row.postcode,
        "postcode_status": row.postcode_status,
        "postcode_source": row.postcode_source,
        "postcode_review_reason": row.postcode_review_reason,
        "city": row.city,
        "asking_price": row.asking_price,
        "property_type": row.property_type,
        "status": row.status,
        "availability": row.availability,
        "rooms": row.rooms,
        "bedrooms": row.bedrooms,
        "bathrooms": row.bathrooms,
        "living_area_m2": row.living_area_m2,
        "plot_area_m2": row.plot_area_m2,
        "volume_m3": row.volume_m3,
        "energy_label": row.energy_label,
        "energy_label_status": row.energy_label_status,
        "energy_label_raw": row.energy_label_raw,
        "energy_label_review_reason": row.energy_label_review_reason,
        "bouwjaar": row.bouwjaar,
        "heating": row.heating,
        "garden": row.garden,
        "parking": row.parking,
        "garage": row.garage,
        "ownership_or_erfpacht": row.ownership_or_erfpacht,
        "vve_active": row.vve_active,
        "vve_monthly_cost": row.vve_monthly_cost,
        "vve_status": row.vve_status,
        "vve_review_reason": row.vve_review_reason,
        "vve_missing_reason": row.vve_missing_reason,
        "description_length_bucket": row.description_length_bucket,
        "residential_classification": row.residential_classification,
        "location_readiness": row.location_readiness.location_status,
        "quality_status": row.quality_status,
        "export_readiness": row.export_readiness,
        "missing_key_fields": _join(row.missing_key_fields),
        "review_fields": _join(row.review_fields),
        "warnings": _join(row.warnings),
        "client_summary_headline": row.client_summary.headline,
        "client_summary_facts": row.client_summary.facts_line,
        "client_summary_financial": row.client_summary.financial_line,
        "client_summary_energy": row.client_summary.energy_line,
        "client_summary_outdoor": row.client_summary.outdoor_line,
        "client_summary_attention": _join(row.client_summary.attention_points),
    }
    return {key: _safe_cell(value) for key, value in values.items()}


@dataclass(frozen=True, slots=True)
class _FieldGap:
    field: str
    usable_count: int
    review_count: int
    missing_count: int
    notes: str


def _field_gaps(rows: Sequence[RealworksPropertyReadinessRow]) -> tuple[_FieldGap, ...]:
    return tuple(_field_gap(field, rows) for field in FIELD_GAP_FIELDS)


def _field_gap(field: str, rows: Sequence[RealworksPropertyReadinessRow]) -> _FieldGap:
    if field == "postcode":
        usable = sum(1 for row in rows if row.postcode)
        missing = len(rows) - usable
        return _FieldGap(field, usable, 0, missing, "location field")
    if field == "postcode_status":
        usable = sum(1 for row in rows if row.postcode_status == "usable")
        missing = sum(1 for row in rows if row.postcode_status == "missing")
        review = len(rows) - usable - missing
        return _FieldGap(field, usable, review, missing, "critical production field")
    if field == "coordinates":
        usable = sum(
            1
            for row in rows
            if row.location_readiness.latitude is not None and row.location_readiness.longitude is not None
        )
        missing = len(rows) - usable
        return _FieldGap(field, usable, 0, missing, "latitude and longitude")
    if field == "vve_status":
        usable = sum(1 for row in rows if row.vve_status in {"usable", "not_applicable"})
        review = sum(1 for row in rows if row.vve_status == "review")
        missing = sum(1 for row in rows if row.vve_status == "missing")
        return _FieldGap(field, usable, review, missing, "missing VvE requires apartment review")
    if field == "energy_label_status":
        usable = sum(1 for row in rows if row.energy_label_status == "usable")
        review = sum(1 for row in rows if row.energy_label_status == "review")
        missing = sum(1 for row in rows if row.energy_label_status == "missing")
        return _FieldGap(field, usable, review, missing, "energy label value/status/raw separated")
    if field == "residential_classification":
        usable = sum(1 for row in rows if row.residential_classification == "residential")
        review = sum(1 for row in rows if row.residential_classification.endswith("_review"))
        missing = sum(1 for row in rows if row.residential_classification.endswith("_blocked"))
        return _FieldGap(field, usable, review, missing, "non-residential rows are not production-ready")

    fact_field = _FACT_FIELD_ALIASES.get(field, field)
    usable = 0
    review = 0
    missing = 0
    for row in rows:
        facts = {fact.field: fact.status for fact in row.facts_record.facts} if row.facts_record else {}
        status = facts.get(fact_field)
        if status == FACT_STATUS_USABLE:
            usable += 1
        elif status == FACT_STATUS_REVIEW:
            review += 1
        else:
            missing += 1
    return _FieldGap(field, usable, review, missing, _field_note(field, usable, review, missing))


@dataclass(frozen=True, slots=True)
class _WarningDetail:
    warning: str
    count: int
    sample_canonical_urls: tuple[str, ...]
    notes: str


def _warning_details(rows: Sequence[RealworksPropertyReadinessRow]) -> tuple[_WarningDetail, ...]:
    urls_by_warning: dict[str, list[str]] = {}
    for row in rows:
        for warning in row.warnings:
            urls_by_warning.setdefault(warning, []).append(row.canonical_url)
    details = [
        _WarningDetail(
            warning=warning,
            count=len(urls),
            sample_canonical_urls=tuple(urls[:3]),
            notes=_warning_note(warning),
        )
        for warning, urls in urls_by_warning.items()
    ]
    return tuple(sorted(details, key=lambda item: (-item.count, item.warning)))


def _problem_score(row: RealworksPropertyReadinessRow) -> int:
    score = 0
    if row.quality_status == "blocked":
        score += 6
    if row.export_readiness == "export_blocked":
        score += 4
    if row.quality_status == "advisor_review":
        score += 2
    score += min(8, len(row.missing_key_fields))
    score += min(6, len(row.review_fields) * 2)
    if "unsupported_property_type_overigog" in row.warnings:
        score += 3
    if "non_residential_property_type" in row.warnings:
        score += 6
    if row.postcode_status == "missing":
        score += 2
    if row.vve_status == "missing":
        score += 2
    if "missing_coordinates" in row.warnings:
        score += 1
    return score


def _issue_summary(row: RealworksPropertyReadinessRow) -> str:
    issues: list[str] = []
    if row.quality_status != "client_ready":
        issues.append(row.quality_status)
    if row.export_readiness != "export_ready":
        issues.append(row.export_readiness)
    issues.extend(f"missing:{field}" for field in row.missing_key_fields)
    issues.extend(f"review:{field}" for field in row.review_fields)
    issues.extend(row.warnings)
    return _join(_dedupe(issues))


def _format_table_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 60))


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in _RAW_MARKERS):
        return ""
    if _contains_image_url(text):
        return ""
    if _is_valid_http_url(text):
        return text
    if len(text) > _LONG_TEXT_LIMIT:
        return "[long_text_omitted]"
    return text


def _contains_image_url(text: str) -> bool:
    for token in text.replace(";", " ").replace(",", " ").split():
        lowered = token.casefold().split("?", 1)[0]
        if lowered.startswith(("http://", "https://")) and lowered.endswith(_IMAGE_EXTENSIONS):
            return True
    return False


def _is_valid_http_url(value: object) -> bool:
    parts = urlsplit(str(value or "").strip())
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def _join(values: Iterable[object]) -> str:
    return "; ".join(str(_safe_cell(value)) for value in values if _safe_cell(value))


def _dedupe(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return tuple(result)


def _metric(result: RealworksPropertyReadinessResult | None, field: str, fallback: int = 0) -> int:
    return int(getattr(result, field, fallback)) if result is not None else fallback


def _counts_text(counts: Counter[str]) -> str:
    return "; ".join(f"{key}={value}" for key, value in sorted(counts.items()) if key)


def _first_attr(rows: Sequence[RealworksPropertyReadinessRow], field: str) -> str:
    return str(getattr(rows[0], field, "")) if rows else ""


def _field_note(field: str, usable: int, review: int, missing: int) -> str:
    if review:
        return "review values require human validation"
    if missing:
        return "missing in readiness rows"
    if usable:
        return "usable in exported rows"
    return ""


def _warning_note(warning: str) -> str:
    notes = {
        "missing_coordinates": "coordinates are not available in current Realworks readiness rows",
        "missing_postcode": "postcode is not available in current Oldenkotte extraction",
        "description_not_stored": "description bucket only; long description text is not exported",
        "cv_ketel_ownership_not_clear": "CV-ketel ownership remains unresolved",
        "energy_label_not_explicit": "energy label signal needs human review",
        "hot_water_not_normalized": "hot-water vocabulary needs later normalization",
        "heating_not_normalized": "heating vocabulary needs later normalization",
        "unsupported_property_type_overigog": "property type is unsupported for automatic client-ready promotion",
        "missing_vve_for_apartment": "apartment has no explicit VvE evidence",
        "non_residential_property_type": "listing appears garage/storage/parking or otherwise non-residential",
    }
    return notes.get(warning, "")


def _datetime_to_utc_iso(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")
