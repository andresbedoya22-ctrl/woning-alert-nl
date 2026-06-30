from __future__ import annotations

import csv
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.pilots.realworks_multi_source_validation import (
    RealworksSourceValidationMetrics,
    RealworksValidationSource,
    validation_source_from_evidence_row,
)
from domek_wonen.pilots.realworks_property_readiness import (
    RealworksPropertyReadinessResult,
    RealworksPropertyReadinessRow,
    run_realworks_property_readiness,
)


MAX_BROADER_AUDIT_SOURCES = 3
DEFAULT_MAX_DETAIL_FETCHES_PER_SOURCE = 15
DEFAULT_TIMEOUT_SECONDS = 15.0
CONTROL_SOURCE_IDS = frozenset({"oldenkotte.com__tilburg", "olden.nl__heusden"})
CONTROL_DOMAINS = frozenset({"oldenkotte.com", "olden.nl"})
OUTPUT_WORKBOOK_NAME = "realworks_broader_bounded_audit_v1.xlsx"
OUTPUT_SUMMARY_NAME = "realworks_broader_bounded_audit_v1_summary.csv"
RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
LONG_TEXT_LIMIT = 500


@dataclass(frozen=True, slots=True)
class RealworksBroaderAuditSource:
    source_id: str
    source_name: str
    domain: str
    listing_url: str
    gemeente: str
    province: str
    platform_guess: str
    delivery_mode: str
    parser_family_candidate: str
    access_policy_status: str
    selection_confidence: str
    selection_evidence_file: str
    selection_reason: str


@dataclass(frozen=True, slots=True)
class RealworksBroaderFamilyAuditDecision:
    sources_attempted: int
    sources_passed: int
    sources_passed_with_review_gaps: int
    sources_needing_hardening: int
    sources_blocked: int
    sources_failed: int
    total_properties_exported: int
    total_export_ready: int
    total_export_review: int
    total_export_blocked: int
    second_wave_sources_validated: int
    family_decision: str
    recommended_next_action: str


@dataclass(frozen=True, slots=True)
class RealworksBroaderBoundedAuditResult:
    sources: tuple[RealworksBroaderAuditSource, ...]
    metrics: tuple[RealworksSourceValidationMetrics, ...]
    readiness_results: tuple[RealworksPropertyReadinessResult, ...]
    family_decision: RealworksBroaderFamilyAuditDecision
    observed_at: datetime
    workbook_path: Path | None = None
    summary_csv_path: Path | None = None
    warnings: tuple[str, ...] = ()


def load_realworks_broader_audit_sources_from_local_evidence(
    *,
    repo_root: Path,
    max_sources: int = MAX_BROADER_AUDIT_SOURCES,
    second_wave_only: bool = True,
) -> tuple[RealworksBroaderAuditSource, ...]:
    rows: list[dict[str, Any]] = []
    evidence_paths = (
        repo_root / "data" / "processed" / "sources_seed_noord_brabant.csv",
        repo_root / "data" / "discovery" / "reference" / "property_discovery_source_overrides.csv",
        repo_root / "data" / "discovery" / "processed" / "sources_seed_with_gemeente.csv",
    )
    for path in evidence_paths:
        rows.extend(_read_csv_rows(path, evidence_file=str(path)))
    return select_realworks_broader_audit_sources(
        rows,
        max_sources=max_sources,
        second_wave_only=second_wave_only,
    )


def select_realworks_broader_audit_sources(
    evidence_rows: Iterable[Mapping[str, Any]],
    *,
    max_sources: int = MAX_BROADER_AUDIT_SOURCES,
    second_wave_only: bool = True,
) -> tuple[RealworksBroaderAuditSource, ...]:
    if max_sources <= 0:
        return ()

    candidates: dict[str, RealworksBroaderAuditSource] = {}
    for row in evidence_rows:
        source = broader_audit_source_from_evidence_row(row)
        if source is None:
            continue
        if second_wave_only and _is_control_source(source):
            continue
        candidates.setdefault(source.domain, source)

    selected = sorted(candidates.values(), key=lambda source: (source.selection_evidence_file, source.domain))
    return tuple(selected[: min(max_sources, MAX_BROADER_AUDIT_SOURCES)])


def broader_audit_source_from_evidence_row(row: Mapping[str, Any]) -> RealworksBroaderAuditSource | None:
    base = validation_source_from_evidence_row(row)
    if base is None:
        return None
    if not _looks_like_listing_index_url(base.listing_url):
        return None

    normalized = {_normalize_key(key): _clean_text(value) for key, value in row.items()}
    return RealworksBroaderAuditSource(
        source_id=base.source_id,
        source_name=base.source_name,
        domain=base.domain,
        listing_url=base.listing_url,
        gemeente=base.gemeente,
        province=_first_value(normalized, ("province", "provincie")),
        platform_guess=base.platform_guess,
        delivery_mode=base.delivery_mode,
        parser_family_candidate=base.parser_family_candidate,
        access_policy_status=base.access_policy_status,
        selection_confidence=_first_value(normalized, ("selection_confidence", "confidence_score", "confidence", "score")),
        selection_evidence_file=_first_value(normalized, ("evidence_file",)),
        selection_reason=base.selection_reason,
    )


def run_realworks_broader_bounded_audit(
    *,
    sources: Sequence[RealworksBroaderAuditSource],
    observed_at: datetime,
    output_dir: Path | None = None,
    max_sources: int = MAX_BROADER_AUDIT_SOURCES,
    max_listing_fetches_per_source: int = 1,
    max_detail_fetches_per_source: int = DEFAULT_MAX_DETAIL_FETCHES_PER_SOURCE,
    timeout_seconds: float = DEFAULT_TIMEOUT_SECONDS,
    fetch_html: Callable[[str], str] | None = None,
) -> RealworksBroaderBoundedAuditResult:
    observed = _to_utc(observed_at)
    capped_sources = tuple(sources[: max(0, min(max_sources, MAX_BROADER_AUDIT_SOURCES))])
    if len(capped_sources) < 2:
        decision = _broader_family_decision(())
        return RealworksBroaderBoundedAuditResult(
            sources=capped_sources,
            metrics=(),
            readiness_results=(),
            family_decision=decision,
            observed_at=observed,
            warnings=("insufficient_additional_sources",),
        )

    readiness_results: list[RealworksPropertyReadinessResult] = []
    metrics: list[RealworksSourceValidationMetrics] = []
    warnings: list[str] = []
    detail_cap = max(0, max_detail_fetches_per_source)

    for source in capped_sources:
        readiness = run_realworks_property_readiness(
            source_id=source.source_id,
            source_domain=source.domain,
            listing_url=source.listing_url,
            max_listing_fetches=max_listing_fetches_per_source,
            max_detail_fetches=detail_cap,
            timeout_seconds=timeout_seconds,
            fetch_html=fetch_html,
            observed_at=observed,
        )
        readiness_results.append(readiness)
        metrics.append(_metrics_for_source(source, readiness))
        warnings.extend(readiness.warnings)

    decision = _broader_family_decision(metrics)
    result = RealworksBroaderBoundedAuditResult(
        sources=capped_sources,
        metrics=tuple(metrics),
        readiness_results=tuple(readiness_results),
        family_decision=decision,
        observed_at=observed,
        warnings=_dedupe(warnings),
    )
    if output_dir is None:
        return result

    output_dir = Path(output_dir)
    workbook_path = output_dir / OUTPUT_WORKBOOK_NAME
    summary_csv_path = output_dir / OUTPUT_SUMMARY_NAME
    write_realworks_broader_audit_workbook(result, workbook_path)
    write_realworks_broader_audit_summary_csv(result, summary_csv_path)
    return RealworksBroaderBoundedAuditResult(
        sources=result.sources,
        metrics=result.metrics,
        readiness_results=result.readiness_results,
        family_decision=result.family_decision,
        observed_at=result.observed_at,
        workbook_path=workbook_path,
        summary_csv_path=summary_csv_path,
        warnings=result.warnings,
    )


def write_realworks_broader_audit_workbook(
    result: RealworksBroaderBoundedAuditResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Sources Summary"
    _write_sources_summary(workbook["Sources Summary"], result.sources, result.metrics)
    _write_manual_verification(workbook.create_sheet("Manual Verification"), result.sources, result.readiness_results)
    _write_field_gaps(workbook.create_sheet("Field Gaps"), result.readiness_results)
    _write_warnings(workbook.create_sheet("Warnings"), result.readiness_results)
    _write_problem_rows(workbook.create_sheet("Problem Rows"), result.readiness_results)
    _write_family_decision(workbook.create_sheet("Family Audit Decision"), result.family_decision)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def write_realworks_broader_audit_summary_csv(
    result: RealworksBroaderBoundedAuditResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = (
        "source_id",
        "source_name",
        "domain",
        "listing_url",
        "validation_status",
        "validation_decision",
        "parser_total",
        "parser_qa_clean",
        "detail_attempted",
        "detail_succeeded",
        "detail_failed",
        "readiness_rows_built",
        "export_ready",
        "export_review",
        "export_blocked",
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for metric in result.metrics:
            writer.writerow({column: _safe_cell(getattr(metric, column)) for column in columns})
    return output_path


def _broader_family_decision(
    metrics: Sequence[RealworksSourceValidationMetrics],
) -> RealworksBroaderFamilyAuditDecision:
    statuses = Counter(metric.validation_status for metric in metrics)
    passed = statuses["passed"]
    passed_with_gaps = statuses["passed_with_review_gaps"]
    validated = passed + passed_with_gaps
    total_exported = sum(metric.readiness_rows_built for metric in metrics)
    total_ready = sum(metric.export_ready for metric in metrics)
    total_review = sum(metric.export_review for metric in metrics)
    total_blocked = sum(metric.export_blocked for metric in metrics)

    if len(metrics) < 2:
        family_decision = "insufficient_additional_sources"
        next_action = "Find at least two additional locally evidenced Realworks sources before auditing."
    elif statuses["blocked"] and statuses["blocked"] == len(metrics):
        family_decision = "blocked_by_access_policy"
        next_action = "Do not proceed until access policy permits at least two additional sources."
    elif statuses["needs_hardening"] > 0 or statuses["failed"] > 0:
        family_decision = "realworks_needs_hardening_v2"
        next_action = "Harden the reusable Realworks family before Noord-Brabant-wide application."
    elif validated >= 2 and all(metric.parser_qa_clean > 0 for metric in metrics) and total_exported > 0:
        family_decision = "realworks_ready_for_noord_brabant_realworks_audit"
        next_action = "Proceed to a broader Noord-Brabant Realworks audit; do not start matching yet."
    else:
        family_decision = "insufficient_additional_sources"
        next_action = "Add stronger additional source evidence before deciding."

    return RealworksBroaderFamilyAuditDecision(
        sources_attempted=len(metrics),
        sources_passed=passed,
        sources_passed_with_review_gaps=passed_with_gaps,
        sources_needing_hardening=statuses["needs_hardening"],
        sources_blocked=statuses["blocked"],
        sources_failed=statuses["failed"],
        total_properties_exported=total_exported,
        total_export_ready=total_ready,
        total_export_review=total_review,
        total_export_blocked=total_blocked,
        second_wave_sources_validated=validated,
        family_decision=family_decision,
        recommended_next_action=next_action,
    )


def _metrics_for_source(
    source: RealworksBroaderAuditSource,
    readiness: RealworksPropertyReadinessResult,
) -> RealworksSourceValidationMetrics:
    rows = tuple(readiness.rows)
    quality_counts = Counter(row.quality_status for row in rows)
    export_counts = Counter(row.export_readiness for row in rows)
    active_counts = Counter(row.active_inventory_eligible for row in rows)
    validation_status = _validation_status(readiness)
    return RealworksSourceValidationMetrics(
        source_id=source.source_id,
        source_name=source.source_name,
        domain=source.domain,
        listing_url=source.listing_url,
        listing_fetch_status=_listing_fetch_status(readiness),
        robots_allowed_listing="listing_blocked_by_robots" not in readiness.warnings,
        parser_total=readiness.listing_parser_total,
        parser_qa_clean=readiness.listing_qa_clean,
        parser_qa_review=readiness.listing_qa_review,
        parser_qa_rejected=readiness.listing_qa_rejected,
        detail_attempted=readiness.detail_attempted,
        detail_succeeded=readiness.detail_succeeded,
        detail_failed=readiness.detail_failed,
        facts_records_built=readiness.facts_records_built,
        readiness_rows_built=readiness.readiness_rows_built,
        client_ready=quality_counts["client_ready"],
        advisor_review=quality_counts["advisor_review"],
        blocked=quality_counts["blocked"],
        export_ready=export_counts["export_ready"],
        export_review=export_counts["export_review"],
        export_blocked=export_counts["export_blocked"],
        active_inventory_eligible_true=active_counts[True],
        active_inventory_eligible_false=active_counts[False],
        db_persistence_action_counts=_counter_pairs(row.db_persistence_action for row in rows),
        postcode_usable=sum(1 for row in rows if row.postcode_status == "usable"),
        postcode_missing=sum(1 for row in rows if row.postcode_status == "missing"),
        vve_missing_for_apartment=sum(1 for row in rows if row.vve_status == "missing"),
        energy_label_usable=sum(1 for row in rows if row.energy_label_status == "usable"),
        energy_label_review_missing=sum(1 for row in rows if row.energy_label_status in {"review", "missing"}),
        non_residential_rows=sum(1 for row in rows if row.residential_classification.startswith("non_residential")),
        source_published_at_usable=sum(1 for row in rows if row.source_published_at_status == "usable"),
        source_published_at_review=sum(1 for row in rows if row.source_published_at_status == "review"),
        source_published_at_missing=sum(1 for row in rows if row.source_published_at_status == "missing"),
        freshness_bucket_counts=_counter_pairs(row.freshness_bucket for row in rows),
        lifecycle_event_counts=_counter_pairs(event for row in rows for event in row.lifecycle_events),
        top_warnings=_top_counter_pairs(row.warnings for row in rows),
        top_field_gaps=_top_field_gaps(rows),
        problem_row_count=len(readiness.problem_rows_compact),
        validation_status=validation_status,
        validation_decision=_validation_decision(validation_status),
    )


def _validation_status(readiness: RealworksPropertyReadinessResult) -> str:
    warnings = set(readiness.warnings)
    if "listing_blocked_by_robots" in warnings or "detail_blocked_by_robots" in warnings:
        return "blocked"
    if "listing_fetch_parse_or_qa_exception" in warnings:
        return "failed"
    if readiness.listing_parser_total <= 0 or readiness.listing_qa_clean <= 0:
        return "needs_hardening"
    if readiness.detail_attempted > 0 and readiness.detail_succeeded == 0:
        return "needs_hardening"
    if readiness.readiness_rows_built <= 0:
        return "needs_hardening"
    if readiness.detail_failed > readiness.detail_succeeded:
        return "needs_hardening"
    if _has_review_gaps(readiness.rows):
        return "passed_with_review_gaps"
    return "passed"


def _validation_decision(validation_status: str) -> str:
    return {
        "passed": "source passed without source-specific code",
        "passed_with_review_gaps": "source passed with explicit review gaps",
        "needs_hardening": "family needs hardening before broader Realworks audit",
        "blocked": "source blocked by access or robots policy",
        "failed": "technical failure unrelated to access policy",
    }.get(validation_status, "unknown validation state")


def _has_review_gaps(rows: Sequence[RealworksPropertyReadinessRow]) -> bool:
    return any(
        row.quality_status != "client_ready"
        or row.missing_key_fields
        or row.review_fields
        or row.vve_status in {"missing", "review"}
        or row.energy_label_status in {"missing", "review"}
        or row.source_published_at_status in {"missing", "review"}
        for row in rows
    )


def _listing_fetch_status(readiness: RealworksPropertyReadinessResult) -> str:
    if "listing_blocked_by_robots" in readiness.warnings:
        return "robots_blocked"
    if "listing_fetch_parse_or_qa_exception" in readiness.warnings:
        return "failed"
    if readiness.listing_parser_total > 0:
        return "success"
    return "empty"


def _write_sources_summary(
    worksheet: Worksheet,
    sources: Sequence[RealworksBroaderAuditSource],
    metrics: Sequence[RealworksSourceValidationMetrics],
) -> None:
    metric_by_source = {metric.source_id: metric for metric in metrics}
    columns = (
        "source_id",
        "source_name",
        "domain",
        "listing_url",
        "gemeente",
        "selection_reason",
        "validation_status",
        "validation_decision",
        "parser_total",
        "parser_qa_clean",
        "detail_attempted",
        "detail_succeeded",
        "detail_failed",
        "readiness_rows_built",
        "export_ready",
        "export_review",
        "export_blocked",
        "manual_review_priority",
    )
    worksheet.append(columns)
    for source in sources:
        metric = metric_by_source.get(source.source_id)
        values = {
            "source_id": source.source_id,
            "source_name": source.source_name,
            "domain": source.domain,
            "listing_url": source.listing_url,
            "gemeente": source.gemeente,
            "selection_reason": source.selection_reason,
            "validation_status": getattr(metric, "validation_status", "not_run"),
            "validation_decision": getattr(metric, "validation_decision", "not run"),
            "parser_total": getattr(metric, "parser_total", 0),
            "parser_qa_clean": getattr(metric, "parser_qa_clean", 0),
            "detail_attempted": getattr(metric, "detail_attempted", 0),
            "detail_succeeded": getattr(metric, "detail_succeeded", 0),
            "detail_failed": getattr(metric, "detail_failed", 0),
            "readiness_rows_built": getattr(metric, "readiness_rows_built", 0),
            "export_ready": getattr(metric, "export_ready", 0),
            "export_review": getattr(metric, "export_review", 0),
            "export_blocked": getattr(metric, "export_blocked", 0),
            "manual_review_priority": _manual_review_priority(metric),
        }
        worksheet.append([_safe_cell(values[column]) for column in columns])


def _write_manual_verification(
    worksheet: Worksheet,
    sources: Sequence[RealworksBroaderAuditSource],
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    columns = (
        "source_id",
        "source_name",
        "domain",
        "property_link",
        "canonical_url",
        "address",
        "postcode",
        "city",
        "asking_price",
        "source_status",
        "status_bucket",
        "availability",
        "property_type",
        "residential_classification",
        "living_area_m2",
        "plot_area_m2",
        "rooms",
        "bedrooms",
        "bathrooms",
        "energy_label",
        "energy_label_status",
        "energy_label_raw",
        "vve_active",
        "vve_monthly_cost",
        "vve_status",
        "source_published_at",
        "source_published_at_status",
        "first_seen_at",
        "observed_at",
        "freshness_bucket",
        "lifecycle_events",
        "quality_status",
        "export_readiness",
        "active_inventory_eligible",
        "db_persistence_action",
        "missing_key_fields",
        "review_fields",
        "warnings",
        "manual_check_result",
        "manual_check_notes",
    )
    source_by_id = {source.source_id: source for source in sources}
    worksheet.append(columns)
    for readiness in readiness_results:
        source = source_by_id.get(readiness.source_id)
        for row in readiness.rows:
            values = {
                "source_id": row.source_id,
                "source_name": source.source_name if source else row.source_id,
                "domain": source.domain if source else row.source_domain,
                "property_link": "Open listing" if _is_http_url(row.canonical_url) else "",
                "canonical_url": row.canonical_url,
                "address": row.address,
                "postcode": row.postcode,
                "city": row.city,
                "asking_price": row.asking_price,
                "source_status": row.source_status,
                "status_bucket": row.status_bucket,
                "availability": row.availability,
                "property_type": row.property_type,
                "residential_classification": row.residential_classification,
                "living_area_m2": row.living_area_m2,
                "plot_area_m2": row.plot_area_m2,
                "rooms": row.rooms,
                "bedrooms": row.bedrooms,
                "bathrooms": row.bathrooms,
                "energy_label": row.energy_label,
                "energy_label_status": row.energy_label_status,
                "energy_label_raw": row.energy_label_raw,
                "vve_active": row.vve_active,
                "vve_monthly_cost": row.vve_monthly_cost,
                "vve_status": row.vve_status,
                "source_published_at": _datetime_cell(row.source_published_at),
                "source_published_at_status": row.source_published_at_status,
                "first_seen_at": _datetime_cell(row.first_seen_at),
                "observed_at": _datetime_cell(row.observed_at),
                "freshness_bucket": row.freshness_bucket,
                "lifecycle_events": _join(row.lifecycle_events),
                "quality_status": row.quality_status,
                "export_readiness": row.export_readiness,
                "active_inventory_eligible": row.active_inventory_eligible,
                "db_persistence_action": row.db_persistence_action,
                "missing_key_fields": _join(row.missing_key_fields),
                "review_fields": _join(row.review_fields),
                "warnings": _join(row.warnings),
                "manual_check_result": "",
                "manual_check_notes": "",
            }
            worksheet.append([_safe_cell(values[column]) for column in columns])
            link_cell = worksheet.cell(worksheet.max_row, columns.index("property_link") + 1)
            if _is_http_url(row.canonical_url):
                link_cell.hyperlink = row.canonical_url
                link_cell.style = "Hyperlink"


def _write_field_gaps(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "field", "usable_count", "review_count", "missing_count"))
    for readiness in readiness_results:
        for field, usable, review, missing in readiness.field_completion_counts:
            worksheet.append((readiness.source_id, field, usable, review, missing))
        for field, count in _top_field_gaps(readiness.rows, limit=100):
            worksheet.append((readiness.source_id, field, 0, 0, count))


def _write_warnings(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "warning", "count"))
    for readiness in readiness_results:
        for warning, count in readiness.warning_counts:
            worksheet.append((readiness.source_id, warning, count))


def _write_problem_rows(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    worksheet.append(("source_id", "canonical_url", "property_link", "address", "quality_status", "export_readiness", "issues"))
    for readiness in readiness_results:
        for row in readiness.problem_rows_compact:
            issues = _join((*row.missing_key_fields, *row.review_fields, *row.warnings))
            worksheet.append(
                (
                    readiness.source_id,
                    _safe_cell(row.canonical_url),
                    "Open listing" if _is_http_url(row.canonical_url) else "",
                    _safe_cell(row.address),
                    row.quality_status,
                    row.export_readiness,
                    _safe_cell(issues),
                )
            )
            link_cell = worksheet.cell(worksheet.max_row, 3)
            if _is_http_url(row.canonical_url):
                link_cell.hyperlink = row.canonical_url
                link_cell.style = "Hyperlink"


def _write_family_decision(worksheet: Worksheet, decision: RealworksBroaderFamilyAuditDecision) -> None:
    worksheet.append(("metric", "value"))
    for field in RealworksBroaderFamilyAuditDecision.__dataclass_fields__:
        worksheet.append((field, _safe_cell(getattr(decision, field))))


def _read_csv_rows(path: Path, *, evidence_file: str) -> tuple[dict[str, Any], ...]:
    if not path.exists():
        return ()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return tuple({**row, "evidence_file": evidence_file} for row in csv.DictReader(handle))


def _is_control_source(source: RealworksBroaderAuditSource) -> bool:
    return source.source_id in CONTROL_SOURCE_IDS or source.domain in CONTROL_DOMAINS


def _manual_review_priority(metric: RealworksSourceValidationMetrics | None) -> str:
    if metric is None:
        return "not_run"
    if metric.validation_status in {"failed", "needs_hardening", "blocked"}:
        return "high"
    if metric.export_blocked or metric.export_review or metric.problem_row_count:
        return "medium"
    return "low"


def _top_field_gaps(rows: Sequence[RealworksPropertyReadinessRow], limit: int = 10) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for row in rows:
        counts.update(row.missing_key_fields)
        counts.update(f"review:{field}" for field in row.review_fields)
        if row.vve_status == "missing":
            counts["vve_active"] += 1
        if row.energy_label_status in {"review", "missing"}:
            counts["energy_label"] += 1
        if row.source_published_at_status in {"review", "missing"}:
            counts["source_published_at"] += 1
    return tuple(counts.most_common(limit))


def _top_counter_pairs(values: Iterable[Iterable[str]], limit: int = 10) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for items in values:
        counts.update(items)
    return tuple(counts.most_common(limit))


def _counter_pairs(values: Iterable[str]) -> tuple[tuple[str, int], ...]:
    return tuple(sorted(Counter(value for value in values if value).items()))


def _format_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[letter].width = max(10, min(max_length + 2, 60))


def _first_value(row: Mapping[str, str], keys: Iterable[str]) -> str:
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return ""


def _is_http_url(value: object) -> bool:
    parts = urlsplit(str(value or "").strip())
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def _looks_like_listing_index_url(value: object) -> bool:
    parts = urlsplit(str(value or "").strip())
    path = (parts.path or "").rstrip("/").casefold()
    if not path:
        return False
    if "/huis-" in path or "/appartement-" in path or "/woning-" in path:
        return False
    return "/aanbod/woningaanbod" in path or path.endswith("/woningaanbod")


def _normalize_key(value: object) -> str:
    return _clean_text(value).casefold().replace("-", "_").replace(" ", "_")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " ".join(_clean_text(item) for item in value if _clean_text(item))
    return str(value).strip()


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in RAW_MARKERS):
        return ""
    if len(text) > LONG_TEXT_LIMIT:
        return "[long_text_omitted]"
    return text


def _join(values: Iterable[object]) -> str:
    return "; ".join(str(_safe_cell(value)) for value in values if _safe_cell(value) not in (None, ""))


def _dedupe(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return tuple(result)


def _to_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def _datetime_cell(value: datetime | None) -> str:
    if value is None:
        return ""
    return _to_utc(value).isoformat().replace("+00:00", "Z")
