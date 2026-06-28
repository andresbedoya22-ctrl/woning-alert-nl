from __future__ import annotations

from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.facts.models import FACT_STATUS_REVIEW, FACT_STATUS_USABLE
from domek_wonen.pilots.kin_full_property_readiness import (
    KINPropertyReadinessRow,
    run_kin_full_property_readiness,
)


PROPERTY_COLUMNS = (
    "source_id",
    "source_domain",
    "canonical_url",
    "property_link",
    "listing_status",
    "address_raw",
    "postcode",
    "city",
    "gemeente",
    "province",
    "latitude",
    "longitude",
    "location_status",
    "location_source",
    "location_confidence",
    "location_warnings",
    "asking_price",
    "property_type",
    "living_area_m2",
    "plot_area_m2",
    "rooms",
    "bedrooms",
    "bathrooms",
    "floors",
    "volume_m3",
    "energy_label",
    "insulation",
    "eigendomssituatie",
    "vve_monthly_cost",
    "vve_active",
    "heating_type",
    "hot_water",
    "cv_ketel_brand",
    "cv_ketel_ownership",
    "garden",
    "main_garden_area_m2",
    "balcony",
    "storage",
    "garage",
    "garage_count",
    "parking",
    "availability_date",
    "open_huis_badge_or_event",
    "headline",
    "facts_line",
    "financial_line",
    "outdoor_line",
    "energy_line",
    "attention_points",
    "missing_key_fields",
    "warnings",
    "export_readiness",
    "quality_status",
)
FACT_FIELDS = (
    "asking_price",
    "property_type",
    "living_area_m2",
    "plot_area_m2",
    "rooms",
    "bedrooms",
    "bathrooms",
    "floors",
    "volume_m3",
    "energy_label",
    "insulation",
    "eigendomssituatie",
    "vve_monthly_cost",
    "vve_active",
    "heating_type",
    "hot_water",
    "cv_ketel_brand",
    "cv_ketel_ownership",
    "garden",
    "main_garden_area_m2",
    "balcony",
    "storage",
    "garage",
    "garage_count",
    "parking",
    "availability_date",
    "open_huis_badge_or_event",
)
LOCATION_GAP_FIELDS = ("postcode", "city", "address_raw", "latitude", "longitude")
WORKSHEET_NAMES = ("KIN Properties", "Summary", "Field Gaps", "Warnings", "Problem Rows")
_RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif", ".svg")


@dataclass(frozen=True, slots=True)
class KINExcelExportResult:
    output_path: Path
    rows_written: int
    worksheet_names: tuple[str, ...]
    export_ready_count: int
    export_review_count: int
    export_blocked_count: int
    client_ready_count: int
    advisor_review_count: int
    missing_field_counts: tuple[tuple[str, int], ...]
    warning_counts: tuple[tuple[str, int], ...]
    generated_at: str
    warnings: tuple[str, ...] = ()


def export_kin_readiness_rows_to_excel(
    *,
    rows: Sequence[KINPropertyReadinessRow],
    output_path: Path,
    source_id: str = "kin_makelaars",
    generated_at: datetime | None = None,
) -> KINExcelExportResult:
    output_path = Path(output_path)
    generated_at_text = _datetime_to_utc_iso(generated_at or datetime.now(UTC))
    rows = tuple(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    properties_ws = workbook.active
    properties_ws.title = "KIN Properties"
    _write_properties_sheet(properties_ws, rows)

    summary_ws = workbook.create_sheet("Summary")
    _write_summary_sheet(summary_ws, rows, source_id=source_id, generated_at=generated_at_text)

    gaps = _field_gaps(rows)
    gaps_ws = workbook.create_sheet("Field Gaps")
    _write_field_gaps_sheet(gaps_ws, gaps)

    warning_counts = _warning_counts(rows)
    warnings_ws = workbook.create_sheet("Warnings")
    _write_warnings_sheet(warnings_ws, warning_counts)

    problem_ws = workbook.create_sheet("Problem Rows")
    _write_problem_rows_sheet(problem_ws, rows)

    for worksheet in workbook.worksheets:
        _format_table_sheet(worksheet)

    workbook.save(output_path)

    missing_counts = tuple((gap.field, gap.missing_count) for gap in gaps if gap.missing_count)
    export_warnings = _export_warnings(rows)
    return KINExcelExportResult(
        output_path=output_path,
        rows_written=len(rows),
        worksheet_names=tuple(workbook.sheetnames),
        export_ready_count=sum(1 for row in rows if _row_export_readiness(row) == "export_ready"),
        export_review_count=sum(1 for row in rows if _row_export_readiness(row) == "export_review"),
        export_blocked_count=sum(1 for row in rows if _row_export_readiness(row) == "export_blocked"),
        client_ready_count=sum(1 for row in rows if row.quality_status == "client_ready"),
        advisor_review_count=sum(1 for row in rows if row.quality_status == "advisor_review"),
        missing_field_counts=missing_counts,
        warning_counts=warning_counts,
        generated_at=generated_at_text,
        warnings=export_warnings,
    )


def run_kin_excel_export(
    *,
    config_path: Path,
    cache_path: Path,
    output_path: Path,
    max_api_pages: int = 25,
    max_details: int = 300,
    max_runtime_seconds: float | None = None,
    force_refresh: bool = False,
) -> KINExcelExportResult:
    if cache_path is None:
        raise ValueError("cache_path_required")
    if output_path is None:
        raise ValueError("output_path_required")
    readiness = run_kin_full_property_readiness(
        config_path=config_path,
        cache_path=Path(cache_path),
        max_api_pages=max_api_pages,
        max_details=max_details,
        max_runtime_seconds=max_runtime_seconds,
        force_refresh=force_refresh,
    )
    return export_kin_readiness_rows_to_excel(
        rows=readiness.rows,
        output_path=Path(output_path),
        source_id=readiness.source_id,
    )


def _write_properties_sheet(worksheet: Worksheet, rows: Sequence[KINPropertyReadinessRow]) -> None:
    worksheet.append(PROPERTY_COLUMNS)
    for row in rows:
        values = _property_row_values(row)
        worksheet.append([values[column] for column in PROPERTY_COLUMNS])
        excel_row = worksheet.max_row
        link_cell = worksheet.cell(excel_row, PROPERTY_COLUMNS.index("property_link") + 1)
        if _is_valid_http_url(row.canonical_url):
            link_cell.hyperlink = row.canonical_url
            link_cell.style = "Hyperlink"


def _write_summary_sheet(
    worksheet: Worksheet,
    rows: Sequence[KINPropertyReadinessRow],
    *,
    source_id: str,
    generated_at: str,
) -> None:
    quality_counts = Counter(row.quality_status for row in rows)
    location_counts = Counter(row.location_status for row in rows)
    summary_rows = (
        ("generated_at", generated_at),
        ("source_id", source_id),
        ("rows_written", len(rows)),
        ("export_ready_count", sum(1 for row in rows if _row_export_readiness(row) == "export_ready")),
        ("export_review_count", sum(1 for row in rows if _row_export_readiness(row) == "export_review")),
        ("export_blocked_count", sum(1 for row in rows if _row_export_readiness(row) == "export_blocked")),
        ("client_ready_count", quality_counts["client_ready"]),
        ("advisor_review_count", quality_counts["advisor_review"]),
        ("insufficient_location_count", quality_counts["insufficient_location"]),
        ("insufficient_facts_count", quality_counts["insufficient_facts"]),
        ("location_usable_count", location_counts["usable"]),
        ("location_review_count", location_counts["review"]),
        ("location_missing_count", location_counts["missing"]),
    )
    worksheet.append(("metric", "value"))
    for item in summary_rows:
        worksheet.append(item)


def _write_field_gaps_sheet(worksheet: Worksheet, gaps: Sequence["_FieldGap"]) -> None:
    worksheet.append(("field", "missing_count", "review_count", "usable_count", "total_rows", "missing_rate", "review_rate"))
    for gap in gaps:
        worksheet.append(
            (
                gap.field,
                gap.missing_count,
                gap.review_count,
                gap.usable_count,
                gap.total_rows,
                gap.missing_rate,
                gap.review_rate,
            )
        )


def _write_warnings_sheet(worksheet: Worksheet, warning_counts: Sequence[tuple[str, int]]) -> None:
    worksheet.append(("warning", "count"))
    for warning, count in warning_counts:
        worksheet.append((warning, count))


def _write_problem_rows_sheet(worksheet: Worksheet, rows: Sequence[KINPropertyReadinessRow]) -> None:
    columns = ("canonical_url", "property_link", "address_raw", "missing_key_fields", "attention_points", "warnings")
    worksheet.append(columns)
    for row in rows:
        if _row_export_readiness(row) == "export_ready" and not row.warnings and not row.missing_key_fields and not row.attention_points:
            continue
        worksheet.append(
            (
                _safe_cell(row.canonical_url),
                "Open listing" if _is_valid_http_url(row.canonical_url) else "",
                _safe_cell(row.address_raw),
                _join(row.missing_key_fields),
                _join(row.attention_points),
                _join(_row_warnings(row)),
            )
        )
        link_cell = worksheet.cell(worksheet.max_row, 2)
        if _is_valid_http_url(row.canonical_url):
            link_cell.hyperlink = row.canonical_url
            link_cell.style = "Hyperlink"


def _property_row_values(row: KINPropertyReadinessRow) -> dict[str, object]:
    facts = {fact.field: fact for fact in row.facts_record.facts}
    values = {
        "source_id": row.source_id,
        "source_domain": row.source_domain,
        "canonical_url": row.canonical_url,
        "property_link": "Open listing" if _is_valid_http_url(row.canonical_url) else "",
        "listing_status": row.listing_status,
        "address_raw": row.address_raw,
        "postcode": row.postcode,
        "city": row.city,
        "gemeente": row.gemeente,
        "province": row.province,
        "latitude": row.latitude,
        "longitude": row.longitude,
        "location_status": row.location_status,
        "location_source": "parsed_listing",
        "location_confidence": _location_confidence(row),
        "location_warnings": _join(_location_warnings(row)),
        "headline": row.summary.headline,
        "facts_line": row.summary.facts_line,
        "financial_line": row.summary.financial_line,
        "outdoor_line": row.summary.outdoor_line,
        "energy_line": row.summary.energy_line,
        "attention_points": _join(row.attention_points),
        "missing_key_fields": _join(row.missing_key_fields),
        "warnings": _join(_row_warnings(row)),
        "export_readiness": _row_export_readiness(row),
        "quality_status": row.quality_status,
    }
    for field in FACT_FIELDS:
        fact = facts.get(field)
        values[field] = _fact_export_value(fact)
    return {key: _safe_cell(value) for key, value in values.items()}


def _fact_export_value(fact: object) -> object:
    if fact is None:
        return ""
    value = getattr(fact, "normalized_value", None)
    if value is None:
        value = getattr(fact, "value", None)
    return value


@dataclass(frozen=True, slots=True)
class _FieldGap:
    field: str
    missing_count: int
    review_count: int
    usable_count: int
    total_rows: int
    missing_rate: float
    review_rate: float


def _field_gaps(rows: Sequence[KINPropertyReadinessRow]) -> tuple[_FieldGap, ...]:
    return tuple(_fact_gap(field, rows) for field in FACT_FIELDS) + tuple(_location_gap(field, rows) for field in LOCATION_GAP_FIELDS)


def _fact_gap(field: str, rows: Sequence[KINPropertyReadinessRow]) -> _FieldGap:
    usable = 0
    review = 0
    missing = 0
    for row in rows:
        statuses = {fact.field: fact.status for fact in row.facts_record.facts}
        if statuses.get(field) == FACT_STATUS_USABLE:
            usable += 1
        elif statuses.get(field) == FACT_STATUS_REVIEW:
            review += 1
        else:
            missing += 1
    return _gap(field, usable=usable, review=review, missing=missing, total=len(rows))


def _location_gap(field: str, rows: Sequence[KINPropertyReadinessRow]) -> _FieldGap:
    usable = 0
    missing = 0
    for row in rows:
        value = getattr(row, field)
        if value in (None, ""):
            missing += 1
        else:
            usable += 1
    return _gap(field, usable=usable, review=0, missing=missing, total=len(rows))


def _gap(field: str, *, usable: int, review: int, missing: int, total: int) -> _FieldGap:
    return _FieldGap(
        field=field,
        missing_count=missing,
        review_count=review,
        usable_count=usable,
        total_rows=total,
        missing_rate=_rate(missing, total),
        review_rate=_rate(review, total),
    )


def _warning_counts(rows: Sequence[KINPropertyReadinessRow]) -> tuple[tuple[str, int], ...]:
    return tuple(sorted(Counter(warning for row in rows for warning in _row_warnings(row)).items()))


def _row_warnings(row: KINPropertyReadinessRow) -> tuple[str, ...]:
    warnings = list(row.warnings)
    if not _is_valid_http_url(row.canonical_url):
        warnings.append("missing_or_invalid_canonical_url")
    return _dedupe(warnings)


def _export_warnings(rows: Sequence[KINPropertyReadinessRow]) -> tuple[str, ...]:
    return _dedupe(warning for row in rows for warning in _row_warnings(row))


def _row_export_readiness(row: KINPropertyReadinessRow) -> str:
    if not _is_valid_http_url(row.canonical_url):
        return "export_blocked"
    return row.export_readiness


def _location_warnings(row: KINPropertyReadinessRow) -> tuple[str, ...]:
    return tuple(warning for warning in row.warnings if warning.startswith("missing_") or warning == "partial_location")


def _location_confidence(row: KINPropertyReadinessRow) -> float:
    if row.location_status == "usable":
        return 0.95 if row.latitude is not None and row.longitude is not None else 0.85
    if row.location_status == "review":
        return 0.65 if row.address_raw and row.city else 0.45
    return 0.0


def _format_table_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, 48))


def _join(values: Iterable[object]) -> str:
    return "; ".join(_safe_cell(value) for value in values if _safe_cell(value))


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in _RAW_MARKERS):
        return ""
    if _contains_image_url(text):
        return ""
    return text


def _contains_image_url(text: str) -> bool:
    for token in text.replace(";", " ").replace(",", " ").split():
        lowered = token.casefold().split("?", 1)[0]
        if lowered.startswith(("http://", "https://")) and lowered.endswith(_IMAGE_EXTENSIONS):
            return True
    return False


def _is_valid_http_url(value: object) -> bool:
    parts = urlsplit(str(value or "").strip())
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def _rate(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round(numerator / denominator, 4)


def _dedupe(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return tuple(result)


def _datetime_to_utc_iso(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")
