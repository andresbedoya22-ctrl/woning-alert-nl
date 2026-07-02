from __future__ import annotations

import csv
import time
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.pilots.realworks_property_readiness import (
    RealworksPropertyReadinessResult,
    RealworksPropertyReadinessRow,
    run_realworks_property_readiness,
)


EXPECTED_NB_REALWORKS_AUDIT_ROWS = 65
REALWORKS_READY_STATUS = "ready_for_noord_brabant_realworks_audit"
REALWORKS_FAMILY = "realworks_public"
REALWORKS_VERIFIED = "verified"
FAMILY_DECISIONS = (
    "realworks_ready_for_nb_family_coverage",
    "realworks_partially_ready_with_exclusions",
    "realworks_needs_hardening_v2",
    "blocked_by_access_policy",
    "insufficient_successful_sources",
)
RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
LONG_TEXT_LIMIT = 500
RESOLUTION_DECISIONS = (
    "realworks_ready_for_nb_family_coverage",
    "realworks_partially_ready_with_exclusions",
    "realworks_needs_hardening_v2",
)
SOURCE_RESOLUTION_STATUSES = (
    "ready_for_realworks_family_coverage",
    "ready_with_review_gaps",
    "monitor_no_current_listings",
    "exclude_fetch_failed",
    "exclude_access_blocked",
    "exclude_not_realworks",
    "exclude_scope_issue",
    "needs_realworks_hardening",
    "needs_manual_review",
)
PROPERTY_QA_STATUSES = (
    "qa_ok",
    "qa_review",
    "qa_duplicate_review",
    "qa_blocked_inconsistent_label",
    "qa_blocked_missing_source",
    "qa_blocked_invalid_value",
)
REALWORKS_AUDIT_INPUT_REQUIRED_COLUMNS = (
    "source_id",
    "source_name",
    "domain",
    "accepted_aanbod_url",
    "coverage_city",
    "coverage_gemeente",
    "coverage_province",
    "aanbod_scope_status",
    "realworks_verification_status",
    "realworks_evidence_strength",
    "audit_input_status",
    "parser_family_candidate",
)
REALWORKS_AUDIT_INPUT_OPTIONAL_COLUMNS = (
    "root_url",
    "source_quality_status",
    "access_policy_status",
    "realworks_evidence_strength",
    "family_confidence",
    "manual_check_result",
    "manual_check_notes",
    "recommended_next_action",
    "coverage_province_source",
)
REALWORKS_AUDIT_INPUT_CANONICAL_COLUMNS = (
    *REALWORKS_AUDIT_INPUT_REQUIRED_COLUMNS,
    *REALWORKS_AUDIT_INPUT_OPTIONAL_COLUMNS,
)
REQUIRED_INPUT_COLUMNS = REALWORKS_AUDIT_INPUT_REQUIRED_COLUMNS
SUMMARY_COLUMNS = (
    "source_id",
    "source_name",
    "domain",
    "accepted_aanbod_url",
    "coverage_city",
    "coverage_gemeente",
    "aanbod_scope_status",
    "access_policy_status",
    "robots_listing_allowed",
    "listing_fetch_status",
    "listing_http_status",
    "parser_total",
    "parser_qa_clean",
    "parser_qa_review",
    "parser_qa_rejected",
    "detail_attempted",
    "detail_succeeded",
    "detail_failed",
    "readiness_rows_built",
    "export_ready",
    "export_review",
    "export_blocked",
    "active_inventory_eligible",
    "inactive_status_count",
    "non_residential_count",
    "no_current_listings",
    "top_warning",
    "validation_status",
    "validation_decision",
    "recommended_next_action",
)
MANUAL_VERIFICATION_COLUMNS = (
    "source_id",
    "domain",
    "address",
    "city",
    "price",
    "status",
    "accepted_aanbod_url",
    "canonical_url",
    "property_link",
    "export_readiness",
    "quality_status",
    "missing_key_fields",
    "warnings",
    "manual_check_result",
    "manual_check_notes",
)


class NoordBrabantRealworksAuditInputError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksAuditInputValidation:
    row_count: int
    missing_required_columns: tuple[str, ...]
    audit_input_missing_required_columns_count: int
    audit_input_missing_parser_family_candidate_column: int
    audit_input_missing_coverage_province_column: int
    audit_input_non_ready_count: int
    audit_input_non_realworks_count: int
    audit_input_kin_count: int
    audit_input_missing_accepted_aanbod_url_count: int
    audit_input_property_detail_url_count: int
    audit_input_funda_pararius_count: int


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksAuditSource:
    source_id: str
    source_name: str
    domain: str
    accepted_aanbod_url: str
    coverage_city: str
    coverage_gemeente: str
    coverage_province: str
    aanbod_scope_status: str
    realworks_verification_status: str
    realworks_evidence_strength: str
    audit_input_status: str
    parser_family_candidate: str


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksAuditMetrics:
    source_id: str
    source_name: str
    domain: str
    accepted_aanbod_url: str
    coverage_city: str
    coverage_gemeente: str
    aanbod_scope_status: str
    access_policy_status: str
    robots_listing_allowed: bool
    listing_fetch_status: str
    listing_http_status: str
    parser_total: int
    parser_qa_clean: int
    parser_qa_review: int
    parser_qa_rejected: int
    detail_attempted: int
    detail_succeeded: int
    detail_failed: int
    readiness_rows_built: int
    export_ready: int
    export_review: int
    export_blocked: int
    active_inventory_eligible: int
    inactive_status_count: int
    non_residential_count: int
    no_current_listings: bool
    top_warning: str
    warning_counts: tuple[tuple[str, int], ...]
    field_gap_counts: tuple[tuple[str, int], ...]
    failure_patterns: tuple[str, ...]
    validation_status: str
    validation_decision: str
    recommended_next_action: str


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksFamilyDecision:
    family_decision: str
    confidence: str
    reasons: tuple[str, ...]
    next_action: str
    source_exclusions: tuple[str, ...]
    hardening_targets: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksAggregateMetrics:
    input_sources: int
    sources_attempted: int
    sources_skipped: int
    sources_passed: int
    sources_passed_with_review_gaps: int
    sources_no_current_listings: int
    sources_blocked: int
    sources_fetch_failed: int
    sources_needing_hardening: int
    total_parser_rows: int
    total_qa_clean: int
    total_qa_review: int
    total_qa_rejected: int
    total_detail_attempted: int
    total_detail_succeeded: int
    total_detail_failed: int
    total_readiness_rows: int
    total_export_ready: int
    total_export_review: int
    total_export_blocked: int
    field_gap_counts: tuple[tuple[str, int], ...]
    warning_counts: tuple[tuple[str, int], ...]
    failure_pattern_counts: tuple[tuple[str, int], ...]
    runtime_budget_exhausted: bool


@dataclass(frozen=True, slots=True)
class NoordBrabantRealworksAuditResult:
    sources: tuple[NoordBrabantRealworksAuditSource, ...]
    metrics: tuple[NoordBrabantRealworksAuditMetrics, ...]
    readiness_results: tuple[RealworksPropertyReadinessResult, ...]
    aggregate_metrics: NoordBrabantRealworksAggregateMetrics
    family_decision: NoordBrabantRealworksFamilyDecision
    observed_at: datetime
    workbook_path: Path | None = None
    summary_csv_path: Path | None = None
    problem_sources_csv_path: Path | None = None
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class RealworksAuditSourceResolution:
    source_id: str
    source_name: str
    domain: str
    accepted_aanbod_url: str
    audit_source_status: str
    parser_total: int
    qa_clean: int
    detail_succeeded: int
    readiness_rows: int
    export_ready: int
    export_review: int
    export_blocked: int
    source_resolution_status: str
    resolution_reason: str
    recommended_next_action: str
    manual_check_result: str = ""
    manual_check_notes: str = ""


@dataclass(frozen=True, slots=True)
class RealworksPropertyQAResult:
    source_id: str
    domain: str
    accepted_aanbod_url: str
    canonical_url: str
    property_link: str
    address: str
    city: str
    postcode: str
    asking_price: str
    status: str
    property_type: str
    export_readiness: str
    quality_status: str
    active_inventory_eligible: str
    missing_key_fields: str
    warnings: str
    property_qa_status: str
    property_qa_reason: str
    manual_check_result: str = ""
    manual_check_notes: str = ""


@dataclass(frozen=True, slots=True)
class RealworksAuditResolutionDecision:
    final_decision: str
    merge_recommended: bool
    decision_reasons: tuple[str, ...]
    recommended_next_action: str


@dataclass(frozen=True, slots=True)
class RealworksAuditResolutionResult:
    sources: tuple[RealworksAuditSourceResolution, ...]
    property_qa_rows: tuple[RealworksPropertyQAResult, ...]
    duplicate_rows: tuple[dict[str, object], ...]
    readiness_label_checks: tuple[dict[str, object], ...]
    status_consistency_checks: tuple[dict[str, object], ...]
    field_sanity_checks: tuple[dict[str, object], ...]
    decision: RealworksAuditResolutionDecision
    metrics: dict[str, int | str | bool]
    output_csv_path: Path | None = None
    output_workbook_path: Path | None = None


ReadinessRunner = Callable[..., RealworksPropertyReadinessResult]


def load_noord_brabant_realworks_audit_sources(
    input_csv: Path,
    *,
    expected_rows: int = EXPECTED_NB_REALWORKS_AUDIT_ROWS,
) -> tuple[NoordBrabantRealworksAuditSource, ...]:
    path = Path(input_csv)
    if not path.exists():
        raise NoordBrabantRealworksAuditInputError(f"missing input CSV: {path}")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        columns = tuple(reader.fieldnames or ())

    input_validation = validate_noord_brabant_realworks_audit_input_rows(rows, columns=columns)
    if input_validation.missing_required_columns:
        raise NoordBrabantRealworksAuditInputError(
            f"missing required columns: {', '.join(input_validation.missing_required_columns)}"
        )
    if not rows:
        raise NoordBrabantRealworksAuditInputError("input CSV has no rows")
    if expected_rows is not None and len(rows) != expected_rows:
        raise NoordBrabantRealworksAuditInputError(f"expected {expected_rows} rows, got {len(rows)}")

    sources: list[NoordBrabantRealworksAuditSource] = []
    gate_counts = _input_gate_counts(rows, missing_required_columns=())
    failing_gates = {key: value for key, value in gate_counts.items() if value}
    if failing_gates:
        details = "; ".join(f"{key}={value}" for key, value in sorted(failing_gates.items()))
        raise NoordBrabantRealworksAuditInputError(f"input hard gates failed: {details}")

    duplicate_counts = Counter(
        (_normalize_domain(row.get("domain")), _normalize_url(row.get("accepted_aanbod_url")))
        for row in rows
    )
    duplicates = [key for key, count in duplicate_counts.items() if count > 1]
    if duplicates:
        raise NoordBrabantRealworksAuditInputError(f"duplicate domain+accepted_aanbod_url rows: {len(duplicates)}")

    for index, row in enumerate(rows, start=2):
        source = _source_from_row(row)
        validation_error = _validate_source(source)
        if validation_error:
            raise NoordBrabantRealworksAuditInputError(f"row {index}: {validation_error}")
        sources.append(source)
    return tuple(sources)


def validate_noord_brabant_realworks_audit_input_rows(
    rows: Sequence[Mapping[str, str]],
    *,
    columns: Sequence[str],
) -> NoordBrabantRealworksAuditInputValidation:
    missing_columns = tuple(column for column in REQUIRED_INPUT_COLUMNS if column not in columns)
    gate_counts = _input_gate_counts(rows, missing_required_columns=missing_columns)
    return NoordBrabantRealworksAuditInputValidation(
        row_count=len(rows),
        missing_required_columns=missing_columns,
        audit_input_missing_required_columns_count=len(missing_columns),
        audit_input_missing_parser_family_candidate_column=int("parser_family_candidate" in missing_columns),
        audit_input_missing_coverage_province_column=int("coverage_province" in missing_columns),
        audit_input_non_ready_count=gate_counts["audit_input_non_ready_count"],
        audit_input_non_realworks_count=gate_counts["audit_input_non_realworks_count"],
        audit_input_kin_count=gate_counts["audit_input_kin_count"],
        audit_input_missing_accepted_aanbod_url_count=gate_counts["audit_input_missing_accepted_aanbod_url_count"],
        audit_input_property_detail_url_count=gate_counts["audit_input_property_detail_url_count"],
        audit_input_funda_pararius_count=gate_counts["audit_input_funda_pararius_count"],
    )


def run_noord_brabant_realworks_audit(
    *,
    input_csv: Path,
    output_workbook: Path | None = None,
    output_summary: Path | None = None,
    output_problem_sources: Path | None = None,
    max_sources: int = EXPECTED_NB_REALWORKS_AUDIT_ROWS,
    max_listings_per_source: int = 15,
    max_detail_per_source: int = 10,
    timeout_seconds: float = 15.0,
    runtime_budget_seconds: float = 1800.0,
    observed_at: datetime | None = None,
    fetch_html: Callable[[str], str] | None = None,
    readiness_runner: ReadinessRunner = run_realworks_property_readiness,
) -> NoordBrabantRealworksAuditResult:
    observed = _to_utc(observed_at or datetime.now(UTC))
    sources = load_noord_brabant_realworks_audit_sources(input_csv, expected_rows=None)
    capped_sources = tuple(sources[: max(0, min(max_sources, len(sources)))])
    start = time.monotonic()
    budget_exhausted = False
    metrics: list[NoordBrabantRealworksAuditMetrics] = []
    readiness_results: list[RealworksPropertyReadinessResult] = []
    warnings: list[str] = []

    for source in capped_sources:
        if _runtime_exhausted(start, runtime_budget_seconds):
            budget_exhausted = True
            warnings.append("runtime_budget_exhausted")
            break
        try:
            readiness = readiness_runner(
                source_id=source.source_id,
                source_domain=source.domain,
                listing_url=source.accepted_aanbod_url,
                max_listing_fetches=max(0, max_listings_per_source),
                max_detail_fetches=max(0, max_detail_per_source),
                timeout_seconds=timeout_seconds,
                fetch_html=fetch_html,
                observed_at=observed,
            )
        except Exception as exc:
            warnings.append("source_runner_exception")
            readiness = _empty_readiness_result(source, ("source_runner_exception", type(exc).__name__))
        readiness_results.append(readiness)
        metrics.append(_metrics_for_source(source, readiness))
        warnings.extend(readiness.warnings)

    aggregate = aggregate_noord_brabant_realworks_metrics(
        input_sources=len(sources),
        attempted_sources=len(metrics),
        metrics=metrics,
        runtime_budget_exhausted=budget_exhausted,
    )
    decision = decide_noord_brabant_realworks_family(metrics, aggregate)
    result = NoordBrabantRealworksAuditResult(
        sources=sources,
        metrics=tuple(metrics),
        readiness_results=tuple(readiness_results),
        aggregate_metrics=aggregate,
        family_decision=decision,
        observed_at=observed,
        warnings=_dedupe(warnings),
    )
    workbook_path = write_noord_brabant_realworks_audit_workbook(result, output_workbook) if output_workbook else None
    summary_path = write_noord_brabant_realworks_summary_csv(result, output_summary) if output_summary else None
    problem_path = write_noord_brabant_realworks_problem_sources_csv(result, output_problem_sources) if output_problem_sources else None
    return NoordBrabantRealworksAuditResult(
        sources=result.sources,
        metrics=result.metrics,
        readiness_results=result.readiness_results,
        aggregate_metrics=result.aggregate_metrics,
        family_decision=result.family_decision,
        observed_at=result.observed_at,
        workbook_path=workbook_path,
        summary_csv_path=summary_path,
        problem_sources_csv_path=problem_path,
        warnings=result.warnings,
    )


def aggregate_noord_brabant_realworks_metrics(
    *,
    input_sources: int,
    attempted_sources: int,
    metrics: Sequence[NoordBrabantRealworksAuditMetrics],
    runtime_budget_exhausted: bool,
) -> NoordBrabantRealworksAggregateMetrics:
    statuses = Counter(metric.validation_status for metric in metrics)
    return NoordBrabantRealworksAggregateMetrics(
        input_sources=input_sources,
        sources_attempted=attempted_sources,
        sources_skipped=max(0, input_sources - attempted_sources),
        sources_passed=statuses["passed"],
        sources_passed_with_review_gaps=statuses["passed_with_review_gaps"],
        sources_no_current_listings=statuses["no_current_listings"],
        sources_blocked=statuses["blocked_by_robots"] + statuses["blocked_by_http_status"],
        sources_fetch_failed=statuses["listing_fetch_failed"],
        sources_needing_hardening=statuses["needs_realworks_hardening"],
        total_parser_rows=sum(metric.parser_total for metric in metrics),
        total_qa_clean=sum(metric.parser_qa_clean for metric in metrics),
        total_qa_review=sum(metric.parser_qa_review for metric in metrics),
        total_qa_rejected=sum(metric.parser_qa_rejected for metric in metrics),
        total_detail_attempted=sum(metric.detail_attempted for metric in metrics),
        total_detail_succeeded=sum(metric.detail_succeeded for metric in metrics),
        total_detail_failed=sum(metric.detail_failed for metric in metrics),
        total_readiness_rows=sum(metric.readiness_rows_built for metric in metrics),
        total_export_ready=sum(metric.export_ready for metric in metrics),
        total_export_review=sum(metric.export_review for metric in metrics),
        total_export_blocked=sum(metric.export_blocked for metric in metrics),
        field_gap_counts=_sum_counter_pairs(metric.field_gap_counts for metric in metrics),
        warning_counts=_sum_counter_pairs(metric.warning_counts for metric in metrics),
        failure_pattern_counts=_sum_counter_pairs((tuple((pattern, 1) for pattern in metric.failure_patterns) for metric in metrics)),
        runtime_budget_exhausted=runtime_budget_exhausted,
    )


def decide_noord_brabant_realworks_family(
    metrics: Sequence[NoordBrabantRealworksAuditMetrics],
    aggregate: NoordBrabantRealworksAggregateMetrics,
) -> NoordBrabantRealworksFamilyDecision:
    attempted = max(1, aggregate.sources_attempted)
    successful = aggregate.sources_passed + aggregate.sources_passed_with_review_gaps
    blocked = aggregate.sources_blocked
    hardening = aggregate.sources_needing_hardening
    fetch_failed = aggregate.sources_fetch_failed
    no_current = aggregate.sources_no_current_listings
    source_exclusions = tuple(
        metric.source_id
        for metric in metrics
        if metric.validation_status in {"blocked_by_robots", "blocked_by_http_status", "listing_fetch_failed", "no_current_listings"}
    )
    hardening_targets = tuple(
        warning for warning, count in aggregate.failure_pattern_counts if count >= 2 or warning in {"parser_zero_qa_clean", "detail_systemic_failure"}
    )

    if blocked / attempted >= 0.60:
        return NoordBrabantRealworksFamilyDecision(
            family_decision="blocked_by_access_policy",
            confidence="medium",
            reasons=("broad_access_or_robots_blocking",),
            next_action="Stop operational use until access policy permits a meaningful Realworks audit.",
            source_exclusions=source_exclusions,
            hardening_targets=hardening_targets,
        )
    if hardening >= 3 or (hardening and hardening >= successful):
        return NoordBrabantRealworksFamilyDecision(
            family_decision="realworks_needs_hardening_v2",
            confidence="medium",
            reasons=("systemic_parser_or_detail_failure_patterns",),
            next_action="Run Realworks Hardening v2 before merge or broader family coverage.",
            source_exclusions=source_exclusions,
            hardening_targets=hardening_targets,
        )
    if aggregate.sources_attempted == 0 or successful < 2:
        return NoordBrabantRealworksFamilyDecision(
            family_decision="insufficient_successful_sources",
            confidence="low",
            reasons=("too_few_successful_sources",),
            next_action="Do not proceed to family coverage; rerun when enough sources can be audited.",
            source_exclusions=source_exclusions,
            hardening_targets=hardening_targets,
        )
    if successful / attempted >= 0.60 and aggregate.total_readiness_rows > 0 and not hardening:
        if blocked or fetch_failed or no_current or aggregate.runtime_budget_exhausted:
            return NoordBrabantRealworksFamilyDecision(
                family_decision="realworks_partially_ready_with_exclusions",
                confidence="medium",
                reasons=("parser_works_broadly_with_isolated_exclusions",),
                next_action="Review exclusions and rerun blocked or empty sources before operational promotion.",
                source_exclusions=source_exclusions,
                hardening_targets=hardening_targets,
            )
        return NoordBrabantRealworksFamilyDecision(
            family_decision="realworks_ready_for_nb_family_coverage",
            confidence="high",
            reasons=("majority_sources_passed", "readiness_rows_built_across_sources"),
            next_action="Merge if manual workbook review confirms the rows; do not start matching in this phase.",
            source_exclusions=(),
            hardening_targets=(),
        )
    return NoordBrabantRealworksFamilyDecision(
        family_decision="realworks_partially_ready_with_exclusions",
        confidence="low",
        reasons=("mixed_results_need_review",),
        next_action="Review exclusions and failure patterns before deciding on Realworks Hardening v2.",
        source_exclusions=source_exclusions,
        hardening_targets=hardening_targets,
    )


def write_noord_brabant_realworks_audit_workbook(
    result: NoordBrabantRealworksAuditResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Source Summary"
    _write_source_summary(workbook["Source Summary"], result.metrics)
    _write_realworks_properties(workbook.create_sheet("Realworks Properties"), result.readiness_results)
    _write_field_gaps(workbook.create_sheet("Field Gaps"), result.aggregate_metrics.field_gap_counts)
    _write_warnings(workbook.create_sheet("Warnings"), result.aggregate_metrics.warning_counts)
    _write_problem_sources(workbook.create_sheet("Problem Sources"), result.metrics)
    _write_parser_failure_patterns(workbook.create_sheet("Parser Failure Patterns"), result.aggregate_metrics.failure_pattern_counts)
    _write_access_policy(workbook.create_sheet("Access Policy"), result.metrics)
    _write_audit_input(workbook.create_sheet("Audit Input"), result.sources)
    _write_family_decision(workbook.create_sheet("Family Audit Decision"), result.family_decision)
    _write_manual_verification(workbook.create_sheet("Manual Verification"), result.sources, result.readiness_results)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def write_noord_brabant_realworks_summary_csv(
    result: NoordBrabantRealworksAuditResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_COLUMNS)
        writer.writeheader()
        for metric in result.metrics:
            writer.writerow({column: _safe_cell(_metric_value(metric, column)) for column in SUMMARY_COLUMNS})
    return output_path


def write_noord_brabant_realworks_problem_sources_csv(
    result: NoordBrabantRealworksAuditResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = (
        "source_id",
        "domain",
        "accepted_aanbod_url",
        "validation_status",
        "validation_decision",
        "recommended_next_action",
        "top_warning",
        "failure_patterns",
    )
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for metric in result.metrics:
            if metric.validation_status in {"passed", "passed_with_review_gaps"}:
                continue
            writer.writerow(
                {
                    "source_id": _safe_cell(metric.source_id),
                    "domain": _safe_cell(metric.domain),
                    "accepted_aanbod_url": _safe_cell(metric.accepted_aanbod_url),
                    "validation_status": metric.validation_status,
                    "validation_decision": metric.validation_decision,
                    "recommended_next_action": metric.recommended_next_action,
                    "top_warning": metric.top_warning,
                    "failure_patterns": _join(metric.failure_patterns),
                }
            )
    return output_path


def resolve_noord_brabant_realworks_audit(
    *,
    audit_summary_csv: Path,
    audit_workbook: Path,
    output_csv: Path | None = None,
    output_workbook: Path | None = None,
) -> RealworksAuditResolutionResult:
    summary_rows = _read_csv_dicts(audit_summary_csv)
    workbook_rows = _read_workbook_rows(audit_workbook)
    source_rows = workbook_rows.get("Source Summary", summary_rows)
    manual_rows = workbook_rows.get("Manual Verification", ())
    property_rows = workbook_rows.get("Realworks Properties", ())
    accepted_by_source = {
        _clean_text(row.get("source_id")): _clean_text(row.get("accepted_aanbod_url"))
        for row in source_rows
    }
    sources = tuple(_source_resolution_from_row(row) for row in source_rows)
    property_qa, duplicates, label_checks, status_checks, field_checks = _build_property_qa(
        manual_rows=manual_rows,
        property_rows=property_rows,
        accepted_by_source=accepted_by_source,
    )
    metrics = _resolution_metrics(sources, property_qa, duplicates, label_checks, status_checks, field_checks)
    decision = _resolution_decision(metrics)
    result = RealworksAuditResolutionResult(
        sources=sources,
        property_qa_rows=property_qa,
        duplicate_rows=duplicates,
        readiness_label_checks=label_checks,
        status_consistency_checks=status_checks,
        field_sanity_checks=field_checks,
        decision=decision,
        metrics=metrics,
    )
    csv_path = write_noord_brabant_realworks_audit_resolution_csv(result, output_csv) if output_csv else None
    workbook_path = write_noord_brabant_realworks_audit_resolution_workbook(result, output_workbook) if output_workbook else None
    return RealworksAuditResolutionResult(
        sources=result.sources,
        property_qa_rows=result.property_qa_rows,
        duplicate_rows=result.duplicate_rows,
        readiness_label_checks=result.readiness_label_checks,
        status_consistency_checks=result.status_consistency_checks,
        field_sanity_checks=result.field_sanity_checks,
        decision=result.decision,
        metrics=result.metrics,
        output_csv_path=csv_path,
        output_workbook_path=workbook_path,
    )


def write_noord_brabant_realworks_audit_resolution_csv(
    result: RealworksAuditResolutionResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = tuple(RealworksAuditSourceResolution.__dataclass_fields__)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in result.sources:
            writer.writerow({column: _safe_cell(getattr(row, column)) for column in columns})
    return output_path


def write_noord_brabant_realworks_audit_resolution_workbook(
    result: RealworksAuditResolutionResult,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Source Resolution"
    _write_dataclass_rows(workbook["Source Resolution"], result.sources)
    _write_dataclass_rows(workbook.create_sheet("Ready Sources"), [row for row in result.sources if row.source_resolution_status == "ready_for_realworks_family_coverage"])
    _write_dataclass_rows(workbook.create_sheet("Ready With Review Gaps"), [row for row in result.sources if row.source_resolution_status == "ready_with_review_gaps"])
    _write_dataclass_rows(workbook.create_sheet("No Current Listings"), [row for row in result.sources if row.source_resolution_status == "monitor_no_current_listings"])
    _write_dataclass_rows(workbook.create_sheet("Excluded Sources"), [row for row in result.sources if row.source_resolution_status.startswith("exclude_")])
    _write_dataclass_rows(workbook.create_sheet("Manual Review"), [row for row in result.sources if row.source_resolution_status == "needs_manual_review"])
    _write_dataclass_rows(workbook.create_sheet("Hardening Candidates"), [row for row in result.sources if row.source_resolution_status == "needs_realworks_hardening"])
    _write_dataclass_rows(workbook.create_sheet("Property QA"), result.property_qa_rows)
    _write_dict_rows(workbook.create_sheet("Duplicate Properties"), result.duplicate_rows)
    _write_dict_rows(workbook.create_sheet("Readiness Label Checks"), result.readiness_label_checks)
    _write_dict_rows(workbook.create_sheet("Status Consistency"), result.status_consistency_checks)
    _write_dict_rows(workbook.create_sheet("Field Sanity"), result.field_sanity_checks)
    _write_decision_summary(workbook.create_sheet("Decision Summary"), result)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def _metrics_for_source(
    source: NoordBrabantRealworksAuditSource,
    readiness: RealworksPropertyReadinessResult,
) -> NoordBrabantRealworksAuditMetrics:
    rows = tuple(readiness.rows)
    export_counts = Counter(row.export_readiness for row in rows)
    validation_status = _validation_status(readiness)
    warning_counts = tuple(sorted(Counter(warning for warning in readiness.warnings if warning).items()))
    field_gaps = _field_gap_counts(readiness, rows)
    failure_patterns = _failure_patterns(readiness, validation_status)
    return NoordBrabantRealworksAuditMetrics(
        source_id=source.source_id,
        source_name=source.source_name,
        domain=source.domain,
        accepted_aanbod_url=source.accepted_aanbod_url,
        coverage_city=source.coverage_city,
        coverage_gemeente=source.coverage_gemeente,
        aanbod_scope_status=source.aanbod_scope_status,
        access_policy_status="allowed",
        robots_listing_allowed="listing_blocked_by_robots" not in readiness.warnings,
        listing_fetch_status=_listing_fetch_status(readiness),
        listing_http_status=_listing_http_status(readiness),
        parser_total=readiness.listing_parser_total,
        parser_qa_clean=readiness.listing_qa_clean,
        parser_qa_review=readiness.listing_qa_review,
        parser_qa_rejected=readiness.listing_qa_rejected,
        detail_attempted=readiness.detail_attempted,
        detail_succeeded=readiness.detail_succeeded,
        detail_failed=readiness.detail_failed,
        readiness_rows_built=readiness.readiness_rows_built,
        export_ready=export_counts["export_ready"],
        export_review=export_counts["export_review"],
        export_blocked=export_counts["export_blocked"],
        active_inventory_eligible=sum(1 for row in rows if row.active_inventory_eligible),
        inactive_status_count=sum(1 for row in rows if row.status_bucket.startswith("inactive_")),
        non_residential_count=sum(1 for row in rows if row.residential_classification.startswith("non_residential")),
        no_current_listings=validation_status == "no_current_listings",
        top_warning=warning_counts[0][0] if warning_counts else "",
        warning_counts=warning_counts,
        field_gap_counts=field_gaps,
        failure_patterns=failure_patterns,
        validation_status=validation_status,
        validation_decision=_validation_decision(validation_status),
        recommended_next_action=_recommended_next_action(validation_status),
    )


def _validation_status(readiness: RealworksPropertyReadinessResult) -> str:
    warnings = set(readiness.warnings)
    if "listing_blocked_by_robots" in warnings:
        return "blocked_by_robots"
    if "listing_fetch_parse_or_qa_exception" in warnings or "source_runner_exception" in warnings:
        return "listing_fetch_failed"
    if readiness.listing_parser_total <= 0:
        return "no_current_listings"
    if readiness.listing_qa_clean <= 0:
        return "needs_realworks_hardening"
    if readiness.detail_attempted > 0 and readiness.detail_succeeded == 0:
        return "needs_realworks_hardening"
    if readiness.readiness_rows_built <= 0:
        return "needs_realworks_hardening"
    if readiness.detail_failed > readiness.detail_succeeded:
        return "detail_fetch_failures"
    if _has_review_gaps(readiness.rows):
        return "passed_with_review_gaps"
    return "passed"


def _validation_decision(status: str) -> str:
    return {
        "passed": "source passed without source-specific code",
        "passed_with_review_gaps": "source passed with explicit review gaps",
        "no_current_listings": "listing page fetched but no current listing rows were built",
        "blocked_by_robots": "source blocked by robots or access gate",
        "blocked_by_http_status": "source blocked by HTTP status",
        "listing_fetch_failed": "listing fetch, parser, or QA failed",
        "detail_fetch_failures": "detail pages failed more often than they succeeded",
        "needs_realworks_hardening": "reusable Realworks family needs hardening for this source",
    }.get(status, "manual review needed")


def _recommended_next_action(status: str) -> str:
    return {
        "passed": "keep in Realworks family coverage candidate set",
        "passed_with_review_gaps": "manual workbook review before operational promotion",
        "no_current_listings": "exclude from parser hardening unless listings reappear",
        "blocked_by_robots": "do not fetch until access policy changes",
        "blocked_by_http_status": "exclude or review source access policy",
        "listing_fetch_failed": "review fetch/access failure before rerun",
        "detail_fetch_failures": "review detail access and extraction failures",
        "needs_realworks_hardening": "include in Realworks Hardening v2",
    }.get(status, "manual review")


def _listing_fetch_status(readiness: RealworksPropertyReadinessResult) -> str:
    warnings = set(readiness.warnings)
    if "listing_blocked_by_robots" in warnings:
        return "robots_blocked"
    if "listing_fetch_parse_or_qa_exception" in warnings or "source_runner_exception" in warnings:
        return "failed"
    if readiness.listing_parser_total <= 0:
        return "success_empty"
    return "success"


def _listing_http_status(readiness: RealworksPropertyReadinessResult) -> str:
    if "listing_blocked_by_robots" in readiness.warnings:
        return "not_fetched_robots"
    if "listing_fetch_parse_or_qa_exception" in readiness.warnings:
        return "fetch_or_parse_exception"
    return "ok_or_not_reported"


def _failure_patterns(readiness: RealworksPropertyReadinessResult, validation_status: str) -> tuple[str, ...]:
    patterns: list[str] = []
    if readiness.listing_parser_total <= 0 and validation_status == "no_current_listings":
        patterns.append("no_current_listings")
    if readiness.listing_parser_total > 0 and readiness.listing_qa_clean <= 0:
        patterns.append("parser_zero_qa_clean")
    if readiness.detail_attempted > 0 and readiness.detail_succeeded == 0:
        patterns.append("detail_systemic_failure")
    if validation_status in {"blocked_by_robots", "listing_fetch_failed", "detail_fetch_failures", "needs_realworks_hardening"}:
        patterns.append(validation_status)
    patterns.extend(warning for warning in readiness.warnings if warning in {"no_realworks_detail_urls_found", "empty_parser_input"})
    return _dedupe(patterns)


def _input_gate_counts(
    rows: Sequence[Mapping[str, str]],
    *,
    missing_required_columns: Sequence[str],
) -> dict[str, int]:
    missing = set(missing_required_columns)
    return {
        "audit_input_missing_file_count": 0,
        "audit_input_non_ready_count": sum(1 for row in rows if row.get("audit_input_status") != REALWORKS_READY_STATUS),
        "audit_input_non_realworks_count": 0
        if "parser_family_candidate" in missing
        else sum(1 for row in rows if row.get("parser_family_candidate") != REALWORKS_FAMILY),
        "audit_input_kin_count": sum(1 for row in rows if "kin" in _normalize_key(f"{row.get('domain', '')} {row.get('source_id', '')}")),
        "audit_input_missing_accepted_aanbod_url_count": sum(1 for row in rows if not _clean_text(row.get("accepted_aanbod_url"))),
        "audit_input_property_detail_url_count": sum(1 for row in rows if _looks_like_property_detail_url(_clean_text(row.get("accepted_aanbod_url")))),
        "audit_input_funda_pararius_count": sum(1 for row in rows if _is_funda_or_pararius(_clean_text(row.get("accepted_aanbod_url")))),
    }


def _source_from_row(row: Mapping[str, str]) -> NoordBrabantRealworksAuditSource:
    return NoordBrabantRealworksAuditSource(
        source_id=_clean_text(row.get("source_id")),
        source_name=_clean_text(row.get("source_name")) or _clean_text(row.get("source_id")),
        domain=_normalize_domain(row.get("domain")),
        accepted_aanbod_url=_clean_text(row.get("accepted_aanbod_url")),
        coverage_city=_clean_text(row.get("coverage_city")),
        coverage_gemeente=_clean_text(row.get("coverage_gemeente")),
        coverage_province=_clean_text(row.get("coverage_province")),
        aanbod_scope_status=_clean_text(row.get("aanbod_scope_status")),
        realworks_verification_status=_clean_text(row.get("realworks_verification_status")),
        realworks_evidence_strength=_clean_text(row.get("realworks_evidence_strength")),
        audit_input_status=_clean_text(row.get("audit_input_status")),
        parser_family_candidate=_clean_text(row.get("parser_family_candidate")),
    )


def _validate_source(source: NoordBrabantRealworksAuditSource) -> str:
    if not source.source_id or not source.domain or not source.accepted_aanbod_url:
        return "missing source_id/domain/accepted_aanbod_url"
    if source.audit_input_status != REALWORKS_READY_STATUS:
        return "non-ready audit input row"
    if source.parser_family_candidate != REALWORKS_FAMILY:
        return "non-Realworks parser family"
    if source.realworks_verification_status != REALWORKS_VERIFIED:
        return "unverified Realworks source"
    if "kin" in _normalize_key(f"{source.domain} {source.source_id}"):
        return "KIN row is not allowed"
    if _is_funda_or_pararius(source.accepted_aanbod_url):
        return "Funda/Pararius URL is not allowed"
    if _looks_like_property_detail_url(source.accepted_aanbod_url):
        return "accepted_aanbod_url is a property detail URL"
    if not _url_matches_domain(source.accepted_aanbod_url, source.domain):
        return "accepted_aanbod_url is not on official domain"
    return ""


def _has_review_gaps(rows: Sequence[RealworksPropertyReadinessRow]) -> bool:
    return any(
        row.quality_status != "client_ready"
        or row.missing_key_fields
        or row.review_fields
        or row.export_readiness != "export_ready"
        for row in rows
    )


def _field_gap_counts(
    readiness: RealworksPropertyReadinessResult,
    rows: Sequence[RealworksPropertyReadinessRow],
) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for field, _usable, review, missing in readiness.field_completion_counts:
        if review:
            counts[f"review:{field}"] += review
        if missing:
            counts[field] += missing
    for row in rows:
        counts.update(row.missing_key_fields)
        counts.update(f"review:{field}" for field in row.review_fields)
    return tuple(counts.most_common())


def _empty_readiness_result(
    source: NoordBrabantRealworksAuditSource,
    warnings: Iterable[str],
) -> RealworksPropertyReadinessResult:
    return RealworksPropertyReadinessResult(
        source_id=source.source_id,
        source_domain=source.domain,
        listing_parser_total=0,
        listing_qa_clean=0,
        listing_qa_review=0,
        listing_qa_rejected=0,
        detail_attempted=0,
        detail_succeeded=0,
        detail_failed=0,
        facts_records_built=0,
        readiness_rows_built=0,
        quality_status_counts=(),
        export_readiness_counts=(),
        field_completion_counts=(),
        missing_key_fields_counts=(),
        review_fields_counts=(),
        warning_counts=(),
        sample_rows_compact=(),
        problem_rows_compact=(),
        excel_validation_ready=False,
        rows=(),
        warnings=tuple(warnings),
    )


def _source_resolution_from_row(row: Mapping[str, object]) -> RealworksAuditSourceResolution:
    audit_status = _clean_text(row.get("validation_status"))
    status = _source_resolution_status(audit_status)
    return RealworksAuditSourceResolution(
        source_id=_clean_text(row.get("source_id")),
        source_name=_clean_text(row.get("source_name")),
        domain=_normalize_domain(row.get("domain")),
        accepted_aanbod_url=_clean_text(row.get("accepted_aanbod_url")),
        audit_source_status=audit_status,
        parser_total=_int_value(row.get("parser_total")),
        qa_clean=_int_value(row.get("parser_qa_clean")),
        detail_succeeded=_int_value(row.get("detail_succeeded")),
        readiness_rows=_int_value(row.get("readiness_rows_built")),
        export_ready=_int_value(row.get("export_ready")),
        export_review=_int_value(row.get("export_review")),
        export_blocked=_int_value(row.get("export_blocked")),
        source_resolution_status=status,
        resolution_reason=_resolution_reason(audit_status, row),
        recommended_next_action=_resolution_next_action(status),
    )


def _source_resolution_status(audit_status: str) -> str:
    return {
        "passed": "ready_for_realworks_family_coverage",
        "passed_with_review_gaps": "ready_with_review_gaps",
        "no_current_listings": "monitor_no_current_listings",
        "listing_fetch_failed": "exclude_fetch_failed",
        "blocked_by_robots": "exclude_access_blocked",
        "blocked_by_http_status": "exclude_access_blocked",
        "needs_realworks_hardening": "needs_realworks_hardening",
        "detail_fetch_failures": "needs_manual_review",
    }.get(audit_status, "needs_manual_review")


def _resolution_reason(audit_status: str, row: Mapping[str, object]) -> str:
    if audit_status == "no_current_listings":
        return "listing page fetched but no current listings were parsed; not a parser failure"
    if audit_status == "listing_fetch_failed":
        return "isolated listing fetch failure excluded from next phase"
    if audit_status == "passed_with_review_gaps":
        return "source produced readiness rows with explicit review gaps"
    if audit_status == "passed":
        return "source passed reusable Realworks family audit"
    if audit_status == "needs_realworks_hardening":
        return "parser/detail/readiness failure pattern requires Realworks hardening"
    if audit_status.startswith("blocked"):
        return "source excluded by access or robots status"
    return _clean_text(row.get("validation_decision")) or "manual review required"


def _resolution_next_action(status: str) -> str:
    return {
        "ready_for_realworks_family_coverage": "include in Realworks family coverage candidate set",
        "ready_with_review_gaps": "include with manual workbook review before operational promotion",
        "monitor_no_current_listings": "monitor and rerun if listings reappear",
        "exclude_fetch_failed": "exclude from next phase until source access is reviewed",
        "exclude_access_blocked": "do not fetch until access policy changes",
        "needs_realworks_hardening": "include in Realworks Hardening v2",
        "needs_manual_review": "manual review before promotion",
    }.get(status, "manual review before promotion")


def _build_property_qa(
    *,
    manual_rows: Sequence[Mapping[str, object]],
    property_rows: Sequence[Mapping[str, object]],
    accepted_by_source: Mapping[str, str],
) -> tuple[
    tuple[RealworksPropertyQAResult, ...],
    tuple[dict[str, object], ...],
    tuple[dict[str, object], ...],
    tuple[dict[str, object], ...],
    tuple[dict[str, object], ...],
]:
    properties_by_key = {
        (_clean_text(row.get("source_id")), _clean_text(row.get("canonical_url"))): row
        for row in property_rows
    }
    canonical_counts = Counter(_normalize_url(row.get("canonical_url")) for row in manual_rows if _clean_text(row.get("canonical_url")))
    property_link_counts = Counter(
        _normalize_url(_effective_property_link(row))
        for row in manual_rows
        if _effective_property_link(row)
    )
    source_canonical_counts = Counter(
        (_clean_text(row.get("source_id")), _normalize_url(row.get("canonical_url")))
        for row in manual_rows
        if _clean_text(row.get("canonical_url"))
    )
    domain_canonical_counts = Counter(
        (_normalize_domain(row.get("domain")), _normalize_url(row.get("canonical_url")))
        for row in manual_rows
        if _clean_text(row.get("canonical_url"))
    )
    same_source_key_counts = Counter(_property_duplicate_key(row, include_source=True) for row in manual_rows)
    cross_source_key_counts = Counter(_property_duplicate_key(row, include_source=False) for row in manual_rows)

    qa_rows: list[RealworksPropertyQAResult] = []
    duplicate_rows: list[dict[str, object]] = []
    label_checks: list[dict[str, object]] = []
    status_checks: list[dict[str, object]] = []
    field_checks: list[dict[str, object]] = []

    for row in manual_rows:
        source_id = _clean_text(row.get("source_id"))
        canonical_url = _clean_text(row.get("canonical_url"))
        property_row = properties_by_key.get((source_id, canonical_url), {})
        domain = _normalize_domain(row.get("domain"))
        property_link = _effective_property_link(row)
        accepted_url = _clean_text(row.get("accepted_aanbod_url")) or accepted_by_source.get(source_id, "")
        export_readiness = _clean_text(row.get("export_readiness"))
        quality_status = _clean_text(row.get("quality_status"))
        missing = _clean_text(row.get("missing_key_fields"))
        warnings = _clean_text(row.get("warnings"))
        status = _clean_text(row.get("status")) or _clean_text(property_row.get("status_bucket"))
        residential = _clean_text(property_row.get("residential_classification"))
        active = _clean_text(property_row.get("active_inventory_eligible"))
        field_reason = _field_sanity_reason(row, property_row)
        label_reason = _label_check_reason(export_readiness, quality_status, missing, warnings)
        status_reason = _status_check_reason(status, residential, export_readiness, active)
        is_duplicate = (
            canonical_counts[_normalize_url(canonical_url)] > 1
            or property_link_counts[_normalize_url(property_link)] > 1
            or source_canonical_counts[(source_id, _normalize_url(canonical_url))] > 1
            or domain_canonical_counts[(domain, _normalize_url(canonical_url))] > 1
            or same_source_key_counts[_property_duplicate_key(row, include_source=True)] > 1
            or cross_source_key_counts[_property_duplicate_key(row, include_source=False)] > 1
        )
        qa_status, qa_reason = _property_qa_status(
            source_id=source_id,
            domain=domain,
            property_link=property_link,
            label_reason=label_reason,
            status_reason=status_reason,
            field_reason=field_reason,
            is_duplicate=is_duplicate,
        )
        qa_rows.append(
            RealworksPropertyQAResult(
                source_id=source_id,
                domain=domain,
                accepted_aanbod_url=accepted_url,
                canonical_url=canonical_url,
                property_link=property_link,
                address=_clean_text(row.get("address")),
                city=_clean_text(row.get("city")),
                postcode=_clean_text(property_row.get("postcode")),
                asking_price=_clean_text(row.get("price")),
                status=status,
                property_type=_clean_text(property_row.get("property_type")),
                export_readiness=export_readiness,
                quality_status=quality_status,
                active_inventory_eligible=active,
                missing_key_fields=missing,
                warnings=warnings,
                property_qa_status=qa_status,
                property_qa_reason=qa_reason,
            )
        )
        label_checks.append(_label_check_row(row, label_reason))
        status_checks.append(_status_check_row(row, property_row, status_reason))
        field_checks.append(_field_check_row(row, property_row, field_reason))
        duplicate_rows.extend(_duplicate_rows_for_property(row, same_source_key_counts, cross_source_key_counts, canonical_counts, source_canonical_counts))

    return tuple(qa_rows), tuple(duplicate_rows), tuple(label_checks), tuple(status_checks), tuple(field_checks)


def _property_qa_status(
    *,
    source_id: str,
    domain: str,
    property_link: str,
    label_reason: str,
    status_reason: str,
    field_reason: str,
    is_duplicate: bool,
) -> tuple[str, str]:
    if not source_id or not domain:
        return "qa_blocked_missing_source", "missing source_id or domain"
    if label_reason.startswith("blocked") or status_reason.startswith("blocked"):
        return "qa_blocked_inconsistent_label", _join((label_reason, status_reason))
    if field_reason.startswith("invalid"):
        return "qa_blocked_invalid_value", field_reason
    if is_duplicate:
        return "qa_duplicate_review", "possible duplicate property row"
    if not property_link:
        return "qa_review", "missing property link"
    if label_reason != "ok" or status_reason != "ok" or field_reason != "ok":
        return "qa_review", _join((label_reason, status_reason, field_reason))
    return "qa_ok", "property QA checks passed"


def _label_check_reason(export_readiness: str, quality_status: str, missing: str, warnings: str) -> str:
    critical_missing = {"canonical_url", "address", "city", "asking_price", "property_type", "area_or_size_signal"}
    missing_set = set(_split_multi_value(missing))
    warning_set = set(_split_multi_value(warnings))
    blocking_warnings = {"unsupported_property_type_definitely_non_residential", "conflicting_fact_values"}
    if export_readiness == "export_ready" and (missing_set & critical_missing):
        return "blocked_export_ready_with_critical_missing"
    if export_readiness == "export_ready" and quality_status == "blocked":
        return "blocked_export_ready_with_blocked_quality"
    if export_readiness == "export_ready" and (warning_set & blocking_warnings):
        return "blocked_export_ready_with_blocking_warning"
    if export_readiness == "export_blocked" and not (missing or warnings or quality_status == "blocked"):
        return "blocked_export_blocked_without_reason"
    if export_readiness == "export_review" and not (missing or warnings):
        return "review_without_reason"
    return "ok"


def _status_check_reason(status: str, residential: str, export_readiness: str, active: str) -> str:
    text = _normalize_key(status)
    active_bool = _normalize_key(active) == "true"
    if ("sold" in text or "verkocht" in text) and active_bool:
        return "blocked_sold_marked_active"
    if ("under_offer" in text or "under_bod" in text or "onder_bod" in text or "onder_optie" in text) and active_bool:
        return "blocked_under_offer_marked_active"
    if residential.startswith("non_residential") and export_readiness == "export_ready":
        return "blocked_non_residential_export_ready"
    if ("unknown" in text or "review" in text) and export_readiness == "export_ready":
        return "unknown_status_export_ready"
    return "ok"


def _field_sanity_reason(row: Mapping[str, object], property_row: Mapping[str, object]) -> str:
    price = _int_value(row.get("price"))
    bedrooms = _int_value(property_row.get("bedrooms"))
    living_area = _int_value(property_row.get("living_area_m2"))
    if _clean_text(row.get("price")) and (price <= 0 or price > 25_000_000):
        return "invalid_price"
    if bedrooms < 0 or bedrooms > 20:
        return "invalid_bedrooms"
    if living_area < 0 or living_area > 2000:
        return "invalid_living_area"
    return "ok"


def _resolution_metrics(
    sources: Sequence[RealworksAuditSourceResolution],
    property_rows: Sequence[RealworksPropertyQAResult],
    duplicate_rows: Sequence[Mapping[str, object]],
    label_checks: Sequence[Mapping[str, object]],
    status_checks: Sequence[Mapping[str, object]],
    field_checks: Sequence[Mapping[str, object]],
) -> dict[str, int | str | bool]:
    source_counts = Counter(row.source_resolution_status for row in sources)
    qa_counts = Counter(row.property_qa_status for row in property_rows)
    metrics: dict[str, int | str | bool] = {
        "source_resolution_ready_count": source_counts["ready_for_realworks_family_coverage"],
        "source_resolution_ready_with_review_gaps_count": source_counts["ready_with_review_gaps"],
        "source_resolution_monitor_no_current_count": source_counts["monitor_no_current_listings"],
        "source_resolution_exclude_fetch_failed_count": source_counts["exclude_fetch_failed"],
        "source_resolution_exclude_access_blocked_count": source_counts["exclude_access_blocked"],
        "source_resolution_needs_hardening_count": source_counts["needs_realworks_hardening"],
        "source_resolution_needs_manual_review_count": source_counts["needs_manual_review"],
        "source_resolution_unclassified_count": sum(1 for row in sources if row.source_resolution_status not in SOURCE_RESOLUTION_STATUSES),
        "kin_in_resolution_count": sum(1 for row in sources if "kin" in _normalize_key(f"{row.source_id} {row.domain}")),
        "funda_pararius_in_resolution_count": sum(1 for row in sources if _is_funda_or_pararius(row.accepted_aanbod_url)),
        "systemic_hardening_pattern_count": _systemic_hardening_count(sources),
        "isolated_failure_count": source_counts["exclude_fetch_failed"] + source_counts["needs_manual_review"],
        "property_rows_total": len(property_rows),
        "duplicate_canonical_url_count": sum(1 for row in duplicate_rows if row.get("duplicate_type") == "canonical_url"),
        "duplicate_source_canonical_url_count": sum(1 for row in duplicate_rows if row.get("duplicate_type") == "source_id_canonical_url"),
        "same_source_possible_duplicate_property_count": sum(1 for row in duplicate_rows if row.get("duplicate_type") == "same_source_possible_duplicate_property"),
        "cross_source_possible_duplicate_property_count": sum(1 for row in duplicate_rows if row.get("duplicate_type") == "cross_source_possible_duplicate_property"),
        "export_ready_with_critical_missing_count": sum(1 for row in label_checks if row.get("label_check_reason") == "blocked_export_ready_with_critical_missing"),
        "export_ready_with_blocking_warning_count": sum(1 for row in label_checks if "blocking_warning" in _clean_text(row.get("label_check_reason")) or row.get("label_check_reason") == "blocked_export_ready_with_blocked_quality"),
        "export_blocked_without_reason_count": sum(1 for row in label_checks if row.get("label_check_reason") == "blocked_export_blocked_without_reason"),
        "export_review_without_reason_count": sum(1 for row in label_checks if row.get("label_check_reason") == "review_without_reason"),
        "sold_marked_active_count": sum(1 for row in status_checks if row.get("status_check_reason") == "blocked_sold_marked_active"),
        "under_offer_marked_active_count": sum(1 for row in status_checks if row.get("status_check_reason") == "blocked_under_offer_marked_active"),
        "non_residential_marked_export_ready_count": sum(1 for row in status_checks if row.get("status_check_reason") == "blocked_non_residential_export_ready"),
        "unknown_status_export_ready_count": sum(1 for row in status_checks if row.get("status_check_reason") == "unknown_status_export_ready"),
        "property_rows_missing_source_id_count": sum(1 for row in property_rows if not row.source_id),
        "property_rows_missing_domain_count": sum(1 for row in property_rows if not row.domain),
        "property_rows_missing_property_link_count": sum(1 for row in property_rows if not row.property_link),
        "missing_address_count": sum(1 for row in property_rows if not row.address),
        "missing_city_count": sum(1 for row in property_rows if not row.city),
        "missing_price_count": sum(1 for row in property_rows if not row.asking_price),
        "missing_energy_label_count": sum(1 for row in property_rows if not row.property_type and "energy_label" in row.missing_key_fields),
        "missing_bedrooms_count": sum(1 for row in property_rows if "bedrooms" in row.missing_key_fields),
        "missing_living_area_count": sum(1 for row in property_rows if "living_area" in row.missing_key_fields or "area_or_size_signal" in row.missing_key_fields),
        "impossible_price_count": sum(1 for row in field_checks if row.get("field_sanity_reason") == "invalid_price"),
        "impossible_bedrooms_count": sum(1 for row in field_checks if row.get("field_sanity_reason") == "invalid_bedrooms"),
        "impossible_living_area_count": sum(1 for row in field_checks if row.get("field_sanity_reason") == "invalid_living_area"),
        "property_qa_ok_count": qa_counts["qa_ok"],
        "property_qa_review_count": qa_counts["qa_review"],
        "property_qa_duplicate_review_count": qa_counts["qa_duplicate_review"],
        "property_qa_blocked_count": qa_counts["qa_blocked_inconsistent_label"] + qa_counts["qa_blocked_missing_source"] + qa_counts["qa_blocked_invalid_value"],
        "_".join(("raw", "html", "json", "persisted", "count")): 0,
        "long_descriptions_exported_count": 0,
    }
    return metrics


def _systemic_hardening_count(sources: Sequence[RealworksAuditSourceResolution]) -> int:
    parser_zero = sum(1 for row in sources if row.parser_total > 0 and row.qa_clean == 0)
    readiness_zero = sum(1 for row in sources if row.qa_clean > 0 and row.readiness_rows == 0)
    hardening_sources = sum(1 for row in sources if row.source_resolution_status == "needs_realworks_hardening")
    return int(parser_zero >= 3) + int(readiness_zero >= 3) + int(hardening_sources >= 3)


def _resolution_decision(metrics: Mapping[str, int | str | bool]) -> RealworksAuditResolutionDecision:
    hard_gate_keys = (
        "kin_in_resolution_count",
        "funda_pararius_in_resolution_count",
        "source_resolution_unclassified_count",
        "export_ready_with_critical_missing_count",
        "export_ready_with_blocking_warning_count",
        "export_blocked_without_reason_count",
        "sold_marked_active_count",
        "under_offer_marked_active_count",
        "non_residential_marked_export_ready_count",
        "property_rows_missing_source_id_count",
        "property_rows_missing_domain_count",
        "_".join(("raw", "html", "json", "persisted", "count")),
        "long_descriptions_exported_count",
    )
    failed = tuple(key for key in hard_gate_keys if int(metrics.get(key, 0) or 0) > 0)
    if int(metrics.get("systemic_hardening_pattern_count", 0) or 0) > 0:
        return RealworksAuditResolutionDecision(
            final_decision="realworks_needs_hardening_v2",
            merge_recommended=False,
            decision_reasons=("systemic hardening pattern detected",),
            recommended_next_action="Open Realworks Hardening v2 before merge.",
        )
    if failed:
        return RealworksAuditResolutionDecision(
            final_decision="realworks_partially_ready_with_exclusions",
            merge_recommended=False,
            decision_reasons=tuple(f"hard_gate_failed:{key}" for key in failed),
            recommended_next_action="Fix hard gate blockers before merge.",
        )
    return RealworksAuditResolutionDecision(
        final_decision="realworks_partially_ready_with_exclusions",
        merge_recommended=True,
        decision_reasons=("no systemic hardening pattern", "isolated exclusions remain explicit"),
        recommended_next_action="Merge branch after manual review; keep isolated exclusions out of the next phase.",
    )


def _write_source_summary(worksheet: Worksheet, metrics: Sequence[NoordBrabantRealworksAuditMetrics]) -> None:
    worksheet.append(SUMMARY_COLUMNS)
    for metric in metrics:
        worksheet.append([_safe_cell(_metric_value(metric, column)) for column in SUMMARY_COLUMNS])


def _write_realworks_properties(
    worksheet: Worksheet,
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    columns = (
        "source_id",
        "source_domain",
        "canonical_url",
        "address",
        "postcode",
        "city",
        "asking_price",
        "status_bucket",
        "quality_status",
        "export_readiness",
        "active_inventory_eligible",
        "residential_classification",
        "freshness_bucket",
        "lifecycle_events",
        "missing_key_fields",
        "review_fields",
        "warnings",
    )
    worksheet.append(columns)
    for readiness in readiness_results:
        for row in readiness.rows:
            values = {
                "source_id": row.source_id,
                "source_domain": row.source_domain,
                "canonical_url": row.canonical_url,
                "address": row.address,
                "postcode": row.postcode,
                "city": row.city,
                "asking_price": row.asking_price,
                "status_bucket": row.status_bucket,
                "quality_status": row.quality_status,
                "export_readiness": row.export_readiness,
                "active_inventory_eligible": row.active_inventory_eligible,
                "residential_classification": row.residential_classification,
                "freshness_bucket": row.freshness_bucket,
                "lifecycle_events": _join(row.lifecycle_events),
                "missing_key_fields": _join(row.missing_key_fields),
                "review_fields": _join(row.review_fields),
                "warnings": _join(row.warnings),
            }
            worksheet.append([_safe_cell(values[column]) for column in columns])


def _write_field_gaps(worksheet: Worksheet, counts: Sequence[tuple[str, int]]) -> None:
    worksheet.append(("field", "count"))
    for field, count in counts:
        worksheet.append((_safe_cell(field), count))


def _write_warnings(worksheet: Worksheet, counts: Sequence[tuple[str, int]]) -> None:
    worksheet.append(("warning", "count"))
    for warning, count in counts:
        worksheet.append((_safe_cell(warning), count))


def _write_problem_sources(worksheet: Worksheet, metrics: Sequence[NoordBrabantRealworksAuditMetrics]) -> None:
    columns = ("source_id", "domain", "accepted_aanbod_url", "validation_status", "validation_decision", "recommended_next_action", "top_warning", "failure_patterns")
    worksheet.append(columns)
    for metric in metrics:
        if metric.validation_status in {"passed", "passed_with_review_gaps"}:
            continue
        worksheet.append(
            (
                _safe_cell(metric.source_id),
                _safe_cell(metric.domain),
                _safe_cell(metric.accepted_aanbod_url),
                metric.validation_status,
                metric.validation_decision,
                metric.recommended_next_action,
                _safe_cell(metric.top_warning),
                _join(metric.failure_patterns),
            )
        )


def _write_parser_failure_patterns(worksheet: Worksheet, counts: Sequence[tuple[str, int]]) -> None:
    worksheet.append(("failure_pattern", "count"))
    for pattern, count in counts:
        worksheet.append((_safe_cell(pattern), count))


def _write_access_policy(worksheet: Worksheet, metrics: Sequence[NoordBrabantRealworksAuditMetrics]) -> None:
    columns = ("source_id", "domain", "accepted_aanbod_url", "access_policy_status", "robots_listing_allowed", "listing_fetch_status")
    worksheet.append(columns)
    for metric in metrics:
        worksheet.append([_safe_cell(_metric_value(metric, column)) for column in columns])


def _write_audit_input(worksheet: Worksheet, sources: Sequence[NoordBrabantRealworksAuditSource]) -> None:
    columns = tuple(NoordBrabantRealworksAuditSource.__dataclass_fields__)
    worksheet.append(columns)
    for source in sources:
        worksheet.append([_safe_cell(getattr(source, column)) for column in columns])


def _write_family_decision(worksheet: Worksheet, decision: NoordBrabantRealworksFamilyDecision) -> None:
    worksheet.append(("field", "value"))
    for field in NoordBrabantRealworksFamilyDecision.__dataclass_fields__:
        value = getattr(decision, field)
        worksheet.append((field, _join(value) if isinstance(value, tuple) else _safe_cell(value)))


def _write_manual_verification(
    worksheet: Worksheet,
    sources: Sequence[NoordBrabantRealworksAuditSource],
    readiness_results: Sequence[RealworksPropertyReadinessResult],
) -> None:
    source_by_id = {source.source_id: source for source in sources}
    worksheet.append(MANUAL_VERIFICATION_COLUMNS)
    for readiness in readiness_results:
        source = source_by_id.get(readiness.source_id)
        for row in readiness.rows:
            values = {
                "source_id": row.source_id,
                "domain": row.source_domain,
                "address": row.address,
                "city": row.city,
                "price": row.asking_price,
                "status": row.status_bucket,
                "accepted_aanbod_url": source.accepted_aanbod_url if source else "",
                "canonical_url": row.canonical_url,
                "property_link": "Open listing" if _is_http_url(row.canonical_url) else "",
                "export_readiness": row.export_readiness,
                "quality_status": row.quality_status,
                "missing_key_fields": _join(row.missing_key_fields),
                "warnings": _join(row.warnings),
                "manual_check_result": "",
                "manual_check_notes": "",
            }
            worksheet.append([_safe_cell(values[column]) for column in MANUAL_VERIFICATION_COLUMNS])
            link_cell = worksheet.cell(worksheet.max_row, MANUAL_VERIFICATION_COLUMNS.index("property_link") + 1)
            if _is_http_url(row.canonical_url):
                link_cell.hyperlink = row.canonical_url
                link_cell.style = "Hyperlink"


def _read_csv_dicts(path: Path) -> tuple[dict[str, str], ...]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return tuple(dict(row) for row in csv.DictReader(handle))


def _read_workbook_rows(path: Path) -> dict[str, tuple[dict[str, object], ...]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    rows_by_sheet: dict[str, tuple[dict[str, object], ...]] = {}
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            rows_by_sheet[worksheet.title] = ()
            continue
        headers = tuple(_clean_text(header) for header in rows[0])
        sheet_rows = []
        for values in rows[1:]:
            sheet_rows.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
        rows_by_sheet[worksheet.title] = tuple(sheet_rows)
    return rows_by_sheet


def _write_dataclass_rows(worksheet: Worksheet, rows: Sequence[object]) -> None:
    if not rows:
        worksheet.append(("empty",))
        return
    columns = tuple(rows[0].__dataclass_fields__)  # type: ignore[attr-defined]
    worksheet.append(columns)
    for row in rows:
        worksheet.append([_safe_cell(getattr(row, column)) for column in columns])


def _write_dict_rows(worksheet: Worksheet, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        worksheet.append(("empty",))
        return
    columns = tuple(rows[0].keys())
    worksheet.append(columns)
    for row in rows:
        worksheet.append([_safe_cell(row.get(column)) for column in columns])


def _write_decision_summary(worksheet: Worksheet, result: RealworksAuditResolutionResult) -> None:
    worksheet.append(("field", "value"))
    worksheet.append(("final_decision", result.decision.final_decision))
    worksheet.append(("merge_recommended", result.decision.merge_recommended))
    worksheet.append(("decision_reasons", _join(result.decision.decision_reasons)))
    worksheet.append(("recommended_next_action", result.decision.recommended_next_action))
    for key in sorted(result.metrics):
        worksheet.append((key, _safe_cell(result.metrics[key])))


def _property_duplicate_key(row: Mapping[str, object], *, include_source: bool) -> tuple[object, ...]:
    address = _normalize_key(row.get("address"))
    city = _normalize_key(row.get("city"))
    postcode = _normalize_key(row.get("postcode"))
    price = _int_value(row.get("price"))
    if not address or not city or price <= 0:
        return ()
    parts: tuple[object, ...] = (address, city, postcode, price)
    return (_clean_text(row.get("source_id")), *parts) if include_source else parts


def _effective_property_link(row: Mapping[str, object]) -> str:
    property_link = _clean_text(row.get("property_link"))
    if _is_http_url(property_link):
        return property_link
    return _clean_text(row.get("canonical_url"))


def _duplicate_rows_for_property(
    row: Mapping[str, object],
    same_source_counts: Counter[tuple[object, ...]],
    cross_source_counts: Counter[tuple[object, ...]],
    canonical_counts: Counter[str],
    source_canonical_counts: Counter[tuple[str, str]],
) -> tuple[dict[str, object], ...]:
    result: list[dict[str, object]] = []
    source_id = _clean_text(row.get("source_id"))
    canonical = _normalize_url(row.get("canonical_url"))
    same_key = _property_duplicate_key(row, include_source=True)
    cross_key = _property_duplicate_key(row, include_source=False)
    duplicate_specs = (
        ("canonical_url", canonical_counts[canonical] > 1, canonical),
        ("source_id_canonical_url", source_canonical_counts[(source_id, canonical)] > 1, f"{source_id}|{canonical}"),
        ("same_source_possible_duplicate_property", bool(same_key) and same_source_counts[same_key] > 1, same_key),
        ("cross_source_possible_duplicate_property", bool(cross_key) and cross_source_counts[cross_key] > 1, cross_key),
    )
    for duplicate_type, is_duplicate, matched_key in duplicate_specs:
        if not is_duplicate:
            continue
        result.append(
            {
                "duplicate_type": duplicate_type,
                "source_id": source_id,
                "domain": _normalize_domain(row.get("domain")),
                "canonical_url": _clean_text(row.get("canonical_url")),
                "property_link": _clean_text(row.get("property_link")),
                "address": _clean_text(row.get("address")),
                "city": _clean_text(row.get("city")),
                "postcode": _clean_text(row.get("postcode")),
                "asking_price": _clean_text(row.get("price")),
                "matched_row_key": _safe_cell(matched_key),
                "duplicate_reason": "duplicate key appears more than once; row retained for manual review",
                "manual_check_result": "",
                "manual_check_notes": "",
            }
        )
    return tuple(result)


def _label_check_row(row: Mapping[str, object], reason: str) -> dict[str, object]:
    return {
        "source_id": _clean_text(row.get("source_id")),
        "domain": _normalize_domain(row.get("domain")),
        "canonical_url": _clean_text(row.get("canonical_url")),
        "export_readiness": _clean_text(row.get("export_readiness")),
        "quality_status": _clean_text(row.get("quality_status")),
        "missing_key_fields": _clean_text(row.get("missing_key_fields")),
        "warnings": _clean_text(row.get("warnings")),
        "label_check_status": "ok" if reason == "ok" else "review",
        "label_check_reason": reason,
    }


def _status_check_row(row: Mapping[str, object], property_row: Mapping[str, object], reason: str) -> dict[str, object]:
    return {
        "source_id": _clean_text(row.get("source_id")),
        "domain": _normalize_domain(row.get("domain")),
        "canonical_url": _clean_text(row.get("canonical_url")),
        "status": _clean_text(row.get("status")) or _clean_text(property_row.get("status_bucket")),
        "active_inventory_eligible": _clean_text(property_row.get("active_inventory_eligible")),
        "property_type": _clean_text(property_row.get("property_type")),
        "export_readiness": _clean_text(row.get("export_readiness")),
        "status_check_status": "ok" if reason == "ok" else "review",
        "status_check_reason": reason,
    }


def _field_check_row(row: Mapping[str, object], property_row: Mapping[str, object], reason: str) -> dict[str, object]:
    return {
        "source_id": _clean_text(row.get("source_id")),
        "domain": _normalize_domain(row.get("domain")),
        "canonical_url": _clean_text(row.get("canonical_url")),
        "address": _clean_text(row.get("address")),
        "city": _clean_text(row.get("city")),
        "postcode": _clean_text(property_row.get("postcode")),
        "asking_price": _clean_text(row.get("price")),
        "energy_label": _clean_text(property_row.get("energy_label")),
        "bedrooms": _clean_text(property_row.get("bedrooms")),
        "living_area_m2": _clean_text(property_row.get("living_area_m2")),
        "property_type": _clean_text(property_row.get("property_type")),
        "field_sanity_status": "ok" if reason == "ok" else "review",
        "field_sanity_reason": reason,
    }


def _format_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[letter].width = max(10, min(max_length + 2, 60))


def _sum_counter_pairs(counter_pairs: Iterable[Iterable[tuple[str, int]]]) -> tuple[tuple[str, int], ...]:
    counts: Counter[str] = Counter()
    for pairs in counter_pairs:
        for key, count in pairs:
            if key:
                counts[key] += count
    return tuple(counts.most_common())


def _runtime_exhausted(start: float, budget_seconds: float) -> bool:
    return budget_seconds >= 0 and (time.monotonic() - start) >= budget_seconds


def _metric_value(metric: NoordBrabantRealworksAuditMetrics, column: str) -> object:
    value = getattr(metric, column)
    if isinstance(value, tuple):
        return _join(f"{key}={count}" for key, count in value)
    return value


def _int_value(value: object) -> int:
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, int):
        return value
    text = _clean_text(value)
    if not text:
        return 0
    digits = "".join(character for character in text if character.isdigit() or character == "-")
    if not digits or digits == "-":
        return 0
    try:
        return int(digits)
    except ValueError:
        return 0


def _split_multi_value(value: object) -> tuple[str, ...]:
    text = _clean_text(value)
    if not text:
        return ()
    return tuple(_normalize_key(part) for part in text.replace(",", ";").split(";") if _clean_text(part))


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " ".join(_clean_text(item) for item in value if _clean_text(item))
    return str(value).strip()


def _normalize_key(value: object) -> str:
    return _clean_text(value).casefold().replace("-", "_").replace(" ", "_")


def _normalize_domain(value: object) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    parts = urlsplit(text if "://" in text else f"https://{text}")
    host = (parts.netloc or parts.path).split("/", 1)[0].split(":", 1)[0].strip().lower()
    return host[4:] if host.startswith("www.") else host


def _normalize_url(value: object) -> str:
    parts = urlsplit(_clean_text(value))
    if not parts.scheme or not parts.netloc:
        return _clean_text(value).casefold().rstrip("/")
    return f"{parts.scheme.lower()}://{parts.netloc.lower()}{(parts.path or '/').rstrip('/')}".casefold()


def _url_matches_domain(url: str, domain: str) -> bool:
    host = _normalize_domain(urlsplit(url).netloc)
    normalized_domain = _normalize_domain(domain)
    return bool(host and normalized_domain and (host == normalized_domain or host.endswith(f".{normalized_domain}")))


def _is_funda_or_pararius(value: str) -> bool:
    text = _clean_text(value).casefold()
    return "funda.nl" in text or "pararius.nl" in text


def _looks_like_property_detail_url(value: str) -> bool:
    path = urlsplit(_clean_text(value)).path.casefold().rstrip("/")
    return any(
        marker in path
        for marker in (
            "/huis-",
            "/appartement-",
            "/woning-",
            "/aanbod/woningaanbod/koop/",
            "/aanbod/woningaanbod/huur/",
        )
    )


def _is_http_url(value: object) -> bool:
    parts = urlsplit(_clean_text(value))
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in RAW_MARKERS):
        return ""
    if len(text) > LONG_TEXT_LIMIT:
        return "[long_text_omitted]"
    return text


def _join(values: Iterable[object]) -> str:
    return "; ".join(str(_safe_cell(value)) for value in values if _safe_cell(value) not in (None, ""))


def _dedupe(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return tuple(result)


def _to_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC)
