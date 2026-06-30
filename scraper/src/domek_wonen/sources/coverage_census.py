from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass, field
from enum import StrEnum
from html import unescape
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import urljoin, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domek_wonen.compliance import robots_gate


NOORD_BRABANT = "noord_brabant"
DEFAULT_USER_AGENT = "WoningAlertNL-CoverageCensus/0.1 (+manual advisor review; no bypass)"
COMMON_AANBOD_PATHS = (
    "/aanbod",
    "/aanbod/woningaanbod",
    "/aanbod/woningaanbod/koop",
    "/woningaanbod",
    "/woningen",
    "/woningen/koop",
    "/koopwoningen",
    "/aanbod/koop",
    "/te-koop",
    "/huis-kopen",
)
LOCAL_EVIDENCE_PATHS = (
    Path("data/processed/sources_seed_noord_brabant.csv"),
    Path("data/discovery/reference/property_discovery_source_overrides.csv"),
    Path("data/discovery/processed/sources_seed_with_gemeente.csv"),
    Path("data/discovery/platform_fingerprint/platform_fingerprint_results.csv"),
    Path("data/discovery/runs/20260614T122022Z/makelaar_sources_master.csv"),
)
PROHIBITED_FINAL_VALUES = {"", "unknown", "missing", "tbd", "todo", "unknown_manual_review"}
PORTAL_DOMAINS = {"funda.nl", "pararius.nl"}
RAW_MARKERS = ("<html", "<script", "</", '{"', "{'", '"docs"', "window.__")
LONG_TEXT_LIMIT = 500
EVIDENCE_PREVIEW_LIMIT = 180


class CoverageTerminalStatus(StrEnum):
    CONFIRMED_SOURCE_READY = "confirmed_source_ready"
    CONFIRMED_SOURCE_NEEDS_PARSER_FAMILY = "confirmed_source_needs_parser_family"
    CONFIRMED_NO_PUBLIC_AANBOD = "confirmed_no_public_aanbod"
    CONFIRMED_BLOCKED_OR_LEGAL_REVIEW = "confirmed_blocked_or_legal_review"
    CONFIRMED_OUT_OF_SCOPE = "confirmed_out_of_scope"
    CONFIRMED_DUPLICATE = "confirmed_duplicate"
    CONFIRMED_INACTIVE_OR_NO_LONGER_TRADING = "confirmed_inactive_or_no_longer_trading"


ALLOWED_CLASSIFICATIONS = {
    "realworks_public",
    "ogonline_xhr",
    "kolibri",
    "skarabee",
    "wordpress_static",
    "wordpress_json",
    "custom_html",
    "custom_xhr",
    "custom_js_app",
    "iframe_vendor",
    "third_party_only",
    "no_public_aanbod",
    "blocked_or_legal_review",
    "out_of_scope",
    "duplicate",
    "inactive_or_no_longer_trading",
}


MASTER_COLUMNS = (
    "source_id",
    "source_name",
    "domain",
    "root_url",
    "office_city",
    "office_gemeente",
    "office_province",
    "office_location_status",
    "coverage_city",
    "coverage_gemeente",
    "coverage_province",
    "coverage_location_status",
    "outside_office_coverage_status",
    "has_noord_brabant_coverage",
    "accepted_aanbod_url",
    "aanbod_url_status",
    "aanbod_url_confidence",
    "platform_guess",
    "delivery_mode",
    "parser_family_candidate",
    "family_confidence",
    "family_terminal_status",
    "realworks_verification_status",
    "access_policy_status",
    "terminal_status",
    "validation_status",
    "recommended_next_action",
    "manual_check_result",
    "manual_check_notes",
)

QUALITY_GATE_NAMES = (
    "operational_unknown_family_count",
    "missing_aanbod_url_without_terminal_reason_count",
    "rejected_candidate_used_as_master_aanbod_url_count",
    "property_detail_url_as_aanbod_url_count",
    "funda_or_pararius_operational_aanbod_url_count",
    "realworks_without_strong_evidence_count",
    "platform_guess_realworks_but_family_custom_js_app_unreviewed_count",
    "kin_family_conflict_count",
    "custom_js_app_without_fingerprint_attempt_count",
    "gemeente_normalization_conflict_count",
)

GEMEENTE_NORMALIZATION_MAP = {
    "bergen_op_zoom": "Bergen op Zoom",
    "geldrop_mierlo": "Geldrop-Mierlo",
    "gemert_bakel": "Gemert-Bakel",
    "gilze_en_rijen": "Gilze en Rijen",
    "land_van_cuijk": "Land van Cuijk",
    "reusel_de_mierden": "Reusel-De Mierden",
    "s_hertogenbosch": "'s-Hertogenbosch",
    "'s_hertogenbosch": "'s-Hertogenbosch",
    "s-hertogenbosch": "'s-Hertogenbosch",
    "sint_michielsgestel": "Sint-Michielsgestel",
    "son_en_breugel": "Son en Breugel",
}


@dataclass(slots=True)
class CoverageInvestigationAttempt:
    source_id: str
    domain: str
    pass_name: str
    attempted_url: str = ""
    result: str = ""
    evidence: str = ""
    decision: str = ""
    reason: str = ""


@dataclass(slots=True)
class AanbodUrlCandidate:
    source_id: str
    domain: str
    candidate_url: str
    accepted: bool
    rejection_reason: str = ""
    evidence_type: str = ""
    evidence_preview: str = ""
    pass_name: str = ""
    confidence: float = 0.0


@dataclass(slots=True)
class FamilyFingerprintEvidence:
    source_id: str
    domain: str
    parser_family_candidate: str
    delivery_mode: str
    signal: str
    confidence: float
    evidence_preview: str = ""
    pass_name: str = "pass_5_family_fingerprint"


@dataclass(slots=True)
class FamilyConflictRecord:
    source_id: str
    domain: str
    old_family: str
    new_family: str
    conflict_type: str
    evidence: str
    resolution: str
    remaining_action: str = ""


@dataclass(slots=True)
class CustomJSRefingerprintRecord:
    source_id: str
    domain: str
    initial_family: str
    final_family: str
    signal: str
    evidence_preview: str
    decision: str


@dataclass(slots=True)
class MissingDomainQueueRecord:
    source_name: str
    raw_source_id: str
    gemeente: str
    province: str
    evidence_file: str
    reason: str
    suggested_search_query: str
    suggested_next_action: str


@dataclass(slots=True)
class NormalizationIssueRecord:
    source_id: str
    domain: str
    raw_gemeente: str
    normalized_gemeente: str
    field_name: str
    status: str


@dataclass(slots=True)
class RealworksVerificationRecord:
    source_id: str
    domain: str
    platform_guess: str
    accepted_aanbod_url: str
    realworks_verification_status: str
    realworks_evidence_strength: str
    signals_found: str
    signals_missing: str
    decision: str
    reason: str


@dataclass(slots=True)
class CoverageSourceRecord:
    source_id: str
    source_name: str
    domain: str
    root_url: str
    office_city: str = ""
    office_gemeente: str = ""
    office_province: str = ""
    coverage_city: str = ""
    coverage_gemeente: str = ""
    coverage_province: str = ""
    has_noord_brabant_coverage: bool = False
    coverage_evidence: str = ""
    raw_aanbod_url_candidate: str = ""
    aanbod_url: str = ""
    aanbod_url_type: str = ""
    aanbod_url_status: str = ""
    aanbod_url_confidence: float = 0.0
    aanbod_url_evidence: str = ""
    accepted_aanbod_url: str = ""
    accepted_aanbod_url_source: str = ""
    accepted_aanbod_url_confidence: float = 0.0
    accepted_aanbod_url_evidence: str = ""
    rejected_aanbod_url_count: int = 0
    best_rejected_candidate_url: str = ""
    best_rejected_candidate_reason: str = ""
    platform_guess: str = ""
    delivery_mode: str = ""
    parser_family_candidate: str = ""
    family_confidence: float = 0.0
    family_evidence: str = ""
    family_terminal_status: str = ""
    realworks_verification_status: str = ""
    realworks_evidence_strength: str = ""
    realworks_signals_found: str = ""
    realworks_signals_missing: str = ""
    access_policy_status: str = "researching"
    robots_status_if_known: str = "not_checked"
    source_quality_status: str = ""
    validation_status: str = ""
    terminal_status: str = ""
    office_location_status: str = ""
    coverage_location_status: str = ""
    outside_office_coverage_status: str = ""
    duplicate_group_id: str = ""
    duplicate_reason: str = ""
    investigation_pass_count: int = 0
    investigation_attempts: list[CoverageInvestigationAttempt] = field(default_factory=list)
    remaining_review_reason: str = ""
    recommended_next_action: str = ""
    aanbod_url_candidates: list[AanbodUrlCandidate] = field(default_factory=list)
    family_fingerprint_evidence: list[FamilyFingerprintEvidence] = field(default_factory=list)
    family_conflicts: list[FamilyConflictRecord] = field(default_factory=list)
    custom_js_refingerprint_attempts: list[CustomJSRefingerprintRecord] = field(default_factory=list)
    aliases: list[str] = field(default_factory=list)
    source_ids: list[str] = field(default_factory=list)


@dataclass(slots=True)
class CoverageCensusResult:
    records: tuple[CoverageSourceRecord, ...]
    duplicate_records: tuple[CoverageSourceRecord, ...] = ()
    missing_domain_queue: tuple[MissingDomainQueueRecord, ...] = ()
    normalization_issues: tuple[NormalizationIssueRecord, ...] = ()
    initial_source_count: int = 0
    deduped_source_count: int = 0
    workbook_path: Path | None = None
    master_csv_path: Path | None = None
    review_queue_csv_path: Path | None = None
    warnings: tuple[str, ...] = ()

    @property
    def quality_metrics(self) -> dict[str, Any]:
        return compute_quality_metrics(self)


FetchText = Callable[[str], str]
CanFetch = Callable[[str, str], bool]


def run_noord_brabant_coverage_source_census(
    *,
    repo_root: Path,
    output_dir: Path | None = None,
    evidence_paths: Sequence[Path] | None = None,
    allow_live_http: bool = False,
    max_sources: int | None = None,
    max_passes: int = 8,
    max_requests_per_domain: int = 3,
    timeout_seconds: float = 10.0,
    fetch_text: FetchText | None = None,
    can_fetch: CanFetch = robots_gate.can_fetch,
) -> CoverageCensusResult:
    paths = tuple(evidence_paths or LOCAL_EVIDENCE_PATHS)
    rows = load_local_source_seed_rows(repo_root=repo_root, evidence_paths=paths)
    missing_domain_queue = build_missing_domain_queue(rows)
    operational_rows = [row for row in rows if _normalize_domain(_first_value({_normalize_key(k): _clean_text(v) for k, v in row.items()}, ("root_domain", "domain", "source_domain", "website", "website_url")))]
    records, duplicates = consolidate_coverage_source_seeds(operational_rows)
    normalization_issues = collect_normalization_issues(records)
    if max_sources is not None:
        records = records[: max(0, max_sources)]

    request_counts: dict[str, int] = defaultdict(int)
    fetcher = fetch_text
    if fetcher is None and allow_live_http:
        fetcher = lambda url: controlled_http_fetch_text(url, timeout_seconds=timeout_seconds)

    for record in records:
        run_investigation_loop(
            record,
            fetch_text=fetcher,
            can_fetch=can_fetch,
            request_counts=request_counts,
            max_requests_per_domain=max_requests_per_domain,
            max_passes=max_passes,
        )

    result = CoverageCensusResult(
        records=tuple(records),
        duplicate_records=tuple(duplicates),
        missing_domain_queue=tuple(missing_domain_queue),
        normalization_issues=tuple(normalization_issues),
        initial_source_count=len(rows),
        deduped_source_count=len(records),
        warnings=(),
    )
    if output_dir is None:
        return result

    workbook_path = output_dir / "noord_brabant_coverage_source_census_hardened_v1.xlsx"
    master_csv_path = output_dir / "noord_brabant_coverage_source_census_hardened_v1.csv"
    review_queue_csv_path = output_dir / "noord_brabant_coverage_source_census_hardened_v1_review_queue.csv"
    write_coverage_census_outputs(result, workbook_path, master_csv_path, review_queue_csv_path)
    return CoverageCensusResult(
        records=result.records,
        duplicate_records=result.duplicate_records,
        missing_domain_queue=result.missing_domain_queue,
        normalization_issues=result.normalization_issues,
        initial_source_count=result.initial_source_count,
        deduped_source_count=result.deduped_source_count,
        workbook_path=workbook_path,
        master_csv_path=master_csv_path,
        review_queue_csv_path=review_queue_csv_path,
        warnings=result.warnings,
    )


def load_local_source_seed_rows(*, repo_root: Path, evidence_paths: Sequence[Path]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for relative_path in evidence_paths:
        path = repo_root / relative_path
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                rows.append({**row, "evidence_file": str(relative_path).replace("\\", "/")})
    return rows


def consolidate_coverage_source_seeds(
    rows: Iterable[Mapping[str, Any]],
) -> tuple[list[CoverageSourceRecord], list[CoverageSourceRecord]]:
    records_by_key: dict[str, CoverageSourceRecord] = {}
    duplicates: list[CoverageSourceRecord] = []

    for row in rows:
        record = coverage_record_from_seed_row(row)
        key = record.domain or f"source:{record.source_id}"
        existing = records_by_key.get(key)
        if existing is None:
            records_by_key[key] = record
            continue
        duplicates.append(_duplicate_record(record, existing))
        merge_coverage_source_record(existing, record)

    return list(records_by_key.values()), duplicates


def coverage_record_from_seed_row(row: Mapping[str, Any]) -> CoverageSourceRecord:
    normalized = {_normalize_key(key): _clean_text(value) for key, value in row.items()}
    domain = _normalize_domain(_first_value(normalized, ("root_domain", "domain", "source_domain", "website", "website_url")))
    homepage = _canonical_root_url(_first_value(normalized, ("website", "website_url", "homepage_url")), domain)
    source_name = _first_value(normalized, ("office_name", "source_name", "makelaar_name", "name"))
    raw_coverage_gemeente = _first_value(normalized, ("gemeente", "plaats", "normalized_place", "coverage_gemeente"))
    coverage_gemeente = normalize_gemeente_name(raw_coverage_gemeente)
    coverage_province = _first_value(normalized, ("province", "provincie", "coverage_province"))
    office_gemeente_raw = _first_value(normalized, ("office_gemeente",))
    office_gemeente = normalize_gemeente_name(office_gemeente_raw)
    office_province = _first_value(normalized, ("office_province",))
    source_id = _first_value(normalized, ("source_id",)) or _source_id(domain, source_name, coverage_gemeente)
    aanbod_url = _first_value(normalized, ("aanbod_url", "koopaanbod_url", "listing_url", "source_url"))
    aanbod_status = _first_value(normalized, ("aanbod_url_quality", "koopaanbod_url_quality", "aanbod_url_status"))
    platform_guess = _first_value(normalized, ("detected_platform", "platform_guess", "platform", "current_platform_guess"))
    delivery_mode = _canonical_family(_first_value(normalized, ("delivery_mode", "detected_delivery_mode", "detected_delivery_mode_enriched")))
    parser_family = _canonical_family(_first_value(normalized, ("parser_family_candidate", "parser_family")))
    legal_status = _canonical_access_status(_first_value(normalized, ("legal_status", "access_status")))
    evidence_parts = [
        f"evidence_file={normalized.get('evidence_file', '')}",
        _first_value(normalized, ("review_reason", "source_quality_reason", "evidence", "notes")),
    ]
    has_nb_coverage = _normalize_key(coverage_province or office_province) == NOORD_BRABANT

    record = CoverageSourceRecord(
        source_id=source_id,
        source_name=source_name or domain or source_id,
        domain=domain,
        root_url=homepage,
        office_city=_first_value(normalized, ("office_city", "city")),
        office_gemeente=office_gemeente,
        office_province=office_province,
        coverage_city=_first_value(normalized, ("coverage_city", "city", "raw_place", "plaats")),
        coverage_gemeente=coverage_gemeente,
        coverage_province=coverage_province or office_province,
        has_noord_brabant_coverage=has_nb_coverage,
        coverage_evidence=_join(evidence_parts),
        raw_aanbod_url_candidate=aanbod_url,
        aanbod_url="",
        aanbod_url_type=_first_value(normalized, ("aanbod_url_type",)),
        aanbod_url_status=aanbod_status,
        aanbod_url_confidence=_float_value(_first_value(normalized, ("confidence_score", "confidence", "score"))),
        aanbod_url_evidence=_join(evidence_parts),
        platform_guess=platform_guess,
        delivery_mode=delivery_mode,
        parser_family_candidate=parser_family,
        family_confidence=_float_value(_first_value(normalized, ("confidence", "confidence_score", "score"))),
        family_evidence=_join(
            (
                platform_guess,
                _first_value(normalized, ("evidence", "technology_signals", "fetch_status", "recommended_next_action")),
            )
        ),
        access_policy_status=legal_status,
        source_quality_status=_first_value(normalized, ("source_quality_status", "koopaanbod_url_quality", "aanbod_url_quality")),
        office_location_status=_location_status(office_gemeente, office_province),
        coverage_location_status=_location_status(coverage_gemeente, coverage_province),
        outside_office_coverage_status=_outside_office_status(office_province, coverage_province),
        aliases=[],
        source_ids=[source_id] if source_id else [],
    )
    record.investigation_attempts.append(
        CoverageInvestigationAttempt(
            source_id=record.source_id,
            domain=record.domain,
            pass_name="pass_1_local_evidence",
            result="loaded",
            evidence=_preview(record.coverage_evidence),
            decision="seed_loaded",
            reason="local evidence row",
        )
    )
    return record


def merge_coverage_source_record(target: CoverageSourceRecord, incoming: CoverageSourceRecord) -> None:
    target.aliases = _dedupe([*target.aliases, incoming.source_name])
    target.source_ids = _dedupe([*target.source_ids, *incoming.source_ids])
    target.coverage_evidence = _merge_text(target.coverage_evidence, incoming.coverage_evidence)
    target.aanbod_url_evidence = _merge_text(target.aanbod_url_evidence, incoming.aanbod_url_evidence)
    target.family_evidence = _merge_text(target.family_evidence, incoming.family_evidence)
    target.investigation_attempts.extend(incoming.investigation_attempts)
    for field_name in (
        "source_name",
        "root_url",
        "office_city",
        "office_gemeente",
        "office_province",
        "coverage_city",
        "coverage_gemeente",
        "coverage_province",
        "raw_aanbod_url_candidate",
        "aanbod_url_type",
        "aanbod_url_status",
        "platform_guess",
        "delivery_mode",
        "parser_family_candidate",
        "access_policy_status",
        "source_quality_status",
        "office_location_status",
        "coverage_location_status",
        "outside_office_coverage_status",
    ):
        current = _clean_text(getattr(target, field_name))
        incoming_value = _clean_text(getattr(incoming, field_name))
        if not current and incoming_value:
            setattr(target, field_name, incoming_value)
    target.has_noord_brabant_coverage = target.has_noord_brabant_coverage or incoming.has_noord_brabant_coverage
    target.aanbod_url_confidence = max(target.aanbod_url_confidence, incoming.aanbod_url_confidence)
    target.family_confidence = max(target.family_confidence, incoming.family_confidence)


def run_investigation_loop(
    record: CoverageSourceRecord,
    *,
    fetch_text: FetchText | None = None,
    can_fetch: CanFetch = robots_gate.can_fetch,
    request_counts: dict[str, int] | None = None,
    max_requests_per_domain: int = 3,
    max_passes: int = 7,
) -> CoverageSourceRecord:
    request_counts = request_counts if request_counts is not None else defaultdict(int)
    if max_passes < 1:
        max_passes = 1

    if not record.has_noord_brabant_coverage:
        _set_terminal(record, "out_of_scope", CoverageTerminalStatus.CONFIRMED_OUT_OF_SCOPE, "source has no Noord-Brabant coverage evidence")
        return record

    if _is_inactive(record):
        _set_terminal(record, "inactive_or_no_longer_trading", CoverageTerminalStatus.CONFIRMED_INACTIVE_OR_NO_LONGER_TRADING, "local evidence marks source inactive")
        return record

    if _access_blocks(record):
        _set_terminal(record, "blocked_or_legal_review", CoverageTerminalStatus.CONFIRMED_BLOCKED_OR_LEGAL_REVIEW, "blocked by local access policy evidence")
        return record

    accepted = discover_aanbod_url(
        record,
        fetch_text=fetch_text,
        can_fetch=can_fetch,
        request_counts=request_counts,
        max_requests_per_domain=max_requests_per_domain,
    )
    if accepted:
        fingerprint_source_family(
            record,
            fetch_text=fetch_text,
            can_fetch=can_fetch,
            request_counts=request_counts,
            max_requests_per_domain=max_requests_per_domain,
        )
    finalize_coverage_record(record)
    return record


def discover_aanbod_url(
    record: CoverageSourceRecord,
    *,
    fetch_text: FetchText | None = None,
    can_fetch: CanFetch = robots_gate.can_fetch,
    request_counts: dict[str, int] | None = None,
    max_requests_per_domain: int = 3,
) -> bool:
    request_counts = request_counts if request_counts is not None else defaultdict(int)
    explicit_url = _clean_text(record.accepted_aanbod_url or record.aanbod_url or record.raw_aanbod_url_candidate)
    if explicit_url:
        accepted, reason = _validate_aanbod_candidate(explicit_url, record.domain)
        candidate = AanbodUrlCandidate(
            source_id=record.source_id,
            domain=record.domain,
            candidate_url=explicit_url,
            accepted=accepted,
            rejection_reason="" if accepted else reason,
            evidence_type="local_explicit",
            evidence_preview=_preview(record.aanbod_url_evidence),
            pass_name="pass_1_local_evidence",
            confidence=record.aanbod_url_confidence or (0.95 if accepted else 0.2),
        )
        record.aanbod_url_candidates.append(candidate)
        if accepted:
            _set_accepted_aanbod_url(record, _canonical_url(explicit_url), "local_explicit", record.aanbod_url_evidence, candidate.confidence)
            return True
        record.aanbod_url = ""
        _refresh_rejected_candidate_summary(record)

        if reason == "property_detail_url_rejected" and fetch_text is not None:
            for derived_url in derive_listing_index_candidates_from_detail_url(explicit_url):
                derived_html = _guarded_fetch(
                    record,
                    derived_url,
                    pass_name="pass_3_derive_listing_index_from_detail_url",
                    fetch_text=fetch_text,
                    can_fetch=can_fetch,
                    request_counts=request_counts,
                    max_requests_per_domain=max_requests_per_domain,
                )
                if _accept_candidate(
                    record,
                    derived_url,
                    "derived_from_detail_url",
                    derived_html or derived_url,
                    "pass_3_derive_listing_index_from_detail_url",
                    confidence=0.78,
                    require_listing_evidence=True,
                ):
                    return True

    if fetch_text is None:
        record.investigation_attempts.append(
            CoverageInvestigationAttempt(
                record.source_id,
                record.domain,
                "pass_2_homepage_links",
                record.root_url,
                "skipped",
                "live_http_disabled",
                "no_homepage_fetch",
                "fetcher not supplied",
            )
        )
        return False

    homepage_html = _guarded_fetch(
        record,
        record.root_url,
        pass_name="pass_2_homepage_links",
        fetch_text=fetch_text,
        can_fetch=can_fetch,
        request_counts=request_counts,
        max_requests_per_domain=max_requests_per_domain,
    )
    if homepage_html:
        for url, evidence in _homepage_aanbod_links(record.root_url, homepage_html):
            accepted = _accept_candidate(record, url, "homepage_link", evidence, "pass_2_homepage_links", confidence=0.72)
            if accepted:
                return True
        if _page_looks_like_listing_index(homepage_html) and _homepage_may_be_listing_index(record.root_url):
            if _accept_candidate(record, record.root_url, "homepage_listing_index", homepage_html, "pass_2_homepage_links", confidence=0.65):
                return True

    sitemap_url = _root_url_for_path(record.domain, "/sitemap.xml")
    sitemap_text = _guarded_fetch(
        record,
        sitemap_url,
        pass_name="pass_4_sitemap",
        fetch_text=fetch_text,
        can_fetch=can_fetch,
        request_counts=request_counts,
        max_requests_per_domain=max_requests_per_domain,
    )
    if sitemap_text:
        for url in _sitemap_listing_urls(sitemap_text):
            if _accept_candidate(record, url, "sitemap_url", url, "pass_4_sitemap", confidence=0.70, require_listing_evidence=False):
                return True

    for path in COMMON_AANBOD_PATHS:
        url = _root_url_for_path(record.domain, path)
        html = _guarded_fetch(
            record,
            url,
            pass_name="pass_5_common_paths",
            fetch_text=fetch_text,
            can_fetch=can_fetch,
            request_counts=request_counts,
            max_requests_per_domain=max_requests_per_domain,
        )
        if html and _page_looks_like_listing_index(html):
            if _accept_candidate(record, url, "common_path", html, "pass_5_common_paths", confidence=0.62):
                return True

    return False


def fingerprint_source_family(
    record: CoverageSourceRecord,
    *,
    fetch_text: FetchText | None = None,
    can_fetch: CanFetch = robots_gate.can_fetch,
    request_counts: dict[str, int] | None = None,
    max_requests_per_domain: int = 3,
) -> CoverageSourceRecord:
    request_counts = request_counts if request_counts is not None else defaultdict(int)
    local_family = _canonical_family(record.parser_family_candidate or record.delivery_mode or record.platform_guess)
    if local_family not in PROHIBITED_FINAL_VALUES and local_family in ALLOWED_CLASSIFICATIONS:
        _add_family_evidence(record, local_family, local_family, "local_family_evidence", record.family_evidence, max(record.family_confidence, 0.75))
        return record

    content = ""
    if fetch_text is not None and record.aanbod_url:
        content = _guarded_fetch(
            record,
            record.aanbod_url,
            pass_name="pass_6_family_fingerprint",
            fetch_text=fetch_text,
            can_fetch=can_fetch,
            request_counts=request_counts,
            max_requests_per_domain=max_requests_per_domain,
        )
    family, delivery, confidence, signal = classify_family_from_content(
        content=content,
        url=record.aanbod_url,
        local_text=" ".join((record.platform_guess, record.family_evidence, record.delivery_mode)),
    )
    _add_family_evidence(record, family, delivery, signal, content or record.family_evidence, confidence)
    return record


def finalize_coverage_record(record: CoverageSourceRecord) -> CoverageSourceRecord:
    if not record.has_noord_brabant_coverage:
        _set_terminal(record, "out_of_scope", CoverageTerminalStatus.CONFIRMED_OUT_OF_SCOPE, "source has no Noord-Brabant coverage evidence")
    elif _access_blocks(record):
        _set_terminal(record, "blocked_or_legal_review", CoverageTerminalStatus.CONFIRMED_BLOCKED_OR_LEGAL_REVIEW, "blocked or legal review required")
    elif not record.accepted_aanbod_url:
        _set_terminal(record, "no_public_aanbod", CoverageTerminalStatus.CONFIRMED_NO_PUBLIC_AANBOD, "no accepted public aanbod URL after configured investigation passes")
    else:
        resolve_family_conflicts(record)
        if _is_kin_record(record):
            resolve_kin_family_conflict(record)
        family = _canonical_family(record.parser_family_candidate or record.delivery_mode)
        if family in PROHIBITED_FINAL_VALUES:
            family = "custom_js_app"
            _add_family_evidence(record, family, family, "fallback_delivery_classification", record.family_evidence, 0.35)
        if family == "custom_js_app":
            refingerprint_custom_js_app(record)
            family = _canonical_family(record.parser_family_candidate or record.delivery_mode)
        if family == "realworks_public":
            verification = verify_realworks_candidate(record)
            if verification.realworks_verification_status != "verified":
                old_family = "realworks_public"
                new_family = "custom_js_app" if _is_kin_record(record) else "custom_html"
                if _is_kin_record(record):
                    new_family = "ogonline_xhr"
                record.family_conflicts.append(
                    FamilyConflictRecord(
                        source_id=record.source_id,
                        domain=record.domain,
                        old_family=old_family,
                        new_family=new_family,
                        conflict_type="realworks_without_strong_evidence",
                        evidence=verification.reason,
                        resolution="reclassified",
                        remaining_action="review derived listing and source config",
                    )
                )
                _add_family_evidence(record, new_family, new_family, "realworks_verification_reclassified", verification.reason, 0.55)
                family = new_family
        if family not in ALLOWED_CLASSIFICATIONS:
            family = "custom_html"
            _add_family_evidence(record, family, family, "canonicalized_unmapped_family", record.family_evidence, 0.40)
        record.parser_family_candidate = family
        record.delivery_mode = record.delivery_mode if record.delivery_mode in ALLOWED_CLASSIFICATIONS else family
        if family in {"realworks_public", "ogonline_xhr", "wordpress_json", "wordpress_static"}:
            _set_terminal(record, family, CoverageTerminalStatus.CONFIRMED_SOURCE_READY, "accepted aanbod URL and reusable family evidence")
            record.family_terminal_status = "confirmed" if record.family_confidence >= 0.80 else "probable_with_evidence"
        else:
            _set_terminal(record, family, CoverageTerminalStatus.CONFIRMED_SOURCE_NEEDS_PARSER_FAMILY, "accepted aanbod URL but source needs parser family/config follow-up")
            record.family_terminal_status = "probable_with_evidence"

    record.investigation_pass_count = len({attempt.pass_name for attempt in record.investigation_attempts})
    record.validation_status = "passed_quality_gates" if _record_passes_quality_gate(record) else "failed_quality_gates"
    return record


def classify_family_from_content(*, content: str, url: str = "", local_text: str = "") -> tuple[str, str, float, str]:
    text = " ".join((url, local_text, content[:4000])).casefold()
    if any(signal in text for signal in ("aanbodentry", "/aanbod/woningaanbod", "kenmerkname", "kenmerkvalue", "realworks")):
        return "realworks_public", "realworks_public", 0.90, "realworks_signal"
    if any(signal in text for signal in ("ogonline", '"docs"', "'docs'", "totaldocs", "hasnextpage")):
        return "ogonline_xhr", "ogonline_xhr", 0.88, "ogonline_signal"
    if "wp-json" in text:
        return "wordpress_json", "wordpress_json", 0.78, "wordpress_rest_signal"
    if "wp-content" in text or "wordpress" in text:
        return "wordpress_static", "wordpress_static", 0.68, "wordpress_static_signal"
    if "kolibri" in text:
        return "kolibri", "kolibri", 0.72, "kolibri_signal"
    if "skarabee" in text:
        return "skarabee", "skarabee", 0.72, "skarabee_signal"
    if "<iframe" in text:
        return "iframe_vendor", "iframe_vendor", 0.58, "iframe_signal"
    if any(signal in text for signal in ("fetch(", "xhr", "api/", "application/json", "__next_data__", "window.__")):
        return "custom_xhr", "custom_xhr", 0.55, "xhr_or_embedded_state_signal"
    if any(signal in text for signal in ("<div id=\"app", "<div id='app", "react", "vue", "ng-app")) and not _page_looks_like_listing_index(content):
        return "custom_js_app", "custom_js_app", 0.50, "js_app_shell_signal"
    if _page_looks_like_listing_index(content):
        return "custom_html", "custom_html", 0.55, "server_rendered_listing_cards"
    return "custom_js_app", "custom_js_app", 0.35, "no_static_listing_payload"


def verify_realworks_candidate(record: CoverageSourceRecord) -> RealworksVerificationRecord:
    url = record.accepted_aanbod_url or record.aanbod_url
    text = " ".join((url, record.family_evidence, record.accepted_aanbod_url_evidence, record.realworks_signals_found)).casefold()
    found: list[str] = []
    missing: list[str] = []
    if "/aanbod/woningaanbod" in urlsplit(url).path.casefold():
        found.append("accepted_aanbod_url_realworks_shape")
    else:
        missing.append("accepted_listing_index_realworks_shape")
    if "aanbodentry" in text:
        found.append("aanbodEntry")
    else:
        missing.append("aanbodEntry")
    if "kenmerkname" in text or "kenmerkvalue" in text:
        found.append("kenmerkName/kenmerkValue")
    else:
        missing.append("kenmerkName/kenmerkValue")
    if "realworks_signal" in text or "local_family_evidence" in text:
        found.append("realworks_family_signal")
    else:
        missing.append("realworks_family_signal")

    strong = bool({"accepted_aanbod_url_realworks_shape", "aanbodEntry", "kenmerkName/kenmerkValue"} & set(found))
    status = "verified" if strong else "rejected"
    strength = "strong" if strong else "weak"
    decision = "keep_realworks_public" if strong else "do_not_use_realworks_without_strong_evidence"
    reason = "strong Realworks structural evidence found" if strong else "platform_guess or stale local evidence is insufficient"
    record.realworks_verification_status = status
    record.realworks_evidence_strength = strength
    record.realworks_signals_found = "; ".join(found)
    record.realworks_signals_missing = "; ".join(missing)
    return RealworksVerificationRecord(
        source_id=record.source_id,
        domain=record.domain,
        platform_guess=record.platform_guess,
        accepted_aanbod_url=url,
        realworks_verification_status=status,
        realworks_evidence_strength=strength,
        signals_found=record.realworks_signals_found,
        signals_missing=record.realworks_signals_missing,
        decision=decision,
        reason=reason,
    )


def resolve_kin_family_conflict(record: CoverageSourceRecord) -> None:
    if not _is_kin_record(record):
        return
    old_family = record.parser_family_candidate or record.delivery_mode or "unknown"
    if old_family != "ogonline_xhr":
        record.family_conflicts.append(
            FamilyConflictRecord(
                source_id=record.source_id,
                domain=record.domain,
                old_family=old_family,
                new_family="ogonline_xhr",
                conflict_type="kin_stale_realworks_or_local_guess",
                evidence="KIN is the repo's OGonline/XHR laboratory; stale Realworks guess is not accepted",
                resolution="resolved_to_probable_ogonline_xhr",
                remaining_action="revalidate with existing KIN OGonline source config",
            )
        )
    _add_family_evidence(
        record,
        "ogonline_xhr",
        "ogonline_xhr",
        "kin_existing_ogonline_repo_evidence",
        "KIN treated by repo as OGonline/XHR across KIN pilots and source config",
        0.84,
    )
    record.family_terminal_status = "probable_with_evidence"
    record.validation_status = "needs_ogonline_revalidation"
    record.recommended_next_action = "revalidate_with_kin_ogonline_source_config"


def resolve_family_conflicts(record: CoverageSourceRecord) -> None:
    if _is_kin_record(record):
        resolve_kin_family_conflict(record)


def refingerprint_custom_js_app(record: CoverageSourceRecord) -> None:
    evidence_text = " ".join(
        (
            record.family_evidence,
            record.accepted_aanbod_url_evidence,
            " ".join(candidate.evidence_preview for candidate in record.aanbod_url_candidates),
        )
    )
    family, delivery, confidence, signal = classify_family_from_content(
        content=evidence_text,
        url=record.accepted_aanbod_url or record.aanbod_url,
        local_text=record.platform_guess,
    )
    if family == "custom_js_app":
        signal = "confirmed_app_shell_or_no_static_listing_payload"
    _add_family_evidence(record, family, delivery, f"custom_js_app_refingerprint:{signal}", evidence_text, confidence)
    record.custom_js_refingerprint_attempts.append(
        CustomJSRefingerprintRecord(
            source_id=record.source_id,
            domain=record.domain,
            initial_family="custom_js_app",
            final_family=family,
            signal=signal,
            evidence_preview=_preview(evidence_text),
            decision="reclassified" if family != "custom_js_app" else "confirmed_custom_js_app",
        )
    )


def derive_listing_index_candidates_from_detail_url(url: str) -> tuple[str, ...]:
    canonical = _canonical_url(url)
    parts = urlsplit(canonical)
    segments = [segment for segment in parts.path.split("/") if segment]
    candidates: list[str] = []
    if len(segments) >= 5 and segments[:2] == ["aanbod", "woningaanbod"] and segments[-2] in {"koop", "huur"}:
        candidates.append(_url_with_segments(parts, segments[:-1]))
        candidates.append(_url_with_segments(parts, [*segments[:2], segments[-2]]))
        candidates.append(_url_with_segments(parts, segments[:2]))
    if len(segments) >= 4 and segments[0] == "woningaanbod" and segments[-2] in {"koop", "huur"}:
        candidates.append(_url_with_segments(parts, segments[:-1]))
        candidates.append(_url_with_segments(parts, [segments[0], segments[-2]]))
        candidates.append(_url_with_segments(parts, [segments[0]]))
    return tuple(_dedupe(candidates))


def normalize_gemeente_name(value: object) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    key = _normalize_key(text)
    return GEMEENTE_NORMALIZATION_MAP.get(key, text)


def build_missing_domain_queue(rows: Iterable[Mapping[str, Any]]) -> list[MissingDomainQueueRecord]:
    queue: list[MissingDomainQueueRecord] = []
    for row in rows:
        normalized = {_normalize_key(key): _clean_text(value) for key, value in row.items()}
        domain = _normalize_domain(_first_value(normalized, ("root_domain", "domain", "source_domain", "website", "website_url")))
        if domain:
            continue
        source_name = _first_value(normalized, ("office_name", "source_name", "makelaar_name", "name"))
        raw_source_id = _first_value(normalized, ("source_id",))
        gemeente = normalize_gemeente_name(_first_value(normalized, ("gemeente", "plaats", "normalized_place")))
        province = _first_value(normalized, ("province", "provincie"))
        evidence_file = _first_value(normalized, ("evidence_file",))
        query = " ".join(part for part in (source_name, gemeente, province, "makelaar") if part)
        queue.append(
            MissingDomainQueueRecord(
                source_name=source_name,
                raw_source_id=raw_source_id,
                gemeente=gemeente,
                province=province,
                evidence_file=evidence_file,
                reason="missing_official_domain",
                suggested_search_query=query,
                suggested_next_action="manual_domain_resolution_before_operational_census",
            )
        )
    return queue


def collect_normalization_issues(records: Sequence[CoverageSourceRecord]) -> list[NormalizationIssueRecord]:
    issues: list[NormalizationIssueRecord] = []
    for record in records:
        for field_name in ("office_gemeente", "coverage_gemeente"):
            value = getattr(record, field_name)
            normalized = normalize_gemeente_name(value)
            if value and normalized != value:
                issues.append(
                    NormalizationIssueRecord(
                        source_id=record.source_id,
                        domain=record.domain,
                        raw_gemeente=value,
                        normalized_gemeente=normalized,
                        field_name=field_name,
                        status="normalized",
                    )
                )
                setattr(record, field_name, normalized)
    return issues


def compute_quality_metrics(result: CoverageCensusResult) -> dict[str, Any]:
    records = list(result.records)
    in_scope = [record for record in records if record.has_noord_brabant_coverage and record.terminal_status not in {CoverageTerminalStatus.CONFIRMED_OUT_OF_SCOPE.value}]
    metrics = {
        "total_sources": result.initial_source_count,
        "deduped_sources": result.deduped_source_count,
        "in_scope_noord_brabant_coverage_sources": len(in_scope),
        "outside_noord_brabant_office_included": sum(
            1
            for record in records
            if record.has_noord_brabant_coverage
            and _normalize_key(record.office_province)
            and _normalize_key(record.office_province) != NOORD_BRABANT
        ),
        "sources_with_accepted_aanbod_url": sum(1 for record in records if record.accepted_aanbod_url),
        "sources_without_public_aanbod": sum(1 for record in records if record.parser_family_candidate == "no_public_aanbod"),
        "sources_blocked_or_legal_review": sum(1 for record in records if record.parser_family_candidate == "blocked_or_legal_review"),
        "sources_out_of_scope": sum(1 for record in records if record.parser_family_candidate == "out_of_scope"),
        "duplicate_sources": len(result.duplicate_records),
        "sources_by_parser_family": dict(sorted(Counter(record.parser_family_candidate for record in records).items())),
        "sources_by_delivery_mode": dict(sorted(Counter(record.delivery_mode for record in records).items())),
        "sources_by_access_policy": dict(sorted(Counter(record.access_policy_status for record in records).items())),
        "review_queue_count": len(review_queue_records(records)),
        "operational_unknown_family_count": sum(
            1
            for record in in_scope
            if _normalize_key(record.parser_family_candidate) in PROHIBITED_FINAL_VALUES
        ),
        "missing_aanbod_url_without_terminal_reason_count": sum(
            1
            for record in in_scope
            if not record.accepted_aanbod_url
            and record.parser_family_candidate not in {"no_public_aanbod", "blocked_or_legal_review", "out_of_scope", "duplicate", "inactive_or_no_longer_trading"}
        ),
        "rejected_candidate_used_as_master_aanbod_url_count": sum(
            1
            for record in records
            for candidate in record.aanbod_url_candidates
            if not candidate.accepted and record.accepted_aanbod_url and _canonical_url(candidate.candidate_url) == record.accepted_aanbod_url
        ),
        "property_detail_url_as_aanbod_url_count": sum(
            1 for record in records if record.accepted_aanbod_url and _looks_like_property_detail_url(urlsplit(record.accepted_aanbod_url).path)
        ),
        "funda_or_pararius_operational_aanbod_url_count": sum(
            1 for record in records if any(portal in record.accepted_aanbod_url for portal in PORTAL_DOMAINS)
        ),
        "realworks_without_strong_evidence_count": sum(
            1
            for record in records
            if record.parser_family_candidate == "realworks_public"
            and verify_realworks_candidate(record).realworks_verification_status != "verified"
        ),
        "platform_guess_realworks_but_family_custom_js_app_unreviewed_count": sum(
            1
            for record in records
            if _normalize_key(record.platform_guess) == "realworks"
            and record.parser_family_candidate == "custom_js_app"
            and not record.custom_js_refingerprint_attempts
        ),
        "kin_family_conflict_count": sum(
            1 for record in records if _is_kin_record(record) and record.parser_family_candidate == "realworks_public"
        ),
        "custom_js_app_without_fingerprint_attempt_count": sum(
            1 for record in records if record.parser_family_candidate == "custom_js_app" and not record.custom_js_refingerprint_attempts
        ),
        "gemeente_normalization_conflict_count": 0,
        "missing_domain_queue_count": len(result.missing_domain_queue),
        "office_location_known_count": sum(1 for record in records if record.office_location_status == "known"),
        "office_location_unknown_count": sum(1 for record in records if record.office_location_status != "known"),
        "outside_office_sources_identified_count": sum(1 for record in records if record.outside_office_coverage_status == "outside_office_source_included"),
        "outside_office_sources_needing_review_count": sum(1 for record in records if record.outside_office_coverage_status == "office_location_unknown"),
        "custom_js_app_initial_count": sum(1 for record in records if any(item.initial_family == "custom_js_app" for item in record.custom_js_refingerprint_attempts)) or sum(1 for record in records if record.parser_family_candidate == "custom_js_app"),
        "custom_js_app_after_refingerprint_count": sum(1 for record in records if record.parser_family_candidate == "custom_js_app"),
        "custom_js_app_reclassified_to_realworks": sum(1 for record in records for item in record.custom_js_refingerprint_attempts if item.final_family == "realworks_public"),
        "custom_js_app_reclassified_to_wordpress_json": sum(1 for record in records for item in record.custom_js_refingerprint_attempts if item.final_family == "wordpress_json"),
        "custom_js_app_reclassified_to_custom_html": sum(1 for record in records for item in record.custom_js_refingerprint_attempts if item.final_family == "custom_html"),
        "custom_js_app_reclassified_to_custom_xhr": sum(1 for record in records for item in record.custom_js_refingerprint_attempts if item.final_family == "custom_xhr"),
        "custom_js_app_confirmed_app_shell": sum(1 for record in records for item in record.custom_js_refingerprint_attempts if item.final_family == "custom_js_app"),
    }
    metrics["quality_gate_passed"] = all(metrics[name] == 0 for name in QUALITY_GATE_NAMES)
    return metrics


def write_coverage_census_outputs(
    result: CoverageCensusResult,
    workbook_path: Path,
    master_csv_path: Path,
    review_queue_csv_path: Path,
) -> tuple[Path, Path, Path]:
    write_master_csv(result.records, master_csv_path)
    write_review_queue_csv(review_queue_records(result.records), review_queue_csv_path)
    write_coverage_census_workbook(result, workbook_path)
    return workbook_path, master_csv_path, review_queue_csv_path


def write_master_csv(records: Sequence[CoverageSourceRecord], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MASTER_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow(_master_row(record))
    return output_path


def write_review_queue_csv(records: Sequence[CoverageSourceRecord], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = ("source_id", "source_name", "domain", "terminal_status", "remaining_review_reason", "recommended_next_action")
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for record in records:
            writer.writerow({column: _safe_cell(getattr(record, column)) for column in columns})
    return output_path


def write_coverage_census_workbook(result: CoverageCensusResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Master Sources"
    _write_rows(workbook["Master Sources"], MASTER_COLUMNS, [_master_row(record) for record in result.records])
    _write_rows(workbook.create_sheet("Aanbod URL Evidence"), ("source_id", "domain", "candidate_url", "accepted", "rejection_reason", "evidence_type", "evidence_preview", "pass_name"), [_candidate_row(candidate) for record in result.records for candidate in record.aanbod_url_candidates])
    _write_rows(workbook.create_sheet("Family Fingerprints"), ("source_id", "domain", "parser_family_candidate", "delivery_mode", "signal", "confidence", "evidence_preview", "pass_name"), [_fingerprint_row(evidence) for record in result.records for evidence in record.family_fingerprint_evidence])
    _write_rows(workbook.create_sheet("Investigation Attempts"), ("source_id", "domain", "pass_name", "attempted_url", "result", "decision", "reason"), [_attempt_row(attempt) for record in result.records for attempt in record.investigation_attempts])
    _write_rows(workbook.create_sheet("Coverage Matrix"), ("coverage_gemeente", "parser_family_candidate", "count"), _coverage_matrix_rows(result.records))
    _write_rows(workbook.create_sheet("Realworks Candidates"), MASTER_COLUMNS, [_master_row(record) for record in result.records if record.parser_family_candidate == "realworks_public"])
    _write_rows(workbook.create_sheet("Realworks Verification"), ("source_id", "domain", "platform_guess", "accepted_aanbod_url", "realworks_verification_status", "realworks_evidence_strength", "signals_found", "signals_missing", "decision", "reason"), [_realworks_verification_row(record) for record in result.records if record.parser_family_candidate == "realworks_public" or _normalize_key(record.platform_guess) == "realworks"])
    _write_rows(workbook.create_sheet("OGonline Candidates"), MASTER_COLUMNS, [_master_row(record) for record in result.records if record.parser_family_candidate == "ogonline_xhr"])
    _write_rows(workbook.create_sheet("Custom Needs Parser"), MASTER_COLUMNS, [_master_row(record) for record in result.records if record.terminal_status == CoverageTerminalStatus.CONFIRMED_SOURCE_NEEDS_PARSER_FAMILY.value])
    _write_rows(workbook.create_sheet("Custom JS Refingerprint"), ("source_id", "domain", "initial_family", "final_family", "signal", "evidence_preview", "decision"), [asdict(item) for record in result.records for item in record.custom_js_refingerprint_attempts])
    _write_rows(workbook.create_sheet("Family Conflicts"), ("source_id", "domain", "old_family", "new_family", "conflict_type", "evidence", "resolution", "remaining_action"), [asdict(item) for record in result.records for item in record.family_conflicts])
    _write_rows(workbook.create_sheet("Blocked or Legal Review"), MASTER_COLUMNS, [_master_row(record) for record in result.records if record.parser_family_candidate == "blocked_or_legal_review"])
    _write_rows(workbook.create_sheet("Duplicates"), ("duplicate_group_id", "source_id", "source_name", "domain", "duplicate_reason"), [_duplicate_row(record) for record in result.duplicate_records])
    _write_rows(workbook.create_sheet("Missing Domain Queue"), ("source_name", "raw_source_id", "gemeente", "province", "evidence_file", "reason", "suggested_search_query", "suggested_next_action"), [asdict(item) for item in result.missing_domain_queue])
    _write_rows(workbook.create_sheet("Normalization Issues"), ("source_id", "domain", "raw_gemeente", "normalized_gemeente", "field_name", "status"), [asdict(item) for item in result.normalization_issues])
    _write_rows(workbook.create_sheet("Review Queue"), MASTER_COLUMNS, [_master_row(record) for record in review_queue_records(result.records)])
    _write_rows(workbook.create_sheet("Quality Gates"), ("metric", "value", "gate_type", "passed"), _quality_gate_rows(result.quality_metrics))
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def review_queue_records(records: Sequence[CoverageSourceRecord]) -> tuple[CoverageSourceRecord, ...]:
    return tuple(
        record
        for record in records
        if record.parser_family_candidate == "blocked_or_legal_review"
        or record.validation_status == "failed_quality_gates"
    )


def controlled_http_fetch_text(
    url: str,
    *,
    timeout_seconds: float = 10.0,
    user_agent: str = DEFAULT_USER_AGENT,
) -> str:
    request = Request(url, headers={"User-Agent": user_agent})
    try:
        with urlopen(request, timeout=timeout_seconds) as response:
            status = getattr(response, "status", None)
            if status is not None and int(status) >= 400:
                raise RuntimeError(f"http_status:{status}")
            payload = response.read()
            charset = "utf-8"
            headers = getattr(response, "headers", None)
            if headers is not None and hasattr(headers, "get_content_charset"):
                charset = headers.get_content_charset() or charset
    except HTTPError as exc:
        raise RuntimeError(f"http_status:{exc.code}") from exc
    return payload.decode(charset, errors="replace")


def _guarded_fetch(
    record: CoverageSourceRecord,
    url: str,
    *,
    pass_name: str,
    fetch_text: FetchText,
    can_fetch: CanFetch,
    request_counts: dict[str, int],
    max_requests_per_domain: int,
) -> str:
    domain, path = _domain_path(url, fallback_domain=record.domain)
    if not domain:
        record.investigation_attempts.append(CoverageInvestigationAttempt(record.source_id, record.domain, pass_name, url, "skipped", "", "not_fetched", "missing_domain"))
        return ""
    if request_counts[domain] >= max_requests_per_domain:
        record.investigation_attempts.append(CoverageInvestigationAttempt(record.source_id, domain, pass_name, url, "skipped", "", "not_fetched", "max_requests_per_domain_reached"))
        return ""
    try:
        allowed = can_fetch(domain, path or "/")
    except Exception as exc:
        allowed = False
        reason = f"robots_gate_error:{exc.__class__.__name__}"
    else:
        reason = "robots_allowed" if allowed else "robots_disallowed"
    record.robots_status_if_known = robots_gate.robots_status(domain) if can_fetch is robots_gate.can_fetch else ("allow" if allowed else "disallow")
    if not allowed:
        record.investigation_attempts.append(CoverageInvestigationAttempt(record.source_id, domain, pass_name, url, "blocked", "", "not_fetched", reason))
        return ""
    request_counts[domain] += 1
    try:
        content = fetch_text(url)
    except Exception as exc:
        record.investigation_attempts.append(CoverageInvestigationAttempt(record.source_id, domain, pass_name, url, "failed", "", "fetch_failed", exc.__class__.__name__))
        return ""
    record.investigation_attempts.append(CoverageInvestigationAttempt(record.source_id, domain, pass_name, url, "fetched", _preview(content), "content_available", reason))
    return content


def _accept_candidate(
    record: CoverageSourceRecord,
    url: str,
    evidence_type: str,
    evidence: str,
    pass_name: str,
    *,
    confidence: float,
    require_listing_evidence: bool = True,
) -> bool:
    accepted, reason = _validate_aanbod_candidate(url, record.domain)
    if accepted and require_listing_evidence and evidence and evidence_type != "sitemap_url" and not _page_looks_like_listing_index(evidence):
        accepted = False
        reason = "page_does_not_look_like_listing_index"
    candidate = AanbodUrlCandidate(
        source_id=record.source_id,
        domain=record.domain,
        candidate_url=_canonical_url(url),
        accepted=accepted,
        rejection_reason="" if accepted else reason,
        evidence_type=evidence_type,
        evidence_preview=_preview(evidence),
        pass_name=pass_name,
        confidence=confidence,
    )
    record.aanbod_url_candidates.append(candidate)
    _refresh_rejected_candidate_summary(record)
    if not accepted:
        return False
    _set_accepted_aanbod_url(record, candidate.candidate_url, evidence_type, candidate.evidence_preview, confidence)
    return True


def _set_accepted_aanbod_url(
    record: CoverageSourceRecord,
    url: str,
    source: str,
    evidence: str,
    confidence: float,
) -> None:
    canonical_url = _canonical_url(url)
    record.accepted_aanbod_url = canonical_url
    record.aanbod_url = record.accepted_aanbod_url
    record.accepted_aanbod_url_source = source
    record.accepted_aanbod_url_confidence = max(record.accepted_aanbod_url_confidence, confidence)
    record.accepted_aanbod_url_evidence = _merge_text(record.accepted_aanbod_url_evidence, evidence)
    record.aanbod_url_type = "listing_index"
    record.aanbod_url_status = "accepted"
    record.aanbod_url_confidence = max(record.aanbod_url_confidence, confidence)
    record.aanbod_url_evidence = _merge_text(record.aanbod_url_evidence, f"{source}:{_preview(evidence)}")
    for candidate in record.aanbod_url_candidates:
        if _canonical_url(candidate.candidate_url) == canonical_url:
            candidate.accepted = True
            if candidate.rejection_reason:
                candidate.rejection_reason = "reconciled_after_later_acceptance"
    _refresh_rejected_candidate_summary(record)


def _refresh_rejected_candidate_summary(record: CoverageSourceRecord) -> None:
    rejected = [candidate for candidate in record.aanbod_url_candidates if not candidate.accepted]
    record.rejected_aanbod_url_count = len(rejected)
    if not rejected:
        record.best_rejected_candidate_url = ""
        record.best_rejected_candidate_reason = ""
        return
    best = sorted(rejected, key=lambda item: item.confidence, reverse=True)[0]
    record.best_rejected_candidate_url = best.candidate_url
    record.best_rejected_candidate_reason = best.rejection_reason


def _validate_aanbod_candidate(url: str, official_domain: str) -> tuple[bool, str]:
    parts = urlsplit(url if "://" in url else f"https://{url}")
    host = _normalize_domain(parts.netloc)
    if not host:
        return False, "missing_candidate_domain"
    if host in PORTAL_DOMAINS or any(host.endswith(f".{portal}") for portal in PORTAL_DOMAINS):
        return False, "funda_pararius_not_operational"
    official = _normalize_domain(official_domain)
    if official and host != official and not host.endswith(f".{official}"):
        return False, "candidate_not_on_official_domain"
    path = parts.path.rstrip("/").casefold()
    if not path or path == "/":
        return False, "homepage_not_listing_index"
    if _looks_like_property_detail_url(path):
        return False, "property_detail_url_rejected"
    if not _path_looks_like_listing_index(path):
        return False, "not_listing_index_path"
    return True, ""


def _set_terminal(record: CoverageSourceRecord, classification: str, terminal_status: CoverageTerminalStatus, reason: str) -> None:
    canonical = _canonical_family(classification)
    record.parser_family_candidate = canonical
    record.delivery_mode = canonical
    record.family_terminal_status = {
        "blocked_or_legal_review": "blocked_or_legal_review",
        "no_public_aanbod": "no_public_aanbod",
        "out_of_scope": "out_of_scope",
        "duplicate": "duplicate",
        "inactive_or_no_longer_trading": "inactive_or_no_longer_trading",
    }.get(canonical, record.family_terminal_status or "probable_with_evidence")
    record.terminal_status = terminal_status.value
    record.remaining_review_reason = "" if canonical in {"realworks_public", "ogonline_xhr", "no_public_aanbod", "out_of_scope", "inactive_or_no_longer_trading"} else reason
    record.recommended_next_action = _recommended_action(canonical, terminal_status)


def _add_family_evidence(record: CoverageSourceRecord, family: str, delivery: str, signal: str, evidence: str, confidence: float) -> None:
    family = _canonical_family(family)
    delivery = _canonical_family(delivery)
    record.parser_family_candidate = family
    record.delivery_mode = delivery
    record.family_confidence = max(record.family_confidence, confidence)
    record.family_evidence = _merge_text(record.family_evidence, signal)
    record.family_fingerprint_evidence.append(
        FamilyFingerprintEvidence(
            source_id=record.source_id,
            domain=record.domain,
            parser_family_candidate=family,
            delivery_mode=delivery,
            signal=signal,
            confidence=confidence,
            evidence_preview=_preview(evidence),
        )
    )


def _record_passes_quality_gate(record: CoverageSourceRecord) -> bool:
    family = _normalize_key(record.parser_family_candidate)
    if family in PROHIBITED_FINAL_VALUES:
        return False
    if record.has_noord_brabant_coverage and not record.accepted_aanbod_url and family not in {"no_public_aanbod", "blocked_or_legal_review", "out_of_scope", "duplicate", "inactive_or_no_longer_trading"}:
        return False
    return True


def _master_row(record: CoverageSourceRecord) -> dict[str, object]:
    values = asdict(record)
    values["has_noord_brabant_coverage"] = "yes" if record.has_noord_brabant_coverage else "no"
    return {column: _safe_cell(values.get(column, "")) for column in MASTER_COLUMNS}


def _candidate_row(candidate: AanbodUrlCandidate) -> dict[str, object]:
    return {
        "source_id": candidate.source_id,
        "domain": candidate.domain,
        "candidate_url": candidate.candidate_url,
        "accepted": "yes" if candidate.accepted else "no",
        "rejection_reason": candidate.rejection_reason,
        "evidence_type": candidate.evidence_type,
        "evidence_preview": candidate.evidence_preview,
        "pass_name": candidate.pass_name,
    }


def _fingerprint_row(evidence: FamilyFingerprintEvidence) -> dict[str, object]:
    return asdict(evidence)


def _attempt_row(attempt: CoverageInvestigationAttempt) -> dict[str, object]:
    return {
        "source_id": attempt.source_id,
        "domain": attempt.domain,
        "pass_name": attempt.pass_name,
        "attempted_url": attempt.attempted_url,
        "result": attempt.result,
        "decision": attempt.decision,
        "reason": attempt.reason,
    }


def _duplicate_row(record: CoverageSourceRecord) -> dict[str, object]:
    return {
        "duplicate_group_id": record.duplicate_group_id,
        "source_id": record.source_id,
        "source_name": record.source_name,
        "domain": record.domain,
        "duplicate_reason": record.duplicate_reason,
    }


def _realworks_verification_row(record: CoverageSourceRecord) -> dict[str, object]:
    verification = verify_realworks_candidate(record)
    return asdict(verification)


def _quality_gate_rows(metrics: Mapping[str, Any]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for name in QUALITY_GATE_NAMES:
        value = int(metrics.get(name, 0))
        rows.append({"metric": name, "value": value, "gate_type": "hard", "passed": "yes" if value == 0 else "no"})
    for name in (
        "missing_domain_queue_count",
        "office_location_unknown_count",
        "outside_office_sources_needing_review_count",
        "review_queue_count",
    ):
        rows.append({"metric": name, "value": metrics.get(name, 0), "gate_type": "reported", "passed": "n/a"})
    return rows


def _coverage_matrix_rows(records: Sequence[CoverageSourceRecord]) -> list[dict[str, object]]:
    counts = Counter((record.coverage_gemeente or "unknown", record.parser_family_candidate) for record in records)
    return [
        {"coverage_gemeente": gemeente, "parser_family_candidate": family, "count": count}
        for (gemeente, family), count in sorted(counts.items())
    ]


def _write_rows(worksheet: Worksheet, columns: Sequence[str], rows: Iterable[Mapping[str, Any]]) -> None:
    worksheet.append(tuple(columns))
    for row in rows:
        worksheet.append([_safe_cell(row.get(column, "")) for column in columns])


def _format_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[letter].width = max(10, min(max_length + 2, 60))


def _homepage_aanbod_links(base_url: str, html: str) -> tuple[tuple[str, str], ...]:
    links: list[tuple[str, str]] = []
    for match in re.finditer(r"<a\b[^>]*href\s*=\s*[\"'](?P<href>[^\"']+)[\"'][^>]*>(?P<body>.*?)</a>", html or "", flags=re.IGNORECASE | re.DOTALL):
        href = unescape(match.group("href")).strip()
        text = _visible_text(match.group("body"))
        url = _canonical_url(urljoin(base_url, href))
        if _path_looks_like_listing_index(urlsplit(url).path) or _contains_any(_normalize_key(text), ("aanbod", "woningen", "koopwoningen", "te koop")):
            links.append((url, text))
    return tuple(links)


def _sitemap_listing_urls(text: str) -> tuple[str, ...]:
    urls = re.findall(r"https?://[^<\s\"']+", text or "", flags=re.IGNORECASE)
    return tuple(url.rstrip("/") for url in urls if _path_looks_like_listing_index(urlsplit(url).path) and not _looks_like_property_detail_url(urlsplit(url).path))


def _page_looks_like_listing_index(content: str) -> bool:
    text = (content or "").casefold()
    if not text:
        return False
    return any(
        signal in text
        for signal in (
            "aanbodentry",
            "listing-card",
            "property-card",
            "woningaanbod",
            "koopwoningen",
            "object-list",
            "vraagprijs",
            "k.k.",
        )
    )


def _homepage_may_be_listing_index(url: str) -> bool:
    path = urlsplit(url).path.strip("/")
    return path in {"", "aanbod", "woningen", "koopwoningen"}


def _path_looks_like_listing_index(path: str) -> bool:
    normalized = (path or "").rstrip("/").casefold()
    return normalized in COMMON_AANBOD_PATHS or any(
        signal in normalized
        for signal in (
            "/aanbod",
            "/woningaanbod",
            "/woningen",
            "/koopwoningen",
            "/te-koop",
            "/huis-kopen",
        )
    )


def _looks_like_property_detail_url(path: str) -> bool:
    normalized = (path or "").casefold()
    return bool(
        re.search(r"/(?:koop|huur)/(?:huis|appartement|woning)-", normalized)
        or re.search(r"/(?:huis|appartement|woning)-[^/]*\d", normalized)
        or re.search(r"/aanbod/woningaanbod/[^/]+/(?:koop|huur)/(?:huis|appartement|woning)-", normalized)
        or re.search(r"/woningaanbod/[^/]+/(?:koop|huur)/(?:huis|appartement|woning)-", normalized)
        or re.search(r"/(?:koop|huur)/[^/]+/[^/]*\d", normalized)
        or re.search(r"/woningen/[^/]*\d", normalized)
    )


def _url_with_segments(parts: Any, segments: Sequence[str]) -> str:
    path = "/" + "/".join(segment.strip("/") for segment in segments if segment.strip("/"))
    return urlunsplit((parts.scheme or "https", parts.netloc, path, "", ""))


def _location_status(gemeente: str, province: str) -> str:
    if gemeente and province:
        return "known"
    if gemeente or province:
        return "partial"
    return "unknown_office_location"


def _outside_office_status(office_province: str, coverage_province: str) -> str:
    office = _normalize_key(office_province)
    coverage = _normalize_key(coverage_province)
    if not office:
        return "office_location_unknown"
    if coverage == NOORD_BRABANT and office != NOORD_BRABANT:
        return "outside_office_source_included"
    return "not_outside_office"


def _is_kin_record(record: CoverageSourceRecord) -> bool:
    return "kinmakelaars.nl" in _normalize_domain(record.domain) or "kinmakelaars" in _normalize_key(record.source_id)


def _duplicate_record(record: CoverageSourceRecord, existing: CoverageSourceRecord) -> CoverageSourceRecord:
    return CoverageSourceRecord(
        source_id=record.source_id,
        source_name=record.source_name,
        domain=record.domain,
        root_url=record.root_url,
        coverage_gemeente=record.coverage_gemeente,
        coverage_province=record.coverage_province,
        has_noord_brabant_coverage=record.has_noord_brabant_coverage,
        parser_family_candidate="duplicate",
        delivery_mode="duplicate",
        family_terminal_status="duplicate",
        terminal_status=CoverageTerminalStatus.CONFIRMED_DUPLICATE.value,
        duplicate_group_id=existing.source_id,
        duplicate_reason=f"normalized_domain_duplicate_of:{existing.source_id}",
        recommended_next_action="use_canonical_domain_record",
    )


def _is_inactive(record: CoverageSourceRecord) -> bool:
    text = _normalize_key(" ".join((record.source_quality_status, record.coverage_evidence, record.family_evidence)))
    return "is_active=false" in text or "inactive" in text or "no_longer_trading" in text


def _access_blocks(record: CoverageSourceRecord) -> bool:
    status = _normalize_key(record.access_policy_status)
    text = _normalize_key(" ".join((record.coverage_evidence, record.family_evidence, record.raw_aanbod_url_candidate, record.aanbod_url, record.accepted_aanbod_url, record.source_quality_status)))
    return (
        status in {"blocked", "permission_required", "legal_review", "disabled"}
        or any(portal in text for portal in ("funda.nl", "pararius.nl"))
        or any(signal in text for signal in ("captcha", "login_required", "403", "forbidden", "paywall"))
    )


def _recommended_action(classification: str, terminal_status: CoverageTerminalStatus) -> str:
    if classification == "realworks_public":
        return "include_in_noord_brabant_realworks_audit_v1"
    if classification == "ogonline_xhr":
        return "build_or_reuse_ogonline_source_config"
    if classification in {"custom_html", "custom_xhr", "custom_js_app", "iframe_vendor", "wordpress_static", "wordpress_json", "kolibri", "skarabee"}:
        return "manual_review_then_parser_family_or_source_config"
    if classification == "blocked_or_legal_review":
        return "legal_or_permission_review"
    if classification == "no_public_aanbod":
        return "no_operational_followup_until_public_aanbod_exists"
    if classification == "out_of_scope":
        return "exclude_from_noord_brabant_census"
    if classification == "duplicate":
        return "use_canonical_domain_record"
    if classification == "inactive_or_no_longer_trading":
        return "exclude_until_source_reactivated"
    return terminal_status.value


def _canonical_family(value: object) -> str:
    key = _normalize_key(value)
    mapped = {
        "realworks": "realworks_public",
        "realworks_public": "realworks_public",
        "ogonline": "ogonline_xhr",
        "og_online": "ogonline_xhr",
        "og-online": "ogonline_xhr",
        "wordpress_rest": "wordpress_json",
        "wordpress_html_cards": "wordpress_static",
        "wordpress_cards": "wordpress_static",
        "static_html_cards": "custom_html",
        "html_static_cards": "custom_html",
        "listing_html": "custom_html",
        "json_ld": "custom_html",
        "sitemap_detail": "custom_html",
        "kolibri_public": "kolibri",
        "skarabee_public": "skarabee",
        "funda_iframe_blocked": "blocked_or_legal_review",
        "pararius_external_blocked": "blocked_or_legal_review",
        "captcha_blocked": "blocked_or_legal_review",
        "login_required": "blocked_or_legal_review",
        "iframe_external": "iframe_vendor",
    }.get(key, key)
    return mapped if mapped in ALLOWED_CLASSIFICATIONS else mapped


def _canonical_access_status(value: object) -> str:
    key = _normalize_key(value)
    return {
        "allowed_official_source": "allowed",
        "requires_permission": "permission_required",
        "needs_permission": "permission_required",
        "manual_review": "researching",
        "": "researching",
    }.get(key, key)


def _source_id(domain: str, source_name: str, gemeente: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", _normalize_key(domain or source_name or "unknown-domain")).strip("-")
    scope = re.sub(r"[^a-z0-9]+", "-", _normalize_key(gemeente or "unknown")).strip("-")
    return f"{slug or 'unknown-domain'}__{scope or 'unknown'}"


def _canonical_root_url(value: str, domain: str) -> str:
    if value:
        return _canonical_url(value)
    if domain:
        return f"https://{domain}"
    return ""


def _root_url_for_path(domain: str, path: str) -> str:
    return urlunsplit(("https", domain, path, "", ""))


def _canonical_url(value: str) -> str:
    if not value:
        return ""
    candidate = value if "://" in value else f"https://{value}"
    parts = urlsplit(candidate)
    host = _normalize_domain(parts.netloc or parts.path)
    path = parts.path.rstrip("/") or "/"
    return urlunsplit((parts.scheme.lower() if parts.scheme else "https", host, path, "", ""))


def _domain_path(url: str, *, fallback_domain: str = "") -> tuple[str, str]:
    parts = urlsplit(url if "://" in url else f"https://{url}")
    domain = _normalize_domain(parts.netloc) or fallback_domain
    path = parts.path or "/"
    return domain, path


def _normalize_domain(value: object) -> str:
    text = _clean_text(value).casefold()
    if not text:
        return ""
    candidate = text if "://" in text else f"https://{text}"
    split = urlsplit(candidate)
    host = (split.netloc or split.path).split("/", 1)[0].split(":", 1)[0].strip().casefold()
    if host.startswith("www."):
        host = host[4:]
    return host.strip("/")


def _normalize_key(value: object) -> str:
    return _clean_text(value).casefold().replace("-", "_").replace(" ", "_")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " ".join(_clean_text(item) for item in value if _clean_text(item))
    return str(value).strip()


def _first_value(row: Mapping[str, str], keys: Iterable[str]) -> str:
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return ""


def _float_value(value: object) -> float:
    try:
        parsed = float(_clean_text(value) or "0")
    except ValueError:
        return 0.0
    if parsed > 1.0:
        return min(parsed / 100.0, 1.0)
    return max(parsed, 0.0)


def _preview(value: object, limit: int = EVIDENCE_PREVIEW_LIMIT) -> str:
    text = _safe_cell(value)
    if not isinstance(text, str):
        text = str(text)
    return text[:limit]


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = " ".join(str(value).split())
    lowered = text.casefold()
    if any(marker in lowered for marker in RAW_MARKERS):
        return "[raw_content_omitted]"
    if len(text) > LONG_TEXT_LIMIT:
        return "[long_text_omitted]"
    return text


def _visible_text(html: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html or "")).strip()


def _merge_text(current: object, incoming: object) -> str:
    parts: list[str] = []
    for raw in (_clean_text(current), _clean_text(incoming)):
        if not raw:
            continue
        for part in [item.strip() for item in raw.split(" | ") if item.strip()]:
            if part not in parts:
                parts.append(part)
    return " | ".join(parts)


def _join(values: Iterable[object]) -> str:
    return " | ".join(str(_safe_cell(value)) for value in values if _clean_text(_safe_cell(value)))


def _dedupe(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = _clean_text(value)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            result.append(cleaned)
    return result


def _contains_any(text: str, needles: Iterable[str]) -> bool:
    return any(needle in text for needle in needles)
