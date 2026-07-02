from __future__ import annotations

from collections import Counter, defaultdict
import csv
from dataclasses import asdict, dataclass, field
from pathlib import Path
import re
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .source_intelligence_models import normalize_domain, normalize_key, normalize_text, slugify


FUNDA_OPERATIONAL_SOURCE = "funda_manual_name_observation_only"
CRITICAL_CITIES = {"tilburg", "den bosch", "breda", "eindhoven", "'s-hertogenbosch", "s-hertogenbosch"}
GENERIC_MULTIPLE_NAME = "Meerdere makelaars"
OUTPUT_COLUMNS = (
    "makelaar_id",
    "display_name",
    "normalized_name",
    "aliases",
    "city_presence",
    "province_presence",
    "funda_seen_city",
    "funda_seen_pages",
    "funda_seen_count",
    "benchmark_count_city",
    "seen_in_existing_master",
    "existing_source_id",
    "existing_domain",
    "official_domain",
    "official_domain_status",
    "official_aanbod_url",
    "official_aanbod_url_status",
    "platform_family_candidate",
    "parser_family_candidate",
    "priority_score",
    "priority_tier",
    "gap_reason",
    "recommended_next_action",
    "manual_check_result",
    "manual_check_notes",
)
QUALITY_GATE_NAMES = (
    "makelaar_universe_rows_without_name_count",
    "duplicate_high_confidence_makelaar_count",
    "funda_property_detail_extracted_count",
    "funda_operational_property_source_count",
    "raw_html_json_persisted_count",
    "long_descriptions_exported_count",
    "browser_automation_used_count",
)
SLOGAN_PATTERNS = (
    re.compile(r"\bvia wie anders\b", re.IGNORECASE),
    re.compile(r"\bbeste makelaar\b", re.IGNORECASE),
    re.compile(r"\bsmart move\b", re.IGNORECASE),
    re.compile(r"\bzelf uw huis verkopen\b", re.IGNORECASE),
)
LEGAL_SUFFIX_RE = re.compile(r"\b(b\.?\s*v\.?|n\.?\s*v\.?)\b", re.IGNORECASE)
QUOTES_RE = re.compile(r"[\"'`“”‘’]+")
MULTISPACE_RE = re.compile(r"\s+")


@dataclass(slots=True)
class FundaMakelaarObservation:
    city: str
    province: str
    funda_search_url: str
    page_number: int
    benchmark_count: str
    makelaar_name_raw: str
    makelaar_name_clean_candidate: str
    name_truncated: bool
    generic_multiple_makelaars: bool
    name_needs_manual_review: bool
    evidence_text: str
    collection_status: str
    collected_at: str


@dataclass(slots=True)
class ExistingMakelaarMatch:
    source_id: str = ""
    source_name: str = ""
    domain: str = ""
    official_domain: str = ""
    official_domain_status: str = ""
    official_aanbod_url: str = ""
    official_aanbod_url_status: str = ""
    platform_family_candidate: str = ""
    parser_family_candidate: str = ""
    recommended_next_action: str = ""
    manual_check_result: str = ""
    manual_check_notes: str = ""


@dataclass(slots=True)
class MakelaarUniverseRow:
    makelaar_id: str
    display_name: str
    normalized_name: str
    aliases: str
    city_presence: str
    province_presence: str
    funda_seen_city: str
    funda_seen_pages: str
    funda_seen_count: int
    benchmark_count_city: str
    seen_in_existing_master: str
    existing_source_id: str
    existing_domain: str
    official_domain: str
    official_domain_status: str
    official_aanbod_url: str
    official_aanbod_url_status: str
    platform_family_candidate: str
    parser_family_candidate: str
    priority_score: int
    priority_tier: str
    gap_reason: str
    recommended_next_action: str
    manual_check_result: str
    manual_check_notes: str


@dataclass(slots=True)
class MakelaarUniverseResult:
    rows: tuple[MakelaarUniverseRow, ...]
    review_queue_rows: tuple[MakelaarUniverseRow, ...]
    quality_metrics: Mapping[str, int]
    per_city_counts: Mapping[str, int]
    missing_optional_inputs: tuple[str, ...] = ()
    effective_input_path: str = ""
    raw_rows_imported: int = 0
    deduped_makelaars: int = 0


@dataclass(slots=True)
class _ObservedGroup:
    display_name: str
    normalized_name: str
    aliases: set[str] = field(default_factory=set)
    cities: Counter[str] = field(default_factory=Counter)
    pages_by_city: defaultdict[str, set[int]] = field(default_factory=lambda: defaultdict(set))
    seen_count: int = 0
    benchmark_counts: dict[str, str] = field(default_factory=dict)
    truncated: bool = False
    generic_multiple: bool = False
    needs_manual_review: bool = False
    slogan_aliases: set[str] = field(default_factory=set)
    matched_sources: list[ExistingMakelaarMatch] = field(default_factory=list)


def load_funda_makelaar_observations(path: Path | str) -> list[FundaMakelaarObservation]:
    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    observations: list[FundaMakelaarObservation] = []
    for row in rows:
        observations.append(
            FundaMakelaarObservation(
                city=normalize_text(row.get("city")),
                province=normalize_text(row.get("province")) or "Noord-Brabant",
                funda_search_url=normalize_text(row.get("funda_search_url")),
                page_number=_parse_int(row.get("page_number")),
                benchmark_count=normalize_text(row.get("benchmark_count")),
                makelaar_name_raw=normalize_text(row.get("makelaar_name_raw")),
                makelaar_name_clean_candidate=normalize_text(row.get("makelaar_name_clean_candidate")),
                name_truncated=_parse_bool(row.get("name_truncated")),
                generic_multiple_makelaars=_parse_bool(row.get("generic_multiple_makelaars")),
                name_needs_manual_review=_parse_bool(row.get("name_needs_manual_review")),
                evidence_text=normalize_text(row.get("evidence_text")),
                collection_status=normalize_text(row.get("collection_status")),
                collected_at=normalize_text(row.get("collected_at")),
            )
        )
    return observations


def build_noord_brabant_makelaar_universe(
    funda_raw_csv_path: Path | str,
    source_completion_csv_path: Path | str | None = None,
    missing_domain_resolution_csv_path: Path | str | None = None,
    source_master_csv_path: Path | str | None = None,
    platform_fingerprint_csv_path: Path | str | None = None,
    source_seed_csv_path: Path | str | None = None,
) -> MakelaarUniverseResult:
    effective_input = _pick_effective_funda_input(Path(funda_raw_csv_path))
    observations = load_funda_makelaar_observations(effective_input)
    existing_index, missing_optional_inputs = _load_existing_index(
        source_completion_csv_path=source_completion_csv_path,
        missing_domain_resolution_csv_path=missing_domain_resolution_csv_path,
        source_master_csv_path=source_master_csv_path,
        platform_fingerprint_csv_path=platform_fingerprint_csv_path,
        source_seed_csv_path=source_seed_csv_path,
    )
    groups = _group_observations(observations, existing_index)
    rows = tuple(_build_universe_row(group) for group in sorted(groups.values(), key=lambda item: (-item.seen_count, item.normalized_name, item.display_name)))
    review_queue = tuple(row for row in rows if row.priority_tier == "manual_review")
    per_city_counts = {city: count for city, count in sorted(_per_city_counts(groups).items())}
    quality_metrics = compute_quality_metrics(rows)
    return MakelaarUniverseResult(
        rows=rows,
        review_queue_rows=review_queue,
        quality_metrics=quality_metrics,
        per_city_counts=per_city_counts,
        missing_optional_inputs=tuple(missing_optional_inputs),
        effective_input_path=str(effective_input),
        raw_rows_imported=len(observations),
        deduped_makelaars=len(rows),
    )


def compute_quality_metrics(rows: Iterable[MakelaarUniverseRow]) -> Mapping[str, int]:
    items = list(rows)
    duplicate_high_confidence = Counter(
        (row.normalized_name, row.official_domain or row.existing_domain)
        for row in items
        if row.priority_tier != "manual_review" and (row.official_domain or row.existing_domain)
    )
    metrics = {
        "makelaar_universe_rows_without_name_count": sum(1 for row in items if not row.normalized_name),
        "duplicate_high_confidence_makelaar_count": sum(1 for count in duplicate_high_confidence.values() if count > 1),
        "funda_property_detail_extracted_count": 0,
        "funda_operational_property_source_count": 0,
        "raw_html_json_persisted_count": 0,
        "long_descriptions_exported_count": 0,
        "browser_automation_used_count": 0,
    }
    return metrics


def write_makelaar_universe_csv(rows: Iterable[MakelaarUniverseRow], output_path: Path | str) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: getattr(row, column) for column in OUTPUT_COLUMNS})
    return path


def write_makelaar_universe_workbook(result: MakelaarUniverseResult, output_path: Path | str) -> Path:
    workbook = Workbook()
    universe_ws = workbook.active
    universe_ws.title = "Makelaar Universe"
    _write_rows(universe_ws, OUTPUT_COLUMNS, [asdict(row) for row in result.rows])
    _write_rows(workbook.create_sheet("Review Queue"), OUTPUT_COLUMNS, [asdict(row) for row in result.review_queue_rows])
    _write_rows(
        workbook.create_sheet("Per City Counts"),
        ("city", "makelaar_count"),
        [{"city": city, "makelaar_count": count} for city, count in result.per_city_counts.items()],
    )
    _write_rows(
        workbook.create_sheet("Quality Gates"),
        ("metric", "value", "gate_type", "passed"),
        [
            {"metric": name, "value": value, "gate_type": "hard", "passed": "yes" if value == 0 else "no"}
            for name, value in result.quality_metrics.items()
        ],
    )
    _write_rows(
        workbook.create_sheet("Run Summary"),
        ("metric", "value"),
        [
            {"metric": "effective_input_path", "value": result.effective_input_path},
            {"metric": "raw_rows_imported", "value": result.raw_rows_imported},
            {"metric": "deduped_makelaars", "value": result.deduped_makelaars},
            {"metric": "missing_optional_inputs", "value": " | ".join(result.missing_optional_inputs)},
        ],
    )
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def _pick_effective_funda_input(raw_path: Path) -> Path:
    merged_snapshot = raw_path.with_name(raw_path.stem + "_merged_snapshot.csv")
    if merged_snapshot.exists():
        return merged_snapshot
    return raw_path


def _load_existing_index(
    source_completion_csv_path: Path | str | None,
    missing_domain_resolution_csv_path: Path | str | None,
    source_master_csv_path: Path | str | None,
    platform_fingerprint_csv_path: Path | str | None,
    source_seed_csv_path: Path | str | None,
) -> tuple[dict[str, object], list[str]]:
    entries_by_name: defaultdict[str, list[ExistingMakelaarMatch]] = defaultdict(list)
    entries_by_domain: dict[str, ExistingMakelaarMatch] = {}
    missing_optional_inputs: list[str] = []

    completion_rows = _read_optional_csv(source_completion_csv_path, missing_optional_inputs)
    master_rows = _read_optional_csv(source_master_csv_path, missing_optional_inputs)
    fingerprint_rows = _read_optional_csv(platform_fingerprint_csv_path, missing_optional_inputs)
    seed_rows = _read_optional_csv(source_seed_csv_path, missing_optional_inputs)
    domain_resolution_rows = _read_optional_csv(missing_domain_resolution_csv_path, missing_optional_inputs)

    fingerprint_by_source_id = {normalize_text(row.get("source_id")): row for row in fingerprint_rows}
    master_by_source_id = {normalize_text(row.get("source_id")): row for row in master_rows}
    seed_by_domain = {_normalize_domain_field(row.get("root_domain") or row.get("domain") or row.get("website")): row for row in seed_rows}
    resolution_by_name = defaultdict(list)
    for row in domain_resolution_rows:
        resolution_by_name[_normalize_makelaar_name(row.get("raw_source_name"))].append(row)

    for row in completion_rows:
        source_id = normalize_text(row.get("source_id"))
        domain = _normalize_domain_field(row.get("domain") or row.get("root_url"))
        source_name = normalize_text(row.get("source_name"))
        master_row = master_by_source_id.get(source_id, {})
        fingerprint_row = fingerprint_by_source_id.get(source_id, {})
        seed_row = seed_by_domain.get(domain, {})
        match = ExistingMakelaarMatch(
            source_id=source_id,
            source_name=source_name,
            domain=domain,
            official_domain=domain,
            official_domain_status="known" if domain else "unknown_explicit",
            official_aanbod_url=normalize_text(row.get("accepted_aanbod_url")) or normalize_text(master_row.get("aanbod_url")) or normalize_text(seed_row.get("koopaanbod_url")),
            official_aanbod_url_status=normalize_text(row.get("aanbod_url_status")) or normalize_text(master_row.get("aanbod_url_quality")) or normalize_text(seed_row.get("koopaanbod_url_quality")) or "unknown_explicit",
            platform_family_candidate=normalize_text(row.get("delivery_mode")) or normalize_text(row.get("platform_guess")) or normalize_text(fingerprint_row.get("detected_platform")),
            parser_family_candidate=normalize_text(row.get("parser_family_candidate")) or normalize_text(row.get("delivery_mode")) or _platform_to_parser_family(normalize_text(fingerprint_row.get("detected_platform"))),
            recommended_next_action=normalize_text(row.get("recommended_next_action")) or normalize_text(fingerprint_row.get("recommended_next_action")),
            manual_check_result=normalize_text(row.get("manual_check_result")),
            manual_check_notes=normalize_text(row.get("manual_check_notes")),
        )
        entries_by_name[_normalize_makelaar_name(source_name)].append(match)
        if domain:
            entries_by_domain[domain] = match

    for row in master_rows:
        source_name = normalize_text(row.get("office_name"))
        domain = _normalize_domain_field(row.get("root_domain") or row.get("website"))
        source_id = normalize_text(row.get("source_id"))
        normalized_name = _normalize_makelaar_name(source_name)
        if not normalized_name:
            continue
        if any(match.source_id == source_id for match in entries_by_name[normalized_name]):
            continue
        fingerprint_row = fingerprint_by_source_id.get(source_id, {})
        match = ExistingMakelaarMatch(
            source_id=source_id,
            source_name=source_name,
            domain=domain,
            official_domain=domain,
            official_domain_status="known" if domain else "unknown_explicit",
            official_aanbod_url=normalize_text(row.get("aanbod_url")),
            official_aanbod_url_status=normalize_text(row.get("aanbod_url_quality")) or "unknown_explicit",
            platform_family_candidate=normalize_text(fingerprint_row.get("detected_platform")),
            parser_family_candidate=_platform_to_parser_family(normalize_text(fingerprint_row.get("detected_platform"))),
            recommended_next_action=normalize_text(fingerprint_row.get("recommended_next_action")),
        )
        entries_by_name[normalized_name].append(match)
        if domain and domain not in entries_by_domain:
            entries_by_domain[domain] = match

    for row in seed_rows:
        source_name = normalize_text(row.get("office_name"))
        normalized_name = _normalize_makelaar_name(source_name)
        if not normalized_name:
            continue
        domain = _normalize_domain_field(row.get("root_domain") or row.get("domain") or row.get("website"))
        if any(match.domain == domain and domain for match in entries_by_name[normalized_name]):
            continue
        entries_by_name[normalized_name].append(
            ExistingMakelaarMatch(
                source_name=source_name,
                domain=domain,
                official_domain=domain,
                official_domain_status="known" if domain else "unknown_explicit",
                official_aanbod_url=normalize_text(row.get("koopaanbod_url")),
                official_aanbod_url_status=normalize_text(row.get("koopaanbod_url_quality")) or "unknown_explicit",
                manual_check_notes=normalize_text(row.get("notes")),
            )
        )

    for normalized_name, resolution_rows in resolution_by_name.items():
        resolved = [row for row in resolution_rows if normalize_text(row.get("resolved_domain"))]
        if not resolved:
            continue
        if normalized_name in entries_by_name:
            continue
        best = resolved[0]
        entries_by_name[normalized_name].append(
            ExistingMakelaarMatch(
                source_id=normalize_text(best.get("resolved_source_id")),
                source_name=normalize_text(best.get("raw_source_name")),
                domain=_normalize_domain_field(best.get("resolved_domain")),
                official_domain=_normalize_domain_field(best.get("resolved_domain")),
                official_domain_status=normalize_text(best.get("resolution_status")) or "known",
                recommended_next_action=normalize_text(best.get("suggested_next_action")),
                manual_check_result=normalize_text(best.get("manual_check_result")),
                manual_check_notes=normalize_text(best.get("manual_check_notes")),
            )
        )

    return {"by_name": entries_by_name, "by_domain": entries_by_domain}, missing_optional_inputs


def _read_optional_csv(path: Path | str | None, missing_optional_inputs: list[str]) -> list[dict[str, str]]:
    if path is None:
        return []
    csv_path = Path(path)
    if not csv_path.exists():
        missing_optional_inputs.append(str(csv_path))
        return []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _group_observations(
    observations: Iterable[FundaMakelaarObservation],
    existing_index: Mapping[str, object],
) -> dict[str, _ObservedGroup]:
    groups: dict[str, _ObservedGroup] = {}
    by_name: Mapping[str, list[ExistingMakelaarMatch]] = existing_index["by_name"]  # type: ignore[index]
    for obs in observations:
        normalized = _normalize_makelaar_name(obs.makelaar_name_clean_candidate or obs.makelaar_name_raw)
        if not normalized and not obs.generic_multiple_makelaars:
            continue
        if obs.generic_multiple_makelaars:
            normalized = _normalize_makelaar_name(GENERIC_MULTIPLE_NAME)
        display_name, slogan_alias = _canonical_display_name(obs.makelaar_name_raw)
        if obs.generic_multiple_makelaars:
            display_name = GENERIC_MULTIPLE_NAME
        key = normalized or _normalize_makelaar_name(display_name)
        group = groups.get(key)
        if group is None:
            group = _ObservedGroup(display_name=display_name, normalized_name=key)
            groups[key] = group
        group.aliases.add(normalize_text(obs.makelaar_name_raw))
        if slogan_alias:
            group.slogan_aliases.add(slogan_alias)
            group.aliases.add(slogan_alias)
        group.cities[obs.city] += 1
        group.pages_by_city[obs.city].add(obs.page_number)
        group.seen_count += 1
        if obs.benchmark_count and obs.city not in group.benchmark_counts:
            group.benchmark_counts[obs.city] = obs.benchmark_count
        group.truncated = group.truncated or obs.name_truncated or _looks_truncated(obs.makelaar_name_raw)
        group.generic_multiple = group.generic_multiple or obs.generic_multiple_makelaars
        group.needs_manual_review = group.needs_manual_review or obs.name_needs_manual_review
        if not group.display_name:
            group.display_name = display_name

    for group in groups.values():
        if group.normalized_name in by_name:
            group.matched_sources.extend(by_name[group.normalized_name])
        if group.truncated:
            group.needs_manual_review = True
    return groups


def _build_universe_row(group: _ObservedGroup) -> MakelaarUniverseRow:
    match = _pick_best_match(group)
    aliases = sorted(alias for alias in group.aliases if alias and alias != group.display_name)
    city_presence = " | ".join(f"{city}:{group.cities[city]}" for city in sorted(group.cities))
    benchmark_count_city = " | ".join(
        f"{city}:{group.benchmark_counts[city]}"
        for city in sorted(group.benchmark_counts)
    )
    funda_seen_city = sorted(group.cities, key=lambda city: (-group.cities[city], city))[0] if group.cities else ""
    funda_seen_pages = " | ".join(
        f"{city}:{','.join(str(page) for page in sorted(group.pages_by_city[city]))}"
        for city in sorted(group.pages_by_city)
    )
    seen_in_existing_master = "yes" if match.source_id or match.domain or match.official_domain else "no"
    official_domain = match.official_domain or match.domain
    official_domain_status = match.official_domain_status or ("known" if official_domain else "unknown_explicit")
    official_aanbod_url = match.official_aanbod_url
    official_aanbod_url_status = match.official_aanbod_url_status or ("known" if official_aanbod_url else "unknown_explicit")
    parser_family_candidate = match.parser_family_candidate
    platform_family_candidate = match.platform_family_candidate
    priority_score = _compute_priority_score(group, match)
    priority_tier = _compute_priority_tier(group, match, priority_score)
    gap_reason = _gap_reason(group, match)
    recommended_next_action = _recommended_next_action(group, match, priority_tier)
    manual_check_result = match.manual_check_result
    manual_check_notes = " | ".join(
        part for part in (
            match.manual_check_notes,
            "truncated_name_from_funda" if group.truncated else "",
            "generic_multiple_bucket" if group.generic_multiple else "",
            "slogan_alias_present" if group.slogan_aliases else "",
            "ambiguous_existing_source_match" if len(group.matched_sources) > 1 else "",
        )
        if part
    )
    makelaar_id = _make_makelaar_id(group.normalized_name, official_domain or match.source_id or funda_seen_city)
    return MakelaarUniverseRow(
        makelaar_id=makelaar_id,
        display_name=group.display_name,
        normalized_name=group.normalized_name,
        aliases=" | ".join(aliases),
        city_presence=city_presence,
        province_presence="Noord-Brabant",
        funda_seen_city=funda_seen_city,
        funda_seen_pages=funda_seen_pages,
        funda_seen_count=group.seen_count,
        benchmark_count_city=benchmark_count_city,
        seen_in_existing_master=seen_in_existing_master,
        existing_source_id=match.source_id,
        existing_domain=match.domain,
        official_domain=official_domain,
        official_domain_status=official_domain_status,
        official_aanbod_url=official_aanbod_url,
        official_aanbod_url_status=official_aanbod_url_status,
        platform_family_candidate=platform_family_candidate,
        parser_family_candidate=parser_family_candidate,
        priority_score=priority_score,
        priority_tier=priority_tier,
        gap_reason=gap_reason,
        recommended_next_action=recommended_next_action,
        manual_check_result=manual_check_result,
        manual_check_notes=manual_check_notes,
    )


def _per_city_counts(groups: Mapping[str, _ObservedGroup]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for group in groups.values():
        for city in group.cities:
            counts[city] += 1
    return counts


def _pick_best_match(group: _ObservedGroup) -> ExistingMakelaarMatch:
    if not group.matched_sources:
        return ExistingMakelaarMatch()
    if len(group.matched_sources) == 1:
        return group.matched_sources[0]
    scored = sorted(
        group.matched_sources,
        key=lambda item: (
            -(1 if item.official_aanbod_url else 0),
            -(1 if item.official_domain else 0),
            -(1 if item.parser_family_candidate else 0),
            item.source_id,
            item.domain,
        ),
    )
    top = scored[0]
    if len(scored) > 1:
        second = scored[1]
        if (
            bool(top.official_aanbod_url) == bool(second.official_aanbod_url)
            and bool(top.official_domain) == bool(second.official_domain)
            and top.domain != second.domain
        ):
            ambiguous = ExistingMakelaarMatch(**asdict(top))
            ambiguous.manual_check_result = "manual_review"
            ambiguous.manual_check_notes = "multiple_existing_domain_candidates"
            return ambiguous
    return top


def _compute_priority_score(group: _ObservedGroup, match: ExistingMakelaarMatch) -> int:
    score = 0
    if any(normalize_key(city) in CRITICAL_CITIES for city in group.cities):
        score += 40
    score += min(group.seen_count, 120)
    if match.official_domain:
        score += 20
    if match.official_aanbod_url:
        score += 30
    if match.parser_family_candidate:
        score += 15
    if match.source_id:
        score += 8
    if group.generic_multiple:
        score -= 80
    if group.truncated:
        score -= 35
    if group.needs_manual_review:
        score -= 25
    if "manual_review" in normalize_key(match.manual_check_result):
        score -= 20
    return score


def _compute_priority_tier(group: _ObservedGroup, match: ExistingMakelaarMatch, score: int) -> str:
    if group.generic_multiple or group.truncated or group.needs_manual_review or "manual_review" in normalize_key(match.manual_check_result):
        return "manual_review"
    if score >= 100 and any(normalize_key(city) in CRITICAL_CITIES for city in group.cities):
        return "P0_city_critical"
    if score >= 70:
        return "P1_high"
    if score >= 35:
        return "P2_medium"
    return "P3_low"


def _gap_reason(group: _ObservedGroup, match: ExistingMakelaarMatch) -> str:
    reasons: list[str] = []
    if group.generic_multiple:
        reasons.append("generic_multiple_makelaars")
    if group.truncated:
        reasons.append("truncated_funda_name")
    if not match.official_domain:
        reasons.append("official_domain_unknown")
    if not match.official_aanbod_url:
        reasons.append("official_aanbod_url_unknown")
    if len(group.matched_sources) > 1:
        reasons.append("ambiguous_existing_match")
    return " | ".join(reasons) or "none"


def _recommended_next_action(group: _ObservedGroup, match: ExistingMakelaarMatch, tier: str) -> str:
    if group.generic_multiple:
        return "exclude_generic_bucket_from_domain_resolution"
    if group.truncated:
        return "manual_name_expansion_review"
    if tier == "manual_review":
        return match.recommended_next_action or "manual_review_needed"
    if not match.official_domain:
        return "manual_official_domain_research_required"
    if not match.official_aanbod_url:
        return "verify_public_aanbod_url"
    if not match.parser_family_candidate:
        return "family_fingerprint_needed"
    return match.recommended_next_action or "ready_for_makelaar_universe_followup"


def _normalize_makelaar_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = LEGAL_SUFFIX_RE.sub(" ", text)
    text = QUOTES_RE.sub("", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("+", " ")
    if " - " in text:
        text = text.split(" - ", 1)[0]
    text = re.sub(r"[!?.,;:]+", " ", text)
    text = MULTISPACE_RE.sub(" ", text).strip().lower()
    return text


def _canonical_display_name(raw_name: str) -> tuple[str, str]:
    raw = normalize_text(raw_name)
    if not raw:
        return "", ""
    if " - " in raw:
        left, right = raw.split(" - ", 1)
        if any(pattern.search(right) for pattern in SLOGAN_PATTERNS) or _looks_marketing_slogan(right):
            return left.strip(), right.strip().strip('"')
    return raw, ""


def _looks_marketing_slogan(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    lower = text.lower()
    return any(token in lower for token in ("smart move", "via wie anders", "jouw regio", "zelf uw huis verkopen"))


def _looks_truncated(value: str) -> bool:
    text = normalize_text(value)
    return text.endswith("...") or text.endswith("…")


def _make_makelaar_id(normalized_name: str, stable_hint: str) -> str:
    base = slugify(normalized_name or stable_hint or "makelaar")
    hint = slugify(stable_hint or "nb")
    return f"{base}__{hint}"


def _platform_to_parser_family(detected_platform: str) -> str:
    platform = normalize_key(detected_platform)
    if "realworks" in platform:
        return "realworks_public"
    if "ogonline" in platform:
        return "ogonline_xhr"
    if "wordpress" in platform:
        return "wordpress_json"
    if platform in {"", "unknown"}:
        return ""
    return platform


def _normalize_domain_field(value: object) -> str:
    return normalize_domain(value)


def _parse_bool(value: object) -> bool:
    return normalize_key(value) in {"1", "true", "yes"}


def _parse_int(value: object) -> int:
    text = normalize_text(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _write_rows(worksheet: Worksheet, columns: Iterable[str], rows: Iterable[Mapping[str, object]]) -> None:
    ordered_columns = tuple(columns)
    worksheet.append(ordered_columns)
    for row in rows:
        worksheet.append([_safe_cell(row.get(column, "")) for column in ordered_columns])


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return normalize_text(value)


def _format_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in worksheet.columns:
        sample = list(column_cells[:100])
        letter = get_column_letter(sample[0].column)
        max_length = max(len(str(cell.value or "")) for cell in sample)
        worksheet.column_dimensions[letter].width = max(10, min(max_length + 2, 60))
